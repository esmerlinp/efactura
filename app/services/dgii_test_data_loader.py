import xml.etree.ElementTree as ET
from datetime import datetime, timezone
import openpyxl
import re


class DgiiTestDataLoader:
    """
    Carga el Set de Pruebas de DGII (Excel) y construye el XML e-CF
    directamente desde los datos del Excel, usando los nombres de header
    (no columnas fijas) para mapear a la estructura XML de DGII.
    """

    ITEM_LINE_COUNT = 62
    PAYMENT_MAX = 7
    TAX_MAX = 4
    TELEFONO_MAX = 3
    CODIGO_MAX = 5
    SUB_RECARGO_MAX = 5
    SUB_DESCUENTO_MAX = 5
    SUBCANTIDAD_MAX = 5
    TIPO_IMPUESTO_PER_ITEM = 2

    @classmethod
    def _col_letter(cls, n):
        result = ""
        while n > 0:
            n -= 1
            result = chr(n % 26 + ord("A")) + result
            n //= 26
        return result

    @classmethod
    def load_workbook(cls, filepath):
        wb = openpyxl.load_workbook(filepath, data_only=True)

        def sheet_to_rows(ws):
            headers = {}
            for cell in ws[1]:
                if cell.value is not None:
                    headers[cls._col_letter(cell.column)] = str(cell.value).strip()
            rows_data = []
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
                row_dict = {}
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() != "#e":
                        row_dict[cls._col_letter(cell.column)] = str(cell.value).strip()
                if row_dict:
                    rows_data.append((row_dict, headers))
            return rows_data

        s1 = wb[wb.sheetnames[0]]
        s2 = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else None
        sheet1_data = sheet_to_rows(s1)
        sheet2_data = sheet_to_rows(s2) if s2 else []
        wb.close()
        return sheet1_data, sheet2_data

    # ── helper: obtener valor por nombre de header ──
    @classmethod
    def _v(cls, row_dict, headers, name, default=None):
        for col, hdr in headers.items():
            if hdr == name and col in row_dict:
                return row_dict[col]
        return default

    # ── helper: construir sub-elemento si el valor existe ──
    @classmethod
    def _add(cls, parent, tag, value):
        if value is not None and str(value).strip() != "":
            ET.SubElement(parent, tag).text = str(value).strip()

    # ═══════════════════════════════════════════════════════════════
    # Construccion del XML principal
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def build_xml_from_row(cls, row_dict, headers):
        root = ET.Element("ECF")
        enc = ET.SubElement(root, "Encabezado")
        ET.SubElement(enc, "Version").text = "1.0"

        # ── IdDoc ──
        id_doc = ET.SubElement(enc, "IdDoc")
        tipo = cls._v(row_dict, headers, "TipoeCF") or "32"
        tipo = tipo.replace("E", "") if tipo.startswith("E") else tipo
        ET.SubElement(id_doc, "TipoeCF").text = tipo
        cls._add(id_doc, "eNCF", cls._v(row_dict, headers, "ENCF"))

        for tag in [
            "FechaVencimientoSecuencia", "IndicadorNotaCredito",
            "IndicadorEnvioDiferido", "IndicadorMontoGravado",
            "TipoIngresos", "TipoPago", "FechaLimitePago", "TerminoPago",
            "TipoCuentaPago", "NumeroCuentaPago", "BancoPago",
            "FechaDesde", "FechaHasta",
        ]:
            cls._add(id_doc, tag, cls._v(row_dict, headers, tag))

        # TablaFormasPago
        formas = []
        for i in range(1, cls.PAYMENT_MAX + 1):
            fp = cls._v(row_dict, headers, f"FormaPago[{i}]")
            mp = cls._v(row_dict, headers, f"MontoPago[{i}]")
            if fp:
                formas.append((fp, mp or "0.00"))
        if formas:
            tfp = ET.SubElement(id_doc, "TablaFormasPago")
            for fp, mp in formas:
                fdp = ET.SubElement(tfp, "FormaDePago")
                ET.SubElement(fdp, "FormaPago").text = fp
                ET.SubElement(fdp, "MontoPago").text = mp

        # ── Emisor ──
        emisor_elem = ET.SubElement(enc, "Emisor")
        # Orden estricto segun XSD: RNCEmisor, RazonSocialEmisor, NombreComercial,
        # Sucursal, DireccionEmisor, Municipio, Provincia,
        # [TablaTelefonoEmisor], CorreoEmisor, WebSite, ActividadEconomica,
        # CodigoVendedor, NumeroFacturaInterna, NumeroPedidoInterno,
        # ZonaVenta, RutaVenta, InformacionAdicionalEmisor, FechaEmision
        for tag in [
            "RNCEmisor", "RazonSocialEmisor", "NombreComercial",
            "Sucursal", "DireccionEmisor", "Municipio", "Provincia",
        ]:
            cls._add(emisor_elem, tag, cls._v(row_dict, headers, tag))

        # TablaTelefonoEmisor — entre Provincia y CorreoEmisor
        tels = []
        for i in range(1, cls.TELEFONO_MAX + 1):
            t = cls._v(row_dict, headers, f"TelefonoEmisor[{i}]")
            if t:
                tels.append(t)
        if tels:
            ttel = ET.SubElement(emisor_elem, "TablaTelefonoEmisor")
            for t in tels:
                ET.SubElement(ttel, "TelefonoEmisor").text = t

        for tag in [
            "CorreoEmisor", "WebSite", "ActividadEconomica",
            "CodigoVendedor", "NumeroFacturaInterna", "NumeroPedidoInterno",
            "ZonaVenta", "RutaVenta", "InformacionAdicionalEmisor", "FechaEmision",
        ]:
            cls._add(emisor_elem, tag, cls._v(row_dict, headers, tag))

        # ── Comprador ──
        # Copia verbatim del Excel: el conjunto de datos de la DGII espera
        # exactamente los valores de la fila (incluye RNCComprador=131880681
        # en E32<250K y RFCE; los casos E43/E47 lo traen vacío).
        comp_children = []
        for tag in [
            "RNCComprador", "IdentificadorExtranjero", "RazonSocialComprador",
            "ContactoComprador", "CorreoComprador", "DireccionComprador",
            "MunicipioComprador", "ProvinciaComprador", "PaisComprador",
            "FechaEntrega", "ContactoEntrega", "DireccionEntrega",
            "TelefonoAdicional", "FechaOrdenCompra", "NumeroOrdenCompra",
            "CodigoInternoComprador", "ResponsablePago",
            "InformacionAdicionalComprador",
        ]:
            v = cls._v(row_dict, headers, tag)
            if v:
                comp_children.append((tag, v))

        if comp_children:
            comp_elem = ET.SubElement(enc, "Comprador")
            for tag, val in comp_children:
                ET.SubElement(comp_elem, tag).text = val

        # ── InformacionesAdicionales ──
        info_children = []
        for tag in [
            "FechaEmbarque", "NumeroEmbarque", "NumeroContenedor",
            "NumeroReferencia", "NombrePuertoEmbarque", "CondicionesEntrega",
            "TotalFob", "Seguro", "Flete", "OtrosGastos", "TotalCif",
            "RegimenAduanero", "NombrePuertoSalida", "NombrePuertoDesembarque",
            "PesoBruto", "PesoNeto", "UnidadPesoBruto", "UnidadPesoNeto",
            "CantidadBulto", "UnidadBulto", "VolumenBulto", "UnidadVolumen",
        ]:
            v = cls._v(row_dict, headers, tag)
            if v:
                info_children.append((tag, v))
        if info_children:
            info_elem = ET.SubElement(enc, "InformacionesAdicionales")
            for tag, val in info_children:
                ET.SubElement(info_elem, tag).text = val

        # ── Transporte ──
        trans_children = []
        for tag in [
            "Conductor", "DocumentoTransporte", "Ficha", "Placa",
            "RutaTransporte", "ZonaTransporte", "NumeroAlbaran",
            "ViaTransporte", "PaisOrigen", "DireccionDestino", "PaisDestino",
            "RNCIdentificacionCompaniaTransportista",
            "NombreCompaniaTransportista", "NumeroViaje",
        ]:
            v = cls._v(row_dict, headers, tag)
            if v:
                trans_children.append((tag, v))
        if trans_children:
            trans_elem = ET.SubElement(enc, "Transporte")
            for tag, val in trans_children:
                ET.SubElement(trans_elem, tag).text = val

        # First part of Totales
        totales_elem = ET.SubElement(enc, "Totales")
        for tag in [
            "MontoGravadoTotal", "MontoGravadoI1", "MontoGravadoI2",
            "MontoGravadoI3", "MontoExento",
            "ITBIS1", "ITBIS2", "ITBIS3",
            "TotalITBIS", "TotalITBIS1", "TotalITBIS2", "TotalITBIS3",
            "MontoImpuestoAdicional",
        ]:
            cls._add(totales_elem, tag, cls._v(row_dict, headers, tag))

        # ImpuestosAdicionales wrapper (segun XSD)
        imp_children = []
        for i in range(1, cls.TAX_MAX + 1):
            tipo_imp = cls._v(row_dict, headers, f"TipoImpuesto[{i}]")
            if not tipo_imp:
                continue
            imp_children.append({
                "TipoImpuesto": tipo_imp,
                "TasaImpuestoAdicional": cls._v(row_dict, headers, f"TasaImpuestoAdicional[{i}]") or "",
                "MontoImpuestoSelectivoConsumoEspecifico": cls._v(row_dict, headers, f"MontoImpuestoSelectivoConsumoEspecifico[{i}]") or "",
                "MontoImpuestoSelectivoConsumoAdvalorem": cls._v(row_dict, headers, f"MontoImpuestoSelectivoConsumoAdvalorem[{i}]") or "",
                "OtrosImpuestosAdicionales": cls._v(row_dict, headers, f"OtrosImpuestosAdicionales[{i}]") or "",
            })
        if imp_children:
            imp_wrapper = ET.SubElement(totales_elem, "ImpuestosAdicionales")
            for imp in imp_children:
                ia = ET.SubElement(imp_wrapper, "ImpuestoAdicional")
                ET.SubElement(ia, "TipoImpuesto").text = imp["TipoImpuesto"]
                if imp["TasaImpuestoAdicional"]:
                    ET.SubElement(ia, "TasaImpuestoAdicional").text = imp["TasaImpuestoAdicional"]
                if imp["MontoImpuestoSelectivoConsumoEspecifico"]:
                    ET.SubElement(ia, "MontoImpuestoSelectivoConsumoEspecifico").text = imp["MontoImpuestoSelectivoConsumoEspecifico"]
                if imp["MontoImpuestoSelectivoConsumoAdvalorem"]:
                    ET.SubElement(ia, "MontoImpuestoSelectivoConsumoAdvalorem").text = imp["MontoImpuestoSelectivoConsumoAdvalorem"]
                if imp["OtrosImpuestosAdicionales"]:
                    ET.SubElement(ia, "OtrosImpuestosAdicionales").text = imp["OtrosImpuestosAdicionales"]

        # Second part of Totales
        for tag in [
            "MontoTotal", "MontoNoFacturable", "MontoPeriodo",
            "SaldoAnterior", "MontoAvancePago", "ValorPagar",
            "TotalITBISRetenido", "TotalISRRetencion",
            "TotalITBISPercepcion", "TotalISRPercepcion",
        ]:
            cls._add(totales_elem, tag, cls._v(row_dict, headers, tag))



        # ── OtraMoneda ──
        mon = cls._v(row_dict, headers, "TipoMoneda")
        if mon and mon != "DOP":
            om_elem = ET.SubElement(enc, "OtraMoneda")
            ET.SubElement(om_elem, "TipoMoneda").text = mon
            tc = cls._v(row_dict, headers, "TipoCambio")
            if tc:
                ET.SubElement(om_elem, "TipoCambio").text = tc

            for tag in [
                "MontoGravadoTotalOtraMoneda", "MontoGravado1OtraMoneda",
                "MontoGravado2OtraMoneda", "MontoGravado3OtraMoneda",
                "MontoExentoOtraMoneda",
                "TotalITBISOtraMoneda", "TotalITBIS1OtraMoneda",
                "TotalITBIS2OtraMoneda", "TotalITBIS3OtraMoneda",
                "MontoImpuestoAdicionalOtraMoneda",
                "MontoTotalOtraMoneda",
            ]:
                cls._add(om_elem, tag, cls._v(row_dict, headers, tag))

            # ImpuestosAdicionalesOtraMoneda
            imp_om_children = []
            for i in range(1, cls.TAX_MAX + 1):
                tipo_om = cls._v(row_dict, headers, f"TipoImpuestoOtraMoneda[{i}]")
                if not tipo_om:
                    continue
                imp_om_children.append({
                    "TipoImpuestoOtraMoneda": tipo_om,
                    "TasaImpuestoAdicionalOtraMoneda": cls._v(row_dict, headers, f"TasaImpuestoAdicionalOtraMoneda[{i}]") or "",
                    "MontoImpuestoSelectivoConsumoEspecificoOtraMoneda": cls._v(row_dict, headers, f"MontoImpuestoSelectivoConsumoEspecificoOtraMoneda[{i}]") or "",
                    "MontoImpuestoSelectivoConsumoAdvaloremOtraMoneda": cls._v(row_dict, headers, f"MontoImpuestoSelectivoConsumoAdvaloremOtraMoneda[{i}]") or "",
                    "OtrosImpuestosAdicionalesOtraMoneda": cls._v(row_dict, headers, f"OtrosImpuestosAdicionalesOtraMoneda[{i}]") or "",
                })
            if imp_om_children:
                imp_om_wrapper = ET.SubElement(om_elem, "ImpuestosAdicionalesOtraMoneda")
                for imp in imp_om_children:
                    ia_om = ET.SubElement(imp_om_wrapper, "ImpuestoAdicionalOtraMoneda")
                    ET.SubElement(ia_om, "TipoImpuestoOtraMoneda").text = imp["TipoImpuestoOtraMoneda"]
                    if imp["TasaImpuestoAdicionalOtraMoneda"]:
                        ET.SubElement(ia_om, "TasaImpuestoAdicionalOtraMoneda").text = imp["TasaImpuestoAdicionalOtraMoneda"]
                    if imp["MontoImpuestoSelectivoConsumoEspecificoOtraMoneda"]:
                        ET.SubElement(ia_om, "MontoImpuestoSelectivoConsumoEspecificoOtraMoneda").text = imp["MontoImpuestoSelectivoConsumoEspecificoOtraMoneda"]
                    if imp["MontoImpuestoSelectivoConsumoAdvaloremOtraMoneda"]:
                        ET.SubElement(ia_om, "MontoImpuestoSelectivoConsumoAdvaloremOtraMoneda").text = imp["MontoImpuestoSelectivoConsumoAdvaloremOtraMoneda"]
                    if imp["OtrosImpuestosAdicionalesOtraMoneda"]:
                        ET.SubElement(ia_om, "OtrosImpuestosAdicionalesOtraMoneda").text = imp["OtrosImpuestosAdicionalesOtraMoneda"]

        # ── DetallesItems ──
        items_data = cls._extract_items(row_dict, headers)
        if items_data:
            di = ET.SubElement(root, "DetallesItems")
            for item_data in items_data:
                cls._build_item_element(di, item_data)

        # ── Paginacion (solo si hay datos de paginas multiples) ──
        cls._build_paginacion(root, items_data, row_dict, headers)

        # ── DescuentosORecargos (document-level) ──
        cls._build_descuentos_recargos(root, row_dict, headers)

        # ── InformacionReferencia ──
        ncf_mod = cls._v(row_dict, headers, "NCFModificado")
        if ncf_mod:
            ir_elem = ET.SubElement(root, "InformacionReferencia")
            ET.SubElement(ir_elem, "NCFModificado").text = ncf_mod
            for tag in [
                "RNCOtroContribuyente", "FechaNCFModificado",
                "CodigoModificacion", "RazonModificacion",
            ]:
                v = cls._v(row_dict, headers, tag)
                if v:
                    ET.SubElement(ir_elem, tag).text = v

        # ── FechaHoraFirma ──
        now = datetime.now(timezone.utc)
        ET.SubElement(root, "FechaHoraFirma").text = now.strftime("%d-%m-%Y %H:%M:%S")

        xml_bytes = ET.tostring(root, encoding="utf-8")
        xml_str = xml_bytes.decode("utf-8")
        xml_str = xml_str.replace("\u00a9", "&copy;")
        xml_str = xml_str.replace("\u20ac", "&euro;")
        xml_str = xml_str.replace("\u00ae", "&reg;")
        return xml_str.encode("utf-8")

    # ═══════════════════════════════════════════════════════════════
    # Items
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def _extract_items(cls, row_dict, headers):
        items = []
        for line_num in range(1, cls.ITEM_LINE_COUNT + 1):
            nl = cls._v(row_dict, headers, f"NumeroLinea[{line_num}]")
            if not nl:
                continue
            item = {"NumeroLinea": nl}

            for field in [
                "IndicadorFacturacion", "NombreItem", "IndicadorBienoServicio",
                "DescripcionItem", "CantidadItem", "UnidadMedida",
                "CantidadReferencia", "UnidadReferencia",
                "PrecioUnitarioItem", "DescuentoMonto", "RecargoMonto",
                "MontoItem", "GradosAlcohol", "PrecioUnitarioReferencia",
                "FechaElaboracion", "FechaVencimientoItem",
                "PesoNetoKilogramo", "PesoNetoMineria",
                "TipoAfiliacion", "Liquidacion",
                "IndicadorAgenteRetencionoPercepcion",
                "MontoITBISRetenido", "MontoISRRetenido",
            ]:
                v = cls._v(row_dict, headers, f"{field}[{line_num}]")
                if v:
                    item[field] = v

            # Codigos
            for ci in range(1, cls.CODIGO_MAX + 1):
                tc = cls._v(row_dict, headers, f"TipoCodigo[{line_num}][{ci}]")
                cod = cls._v(row_dict, headers, f"CodigoItem[{line_num}][{ci}]")
                if tc and cod:
                    item.setdefault("codigos", []).append((tc, cod))

            # Subcantidades
            sc_list = []
            for si in range(1, cls.SUBCANTIDAD_MAX + 1):
                sc = cls._v(row_dict, headers, f"Subcantidad[{line_num}][{si}]")
                sco = cls._v(row_dict, headers, f"CodigoSubcantidad[{line_num}][{si}]")
                if sc:
                    sc_list.append((sc, sco or ""))
            if sc_list:
                item["subcantidades"] = sc_list

            # SubDescuentos
            desc_list = []
            for di in range(1, cls.SUB_DESCUENTO_MAX + 1):
                d_tipo = cls._v(row_dict, headers, f"TipoSubDescuento[{line_num}][{di}]")
                d_porc = cls._v(row_dict, headers, f"SubDescuentoPorcentaje[{line_num}][{di}]")
                d_monto = cls._v(row_dict, headers, f"MontoSubDescuento[{line_num}][{di}]")
                if d_tipo:
                    desc_list.append((d_tipo, d_porc or "", d_monto or ""))
            if desc_list:
                item["subdescuentos"] = desc_list

            # SubRecargos
            rec_list = []
            for ri in range(1, cls.SUB_RECARGO_MAX + 1):
                r_tipo = cls._v(row_dict, headers, f"TipoSubRecargo[{line_num}][{ri}]")
                r_porc = cls._v(row_dict, headers, f"SubRecargoPorcentaje[{line_num}][{ri}]")
                r_monto = cls._v(row_dict, headers, f"MontosubRecargo[{line_num}][{ri}]")
                if r_tipo:
                    rec_list.append((r_tipo, r_porc or "", r_monto or ""))
            if rec_list:
                item["subrecargos"] = rec_list

            # TipoImpuesto
            for ti in range(1, cls.TIPO_IMPUESTO_PER_ITEM + 1):
                tv = cls._v(row_dict, headers, f"TipoImpuesto[{line_num}][{ti}]")
                if tv:
                    item.setdefault("tiposImpuesto", []).append(tv)

            # OtraMoneda por item
            for om_field in [
                "PrecioOtraMoneda", "DescuentoOtraMoneda",
                "RecargoOtraMoneda", "MontoItemOtraMoneda",
            ]:
                v = cls._v(row_dict, headers, f"{om_field}[{line_num}]")
                if v:
                    item.setdefault("otraMoneda", {})[om_field.replace("OtraMoneda", "")] = v

            items.append(item)
        return items

    @classmethod
    def _build_item_element(cls, parent, item_data):
        item_elem = ET.SubElement(parent, "Item")
        ET.SubElement(item_elem, "NumeroLinea").text = item_data.get("NumeroLinea", "1")

        # TablaCodigosItem (wrapper segun XSD)
        codigos = item_data.get("codigos", [])
        if codigos:
            tci = ET.SubElement(item_elem, "TablaCodigosItem")
            for tc, cod in codigos:
                csi = ET.SubElement(tci, "CodigosItem")
                ET.SubElement(csi, "TipoCodigo").text = tc
                ET.SubElement(csi, "CodigoItem").text = cod

        # IndicadorFacturacion
        cls._add(item_elem, "IndicadorFacturacion", item_data.get("IndicadorFacturacion"))

        # Retencion — antes de NombreItem
        ret_ind = item_data.get("IndicadorAgenteRetencionoPercepcion")
        has_isr = item_data.get("MontoISRRetenido") or item_data.get("MontoITBISRetenido")
        if ret_ind or has_isr:
            ret_elem = ET.SubElement(item_elem, "Retencion")
            cls._add(ret_elem, "IndicadorAgenteRetencionoPercepcion", ret_ind)
            cls._add(ret_elem, "MontoITBISRetenido", item_data.get("MontoITBISRetenido"))
            cls._add(ret_elem, "MontoISRRetenido", item_data.get("MontoISRRetenido"))

        # NombreItem y campos base
        simple_fields = [
            "NombreItem", "IndicadorBienoServicio",
            "DescripcionItem", "CantidadItem", "UnidadMedida",
            "CantidadReferencia", "UnidadReferencia",
        ]
        for field in simple_fields:
            cls._add(item_elem, field, item_data.get(field))

        # TablaSubcantidad (wrapper segun XSD)
        subcants = item_data.get("subcantidades", [])
        if subcants:
            ts = ET.SubElement(item_elem, "TablaSubcantidad")
            for sc, sco in subcants:
                sci = ET.SubElement(ts, "SubcantidadItem")
                ET.SubElement(sci, "Subcantidad").text = sc
                if sco:
                    ET.SubElement(sci, "CodigoSubcantidad").text = sco

        # Campos despues de Subcantidad
        after_fields = [
            "GradosAlcohol", "PrecioUnitarioReferencia",
            "FechaElaboracion", "FechaVencimientoItem",
            "PesoNetoKilogramo", "PesoNetoMineria",
            "TipoAfiliacion", "Liquidacion",
        ]
        for field in after_fields:
            cls._add(item_elem, field, item_data.get(field))

        # PrecioUnitarioItem
        cls._add(item_elem, "PrecioUnitarioItem", item_data.get("PrecioUnitarioItem"))

        # DescuentoMonto
        cls._add(item_elem, "DescuentoMonto", item_data.get("DescuentoMonto"))

        # TablaSubDescuento (wrapper segun XSD)
        descs = item_data.get("subdescuentos", [])
        if descs:
            tsd = ET.SubElement(item_elem, "TablaSubDescuento")
            for tipo, porc, monto in descs:
                sd_elem = ET.SubElement(tsd, "SubDescuento")
                ET.SubElement(sd_elem, "TipoSubDescuento").text = tipo
                if porc:
                    ET.SubElement(sd_elem, "SubDescuentoPorcentaje").text = porc
                if monto:
                    ET.SubElement(sd_elem, "MontoSubDescuento").text = monto

        # RecargoMonto
        cls._add(item_elem, "RecargoMonto", item_data.get("RecargoMonto"))

        # TablaSubRecargo (wrapper segun XSD)
        recs = item_data.get("subrecargos", [])
        if recs:
            tsr = ET.SubElement(item_elem, "TablaSubRecargo")
            for tipo, porc, monto in recs:
                sr_elem = ET.SubElement(tsr, "SubRecargo")
                ET.SubElement(sr_elem, "TipoSubRecargo").text = tipo
                if porc:
                    ET.SubElement(sr_elem, "SubRecargoPorcentaje").text = porc
                if monto:
                    ET.SubElement(sr_elem, "MontoSubRecargo").text = monto

        # TablaImpuestoAdicional por item
        timps = item_data.get("tiposImpuesto", [])
        if timps:
            tia = ET.SubElement(item_elem, "TablaImpuestoAdicional")
            for tv in timps:
                ia_elem = ET.SubElement(tia, "ImpuestoAdicional")
                ET.SubElement(ia_elem, "TipoImpuesto").text = tv

        # OtraMonedaDetalle (wrapper segun XSD: es OtraMonedaDetalle, no OtraMoneda)
        om = item_data.get("otraMoneda", {})
        if om:
            om_elem = ET.SubElement(item_elem, "OtraMonedaDetalle")
            cls._add(om_elem, "PrecioOtraMoneda", om.get("Precio"))
            cls._add(om_elem, "DescuentoOtraMoneda", om.get("Descuento"))
            cls._add(om_elem, "RecargoOtraMoneda", om.get("Recargo"))
            cls._add(om_elem, "MontoItemOtraMoneda", om.get("MontoItem"))

        # MontoItem — siempre al final del Item
        cls._add(item_elem, "MontoItem", item_data.get("MontoItem"))

    @classmethod
    def _build_descuentos_recargos(cls, root, row_dict, headers):
        """Construye DescuentosORecargos (descuentos/recargos a nivel documento)."""
        entries = []
        for i in range(1, 21):  # up to 20 entries per XSD
            nl = cls._v(row_dict, headers, f"NumeroLineaDoR[{i}]")
            if not nl:
                continue
            entry = {"NumeroLinea": nl}
            for tag in [
                "TipoAjuste", "IndicadorNorma1007",
                "DescripcionDescuentooRecargo", "TipoValor",
                "ValorDescuentooRecargo", "MontoDescuentooRecargo",
                "MontoDescuentooRecargoOtraMoneda",
                "IndicadorFacturacionDescuentooRecargo",
            ]:
                v = cls._v(row_dict, headers, f"{tag}[{i}]")
                if v:
                    entry[tag] = v
            entries.append(entry)

        if entries:
            doR_elem = ET.SubElement(root, "DescuentosORecargos")
            for entry in entries:
                dr = ET.SubElement(doR_elem, "DescuentoORecargo")
                for tag, val in entry.items():
                    ET.SubElement(dr, tag).text = val

    # ═══════════════════════════════════════════════════════════════
    # Paginacion
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def _build_paginacion(cls, root, items_data, row_dict, headers):
        """Solo incluir Paginacion si hay mas de 1 pagina de items (> pagina 1)."""
        num_items = len(items_data)
        has_pagina2 = bool(cls._v(row_dict, headers, "PaginaNo[2]"))
        if num_items <= 0 or (not has_pagina2 and num_items <= 60):
            return  # 1 sola pagina, omitir Paginacion

        pag = ET.SubElement(root, "Paginacion")
        pagina = ET.SubElement(pag, "Pagina")
        ET.SubElement(pagina, "PaginaNo").text = "1"
        ET.SubElement(pagina, "NoLineaDesde").text = "1"
        ET.SubElement(pagina, "NoLineaHasta").text = str(num_items) if num_items > 0 else "1"

        for tag in [
            "SubtotalMontoGravadoPagina", "SubtotalMontoGravado1Pagina",
            "SubtotalMontoGravado2Pagina", "SubtotalMontoGravado3Pagina",
            "SubtotalExentoPagina", "SubtotalItbisPagina",
            "SubtotalItbis1Pagina", "SubtotalItbis2Pagina", "SubtotalItbis3Pagina",
            "SubtotalImpuestoAdicionalPagina", "MontoSubtotalPagina",
            "SubtotalMontoNoFacturablePagina",
        ]:
            v = cls._v(row_dict, headers, f"{tag}[1]")
            if v:
                ET.SubElement(pagina, tag).text = v

    # ═══════════════════════════════════════════════════════════════
    # RFCE (Resumen Factura Consumo Electronica) — XSD RFCE 32 v1.0
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def build_rfce_xml_from_row(cls, row_dict, headers, codigo_seguridad=None):
        """Construye XML RFCE usando los datos de la Hoja2 del Excel."""
        root = ET.Element("RFCE")
        enc = ET.SubElement(root, "Encabezado")
        ET.SubElement(enc, "Version").text = "1.0"

        # ── IdDoc ──
        id_doc = ET.SubElement(enc, "IdDoc")
        ET.SubElement(id_doc, "TipoeCF").text = "32"
        cls._add(id_doc, "eNCF", cls._v(row_dict, headers, "ENCF"))
        cls._add(id_doc, "TipoIngresos", cls._v(row_dict, headers, "TipoIngresos"))
        cls._add(id_doc, "TipoPago", cls._v(row_dict, headers, "TipoPago"))

        formas = []
        for i in range(1, 8):
            fp = cls._v(row_dict, headers, f"FormaPago[{i}]")
            mp = cls._v(row_dict, headers, f"MontoPago[{i}]")
            if fp:
                formas.append((fp, mp or "0.00"))
        if formas:
            tfp = ET.SubElement(id_doc, "TablaFormasPago")
            for fp, mp in formas:
                fdp = ET.SubElement(tfp, "FormaDePago")
                ET.SubElement(fdp, "FormaPago").text = fp
                ET.SubElement(fdp, "MontoPago").text = mp

        # ── Emisor ──
        emisor_elem = ET.SubElement(enc, "Emisor")
        cls._add(emisor_elem, "RNCEmisor", cls._v(row_dict, headers, "RNCEmisor"))
        cls._add(emisor_elem, "RazonSocialEmisor", cls._v(row_dict, headers, "RazonSocialEmisor"))
        cls._add(emisor_elem, "FechaEmision", cls._v(row_dict, headers, "FechaEmision"))

        # ── Comprador ──
        # Copia verbatim de la Hoja 2 (RFCE): el conjunto de datos de la DGII
        # espera RNCComprador=131880681 y RazonSocialComprador tal cual el Excel.
        comp_children = []
        for tag in ["RNCComprador", "IdentificadorExtranjero", "RazonSocialComprador"]:
            v = cls._v(row_dict, headers, tag)
            if v:
                comp_children.append((tag, v))
        if comp_children:
            comp_elem = ET.SubElement(enc, "Comprador")
            for tag, val in comp_children:
                ET.SubElement(comp_elem, tag).text = val

        # ── Totales ──
        totales_elem = ET.SubElement(enc, "Totales")
        for tag in [
            "MontoGravadoTotal", "MontoGravadoI1", "MontoGravadoI2", "MontoGravadoI3",
            "MontoExento", "TotalITBIS", "TotalITBIS1", "TotalITBIS2", "TotalITBIS3",
            "MontoImpuestoAdicional",
        ]:
            cls._add(totales_elem, tag, cls._v(row_dict, headers, tag))

        # ImpuestosAdicionales wrapper
        imp_children = []
        for i in range(1, 5):
            tipo_imp = cls._v(row_dict, headers, f"TipoImpuesto[{i}]")
            if not tipo_imp:
                continue
            imp_children.append({
                "TipoImpuesto": tipo_imp,
                "MontoImpuestoSelectivoConsumoEspecifico": cls._v(row_dict, headers, f"MontoImpuestoSelectivoConsumoEspecifico[{i}]") or "",
                "MontoImpuestoSelectivoConsumoAdvalorem": cls._v(row_dict, headers, f"MontoImpuestoSelectivoConsumoAdvalorem[{i}]") or "",
                "OtrosImpuestosAdicionales": cls._v(row_dict, headers, f"OtrosImpuestosAdicionales[{i}]") or "",
            })
        if imp_children:
            imp_elem = ET.SubElement(totales_elem, "ImpuestosAdicionales")
            for imp in imp_children:
                ia = ET.SubElement(imp_elem, "ImpuestoAdicional")
                ET.SubElement(ia, "TipoImpuesto").text = imp["TipoImpuesto"]
                if imp["MontoImpuestoSelectivoConsumoEspecifico"]:
                    ET.SubElement(ia, "MontoImpuestoSelectivoConsumoEspecifico").text = imp["MontoImpuestoSelectivoConsumoEspecifico"]
                if imp["MontoImpuestoSelectivoConsumoAdvalorem"]:
                    ET.SubElement(ia, "MontoImpuestoSelectivoConsumoAdvalorem").text = imp["MontoImpuestoSelectivoConsumoAdvalorem"]
                if imp["OtrosImpuestosAdicionales"]:
                    ET.SubElement(ia, "OtrosImpuestosAdicionales").text = imp["OtrosImpuestosAdicionales"]

        for tag in [
            "MontoTotal", "MontoNoFacturable", "MontoPeriodo",
        ]:
            cls._add(totales_elem, tag, cls._v(row_dict, headers, tag))



        # ── CodigoSeguridadeCF (requerido, del E32 original firmado) ──
        if codigo_seguridad:
            ET.SubElement(enc, "CodigoSeguridadeCF").text = codigo_seguridad[:6]
        else:
            # Fallback: SHA-256 del ENCF + RNC
            import hashlib
            encf_rfce = cls._v(row_dict, headers, "ENCF") or ""
            rnc_rfce = cls._v(row_dict, headers, "RNCEmisor") or ""
            raw = f"{rnc_rfce}{encf_rfce}".encode("utf-8")
            codigo = hashlib.sha256(raw).hexdigest()[:6].upper()
            ET.SubElement(enc, "CodigoSeguridadeCF").text = codigo

        # RFCE no lleva FechaHoraFirma — el XSD solo tiene Encabezado + <any> (firma)

        xml_bytes = ET.tostring(root, encoding="utf-8")
        xml_str = xml_bytes.decode("utf-8")
        return xml_str.encode("utf-8")

    @classmethod
    def get_rfce_rows(cls, sheet2_data):
        return sheet2_data

    # ═══════════════════════════════════════════════════════════════
    # ACECF (Aprobacion Comercial Electronica) — XSD ACECF v1.0
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def build_acecf_xml_from_row(cls, row_dict, headers):
        """Construye XML ACECF desde una fila del Excel de Aprobaciones Comerciales."""
        root = ET.Element("ACECF")
        detalle = ET.SubElement(root, "DetalleAprobacionComercial")

        for tag in [
            "Version", "RNCEmisor", "eNCF", "FechaEmision",
        ]:
            cls._add(detalle, tag, cls._v(row_dict, headers, tag))

        # MontoTotal requiere 2 decimales — va entre FechaEmision y RNCComprador
        mt = cls._v(row_dict, headers, "MontoTotal")
        if mt:
            try:
                mt = f"{float(mt):.2f}"
            except ValueError:
                pass
            cls._add(detalle, "MontoTotal", mt)

        for tag in [
            "RNCComprador", "Estado",
            "DetalleMotivoRechazo", "FechaHoraAprobacionComercial",
        ]:
            cls._add(detalle, tag, cls._v(row_dict, headers, tag))

        xml_bytes = ET.tostring(root, encoding="utf-8")
        return xml_bytes
