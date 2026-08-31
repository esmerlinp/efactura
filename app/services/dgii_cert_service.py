import base64
import hashlib
import json
import os
import shutil
import time
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from io import BytesIO

from app.services.db_service import db_firestore, DatabaseService
from app.services.dgii import DGIIService
from app.services.dgii_signer import DgiiSigner
from app.services.dgii_test_data_loader import DgiiTestDataLoader
from app.services.dgii_direct import DgiiDirectService
from app.services.dgii_xml_builder import DgiiXmlBuilder
from app.services.supplier_invoice_service import SupplierInvoiceService
from app.models.certificacion import CertificacionProcess, CertStep, CertRun, CaseResult
from config import Config

RFCE_THRESHOLD = 250000.00

DGII_ORDER_GROUP1 = ["31", "32", "41", "43", "44", "45", "46", "47"]
DGII_ORDER_GROUP2 = ["33", "34"]

CERT_COLLECTION = "certificacion"

# Ambiente de DGII para certificación — siempre CerteCF, independiente de .env
CERT_DGII_ENVIRONMENT = "CerteCF"

ECF_BASE = f"https://ecf.dgii.gov.do/{CERT_DGII_ENVIRONMENT}"
FC_BASE = f"https://fc.dgii.gov.do/{CERT_DGII_ENVIRONMENT}"


def _now():
    return datetime.now(timezone.utc).isoformat()


def _get_evidence_dir(company_id, step, run_number):
    return f"uploads/certificacion/{company_id}/step{step}/run{run_number}"


def _ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def _get_cert_doc_path(company_id):
    return f"companies/{company_id}/{CERT_COLLECTION}/process"


def _get_run_doc_path(company_id, step, run_number):
    return f"companies/{company_id}/{CERT_COLLECTION}/process/runs/step{step}_run{run_number}"


class DgiiCertService:

    # ═══════════════════════════════════════════════════════════════
    # Persistencia
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def _get_firestore_doc(cls, path):
        doc_ref = db_firestore.document(path)
        snap = doc_ref.get()
        if snap.exists:
            return snap.to_dict()
        return None

    @classmethod
    def _set_firestore_doc(cls, path, data, merge=False):
        doc_ref = db_firestore.document(path)
        doc_ref.set(data, merge=merge)

    @classmethod
    def get_process(cls, company_id):
        path = _get_cert_doc_path(company_id)
        doc = cls._get_firestore_doc(path)
        if doc:
            doc["id"] = company_id
            cls._migrate_light_runs(company_id, doc)
            return doc
        return {
            "id": company_id,
            "current_step": 1,
            "steps": {},
            "created_at": _now(),
            "updated_at": _now(),
        }

    @classmethod
    def _light_run_summary(cls, run_dict):
        """Copia ligera del run_dict para el doc de proceso (evita superar 1MB)."""
        light = dict(run_dict or {})
        test_set = light.pop("test_set", None)
        if test_set:
            light["test_set_summary"] = {
                "total": test_set.get("total", 0),
                "set_errors": (test_set.get("set_errors") or [])[:20],
                "blocks": [
                    {
                        "index": b.get("index"),
                        "tipo": b.get("tipo"),
                        "label": b.get("label"),
                        "count": b.get("count"),
                        "status": b.get("status"),
                        "sent": b.get("sent_count", 0),
                        "failed": b.get("failed_count", 0),
                    }
                    for b in (test_set.get("blocks") or [])
                ],
            }
        if isinstance(light.get("cases"), list):
            light["cases"] = light["cases"][:100]
        return light

    @classmethod
    def _migrate_light_runs(cls, company_id, process):
        """Limpia runs viejos que llevan test_set completo dentro del doc de proceso."""
        changed = False
        for step_key, step in (process.get("steps") or {}).items():
            if not isinstance(step, dict):
                continue
            runs = step.get("runs") or []
            for i, run in enumerate(runs):
                if isinstance(run, dict) and run.get("test_set"):
                    runs[i] = cls._light_run_summary(run)
                    changed = True
            if changed:
                step["runs"] = runs
        if changed:
            try:
                cls.save_process(company_id, process)
            except Exception as e:
                print(f"⚠️ No se pudo migrar el doc de proceso a runs ligeros: {e}")

    @classmethod
    def save_process(cls, company_id, process_dict):
        process_dict["updated_at"] = _now()
        path = _get_cert_doc_path(company_id)
        cls._set_firestore_doc(path, process_dict, merge=True)

    @classmethod
    def is_certification_completed(cls, company_id):
        process = cls.get_process(company_id)
        steps = process.get("steps", {})
        step_15 = steps.get("15", {})
        return step_15.get("status") == "completed"

    @classmethod
    def is_certification_locked(cls, company_id):
        return cls.is_certification_completed(company_id)

    @classmethod
    def get_step_status(cls, company_id):
        process = cls.get_process(company_id)
        current_step = process.get("current_step", 1)
        steps = process.get("steps", {})
        return {
            "current_step": current_step,
            "steps": steps,
        }

    @classmethod
    def init_step(cls, company_id, step_num):
        process = cls.get_process(company_id)
        steps = process.get("steps", {})
        step_key = str(step_num)

        existing = steps.get(step_key, {})
        current_run = existing.get("current_run", 0) + 1
        runs = existing.get("runs", [])

        step_data = {
            "status": "in_progress",
            "current_run": current_run,
            "runs": runs,
            "started_at": _now(),
            "completed_at": None,
        }
        steps[step_key] = step_data
        process["steps"] = steps
        process["current_step"] = step_num
        cls.save_process(company_id, process)

        return current_run, step_data

    @classmethod
    def complete_step(cls, company_id, step_num, run_number, run_dict):
        process = cls.get_process(company_id)
        steps = process.get("steps", {})
        step_key = str(step_num)
        step_data = steps.get(step_key, {})

        run_dict["status"] = "completed"
        run_dict["completed_at"] = _now()

        runs = step_data.get("runs", [])
        runs.append(cls._light_run_summary(run_dict))
        step_data["runs"] = runs
        step_data["status"] = "completed"
        step_data["completed_at"] = _now()

        run_path = _get_run_doc_path(company_id, step_num, run_number)
        cls._set_firestore_doc(run_path, run_dict)

        steps[step_key] = step_data
        process["steps"] = steps
        cls.save_process(company_id, process)

    @classmethod
    def fail_step(cls, company_id, step_num, run_number, run_dict):
        process = cls.get_process(company_id)
        steps = process.get("steps", {})
        step_key = str(step_num)
        step_data = steps.get(step_key, {})

        run_dict["status"] = "failed"
        run_dict["completed_at"] = _now()

        runs = step_data.get("runs", [])
        runs.append(cls._light_run_summary(run_dict))
        step_data["runs"] = runs
        step_data["status"] = "failed"
        step_data["completed_at"] = _now()

        run_path = _get_run_doc_path(company_id, step_num, run_number)
        cls._set_firestore_doc(run_path, run_dict)

        steps[step_key] = step_data
        process["steps"] = steps
        cls.save_process(company_id, process)

    @classmethod
    def save_run_progress(cls, company_id, step_num, run_number, run_dict):
        run_dict["updated_at"] = _now()
        run_path = _get_run_doc_path(company_id, step_num, run_number)
        cls._set_firestore_doc(run_path, run_dict)

    @classmethod
    def mark_step_skipped(cls, company_id, step_num):
        process = cls.get_process(company_id)
        steps = process.get("steps", {})
        step_key = str(step_num)
        steps[step_key] = {
            "status": "completed",
            "current_run": 0,
            "runs": [],
            "completed_at": _now(),
        }
        process["steps"] = steps
        cls.save_process(company_id, process)

    @classmethod
    def set_current_step_manual(cls, company_id, step_num):
        process = cls.get_process(company_id)
        process["current_step"] = step_num
        cls.save_process(company_id, process)

    @classmethod
    def get_run(cls, company_id, step_num, run_number):
        run_path = _get_run_doc_path(company_id, step_num, run_number)
        return cls._get_firestore_doc(run_path)

    # ═══════════════════════════════════════════════════════════════
    # Validacion de certificado
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def validate_certificate(cls, company_profile):
        cert_content = company_profile.get("certificateContent") or company_profile.get("certificate_content") or ""
        cert_password = company_profile.get("certificatePassword") or company_profile.get("certificate_password") or ""
        cert_name = company_profile.get("certificateName") or company_profile.get("certificate_name") or ""

        if not cert_content:
            return {"valid": False, "error": "No hay certificado digital cargado."}

        try:
            cert_bundle = DgiiSigner.export_pem_bundle(company_profile)
            if not cert_bundle:
                return {"valid": False, "error": "No se pudo leer el certificado."}
        except Exception as e:
            return {"valid": False, "error": f"Error al leer el certificado: {str(e)}"}

        try:
            from cryptography import x509
            from cryptography.hazmat.backends import default_backend
            cert_data = base64.b64decode(cert_content)
            cert = x509.load_pkcs12(cert_data, cert_password.encode(), default_backend())
            not_after = cert.certificate.not_valid_after_utc.isoformat()
            subject = str(cert.certificate.subject)
            return {
                "valid": True,
                "not_after": not_after,
                "subject": subject,
                "name": cert_name,
            }
        except Exception:
            return {
                "valid": True,
                "name": cert_name,
            }

    # ═══════════════════════════════════════════════════════════════
    # Endpoints y autenticación específicos de certificación (CerteCF)
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def _cert_endpoints(cls):
        return {
            "semilla": f"{ECF_BASE}/autenticacion/api/autenticacion/semilla",
            "validar_semilla": f"{ECF_BASE}/autenticacion/api/autenticacion/validarsemilla",
            "recepcion": f"{ECF_BASE}/recepcion/api/facturaselectronicas",
            "rfce_recepcion": f"{FC_BASE}/recepcionfc/api/recepcion/ecf",
            "consulta_resultado": f"{ECF_BASE}/consultaresultado/api/consultas/estado",
            "consulta_estado": f"{ECF_BASE}/consultaestado/api/consultas/estado",
            "aprobacion_comercial": f"{ECF_BASE}/aprobacioncomercial/api/aprobacioncomercial",
        }

    @classmethod
    def _get_cert_token(cls, company_profile):
        import requests
        from xml.sax.saxutils import escape

        endpoints = cls._cert_endpoints()
        semilla_url = endpoints.get("semilla")
        validar_url = endpoints.get("validar_semilla")

        if not semilla_url or not validar_url:
            return None, "Certification endpoints not configured"

        cert_path = DgiiDirectService._prepare_tls_cert(company_profile)

        try:
            headers = DgiiDirectService._build_headers()
            response = requests.get(
                semilla_url, headers=headers, cert=cert_path,
                timeout=Config.DGII_HTTP_TIMEOUT,
            )
            response_data = DgiiDirectService._safe_json(response)
            response_text = response.text if response is not None else ""

            if response.status_code < 200 or response.status_code >= 300:
                return None, f"DGII auth error HTTP {response.status_code}"

            token = DgiiDirectService._extract_token(response_data, response_text)
            if token:
                return token, None

            seed, fecha_seed = DgiiDirectService._extract_seed_xml(response_data, response_text)
            if not seed:
                return None, "No se pudo obtener la semilla de autenticacion (CerteCF)."

            fecha_str = fecha_seed if fecha_seed else datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
            semilla_xml = (
                '<?xml version="1.0" encoding="utf-8"?>'
                f"<SemillaModel>"
                f"<valor>{escape(seed)}</valor>"
                f"<fecha>{escape(fecha_str)}</fecha>"
                f"</SemillaModel>"
            ).encode("utf-8")

            signed_semilla = DgiiSigner.sign_xml(semilla_xml, company_profile)

            token_response = DgiiDirectService._multipart_post(
                validar_url, signed_semilla, token=None,
                filename="signed_seed.xml", cert_path=cert_path,
            )

            if token_response is None or token_response.status_code < 200 or token_response.status_code >= 300:
                return None, f"Error validando semilla (CerteCF): HTTP {token_response.status_code if token_response else 'N/A'}"

            t_data = DgiiDirectService._safe_json(token_response)
            t_text = token_response.text if token_response else ""
            token = DgiiDirectService._extract_token(t_data, t_text)
            if token:
                return token, None

            return None, f"No se pudo extraer token de la respuesta (CerteCF)."
        except Exception as e:
            return None, f"Error autenticando con DGII CerteCF: {str(e)}"

    # ═══════════════════════════════════════════════════════════════
    # Envío a DGII (forzando CerteCF)
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def _send_ecf(cls, company_profile, signed_xml, token, caso):
        try:
            from app.services.dgii_signer import DgiiSigner
            from cryptography import x509
            
            endpoints = cls._cert_endpoints()
            url = endpoints.get("recepcion")
            cert_path = DgiiDirectService._prepare_tls_cert(company_profile)
            safe_encf = caso['encf'].strip()
            
            company_rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
            
            # The DGII CerteCF environment expects strictly RNC + eNCF + .xml
            filename = f"{company_rnc}{safe_encf}.xml"
            print(f"DEBUG_SEND_ECF: filename='{filename}' len={len(filename)}", flush=True)
            response = DgiiDirectService._multipart_post(
                url, signed_xml, token=token, filename=filename, cert_path=cert_path
            )
            text = response.text if response else ""
            data = DgiiDirectService._safe_json(response) if response else None

            track_id = DgiiDirectService._extract_track_id(data, text)
            dgii_status = DgiiDirectService._extract_status(data, text)
            print(f"DEBUG_SEND_ECF_RES: status_code={response.status_code if response else 'None'} dgii_status='{dgii_status}' text='{text}'", flush=True)

            is_success = response is not None and 200 <= response.status_code < 300
            
            # DGII CerteCF sometimes returns 200 even for Rejections
            if "secuencia ya ha sido utilizado" in text.lower() or "secuencia ya ha sido utilizado" in str(data).lower():
                is_success = True
                dgii_status = "ACCEPTED_PREVIOUSLY"
            elif dgii_status == "REJECTED":
                is_success = False

            if is_success and not dgii_status:
                dgii_status = "ACCEPTED"

            return {
                "success": is_success,
                "track_id": track_id,
                "dgii_status": dgii_status or "UNKNOWN",
                "response_data": data or {},
            }
        except Exception as e:
            return {"success": False, "error_message": str(e)}

    @classmethod
    def _send_rfce(cls, company_profile, signed_xml, token, caso):
        try:
            endpoints = cls._cert_endpoints()
            url = endpoints.get("rfce_recepcion")
            if not url:
                return {"success": False, "error_message": "RFCE endpoint not configured"}
            cert_path = DgiiDirectService._prepare_tls_cert(company_profile)
            safe_encf = caso['encf'].strip()
            company_rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
            filename = f"{company_rnc}{safe_encf}.xml"
            
            response = DgiiDirectService._multipart_post(
                url, signed_xml, token=token, filename=filename, cert_path=cert_path
            )
            text = response.text if response else ""
            data = DgiiDirectService._safe_json(response) if response else None

            track_id = DgiiDirectService._extract_track_id(data, text)
            dgii_status = DgiiDirectService._extract_status(data, text)

            is_success = response is not None and 200 <= response.status_code < 300
            
            if "secuencia ya ha sido utilizado" in text.lower() or "secuencia ya ha sido utilizado" in str(data).lower():
                is_success = True
                dgii_status = "ACCEPTED_PREVIOUSLY"
            elif dgii_status == "REJECTED":
                is_success = False

            return {
                "success": is_success,
                "track_id": track_id,
                "dgii_status": dgii_status or "UNKNOWN",
                "response_data": data or {},
            }
        except Exception as e:
            return {"success": False, "error_message": str(e)}

    # ═══════════════════════════════════════════════════════════════
    # Paso 2: Pruebas de Datos e-CF
    # ═══════════════════════════════════════════════════════════════


    @classmethod
    def force_dgii_reset(cls, company_id, company_profile, parsed_data, group, count):
        from app.services.dgii_direct import DgiiDirectService
        grupos = parsed_data.get("_grupos_raw", {})
        cases = grupos.get(str(group), [])
        if not cases:
            return {"success": False, "error": "No hay casos en este grupo"}
            
        base_caso = cases[0]
        row_dict = dict(base_caso["row_dict"])
        headers = dict(base_caso["headers"])
        tipo = base_caso.get("tipo", "31")
        
        token = DgiiDirectService.get_dgii_token(company_profile, sandbox=True)
        if not token:
            return {"success": False, "error": "No se pudo obtener el token de DGII"}
            
        results = []
        for i in range(count):
            encf_key = next((k for k, v in headers.items() if v.strip() == 'ENCF'), None)
            fake_encf = f"E{tipo}000000099{i}"
            if encf_key:
                row_dict[encf_key] = fake_encf
            else:
                row_dict["ENCF"] = fake_encf
                headers["ENCF"] = "ENCF"
                
            raw_xml = DgiiTestDataLoader.build_xml_from_row(row_dict, headers)
            signed_xml = DgiiSigner.sign_xml(raw_xml, company_profile)
            
            res = cls._send_ecf(company_profile, signed_xml, token, {"encf": fake_encf})
            results.append(res)
            
        return {"success": True, "results": results}

    @classmethod
    def parse_step2_excel(cls, excel_path):
        sheet1_rows, sheet2_rows = DgiiTestDataLoader.load_workbook(excel_path)

        rfce_map = {}
        for row_dict, headers in sheet2_rows:
            encf = DgiiTestDataLoader._v(row_dict, headers, "ENCF") or row_dict.get("D", "")
            if encf:
                rfce_map[str(encf).strip()] = (row_dict, headers)

        casos_map = {}
        for row_dict, headers in sheet1_rows:
            tipo = row_dict.get("C", "?")
            encf = str(row_dict.get("D", f"E{tipo}??????")).strip()
            total_str = row_dict.get(
                "EW",
                row_dict.get(
                    next((c for c, h in headers.items() if h.strip() == "MontoTotal"), ""), "0"
                ),
            )
            total = float(total_str) if total_str else 0.0
            casos_map[encf] = {"row_dict": row_dict, "headers": headers, "tipo": tipo, "total": total}

        grupos = {"1": [], "2": [], "3": [], "4": []}
        for encf, caso in casos_map.items():
            tipo = caso["tipo"]
            total = caso["total"]
            is_e32 = tipo == "32"
            is_rfce = is_e32 and total < RFCE_THRESHOLD

            if is_rfce:
                if encf in rfce_map:
                    rfce_row, rfce_hdrs = rfce_map[encf]
                    grupos["3"].append({**caso, "encf": encf, "tag": "rfce", "rfce_row_dict": rfce_row, "rfce_headers": rfce_hdrs})
                else:
                    grupos["3"].append({**caso, "encf": encf, "tag": "rfce"})
                grupos["4"].append({**caso, "encf": encf, "tag": "manual_upload"})
            elif tipo in DGII_ORDER_GROUP1:
                grupos["1"].append({**caso, "encf": encf, "tag": "e-cf"})
            elif tipo in DGII_ORDER_GROUP2:
                grupos["2"].append({**caso, "encf": encf, "tag": "e-cf"})

        def sort_key(item):
            t = item["tipo"]
            if t in DGII_ORDER_GROUP1:
                return (DGII_ORDER_GROUP1.index(t), item["encf"])
            elif t in DGII_ORDER_GROUP2:
                return (10 + DGII_ORDER_GROUP2.index(t), item["encf"])
            return (99, item["encf"])

        for g in ["1", "2", "3", "4"]:
            grupos[g].sort(key=sort_key)

        preview = []
        for g in ["1", "2", "3", "4"]:
            for caso in grupos[g]:
                preview.append({
                    "encf": caso["encf"],
                    "tipo": caso["tipo"],
                    "total": caso["total"],
                    "grupo": int(g),
                    "tag": caso["tag"],
                })

        return {
            "total_cases": len(preview),
            "grupo_1_count": len(grupos["1"]),
            "grupo_2_count": len(grupos["2"]),
            "grupo_3_count": len(grupos["3"]),
            "grupo_4_count": len(grupos["4"]),
            "casos": preview,
            "_grupos_raw": grupos,
        }

    @classmethod
    def _cached_e32_matches(cls, cached_signed_bytes, fresh_raw_bytes):
        """
        True si el E32 firmado cacheado tiene el mismo contenido unsigned que el
        raw recién generado, ignorando FechaHoraFirma (cambia en cada build) y el
        bloque ds:Signature (firma). Compara c14n para no depender del orden de
        atributos/espacios del serializador.
        """
        try:
            from lxml import etree
        except Exception:
            return False

        DS = "{http://www.w3.org/2000/09/xmldsig#}"

        def _normalize(data_bytes, is_signed):
            root = etree.fromstring(data_bytes)
            if is_signed:
                for sig in root.findall(f".//{DS}Signature"):
                    parent = sig.getparent()
                    if parent is not None:
                        parent.remove(sig)
            fhf = root.find("FechaHoraFirma")
            if fhf is not None:
                fhf.text = ""
            return etree.tostring(root, method="c14n")

        try:
            return _normalize(cached_signed_bytes, True) == _normalize(fresh_raw_bytes, False)
        except Exception:
            return False

    @classmethod
    def process_step2_generate(cls, company_id, company_profile, parsed_data, selected_groups=None,
                               dry_run=False, run_number=1, resume_run=False, force_rerun=False):
        if selected_groups is None:
            selected_groups = ["1", "2", "3", "4"]
        selected_groups = [str(g) for g in selected_groups]

        grupos = parsed_data.get("_grupos_raw", {})
        evidence_dir = _get_evidence_dir(company_id, 2, run_number)
        xml_dir = os.path.join(evidence_dir, "xml")
        _ensure_dir(xml_dir)

        token = None
        results = []
        if resume_run:
            existing = cls.get_run(company_id, 2, run_number) or {}
            results = existing.get("cases", [])
            
        success_map = {f"{c['encf']}_{c.get('grupo')}": c for c in results if c.get("success")}
        if force_rerun:
            for k in list(success_map.keys()):
                if any(k.endswith(f"_{g}") for g in selected_groups):
                    del success_map[k]

        all_cases = []
        total_in_excel = 0

        for g in ["1", "2", "3", "4"]:
            for caso in grupos.get(g, []):
                total_in_excel += 1

                if g not in selected_groups:
                    continue
                if f"{caso['encf']}_{g}" in success_map and not dry_run:
                    continue
                all_cases.append((g, caso))

        total = len(all_cases)
        accepted = rejected = pending_count = 0

        for idx, (g, caso) in enumerate(all_cases, 1):
            encf = caso["encf"]
            tipo = caso["tipo"]
            row_dict = caso["row_dict"]      # Sheet 1 data — used for full e-CF 32 XML
            headers = caso["headers"]         # Sheet 1 headers
            rfce_row_dict = caso.get("rfce_row_dict", row_dict)  # Sheet 2 data — used for RFCE XML
            rfce_headers = caso.get("rfce_headers", headers)     # Sheet 2 headers
            tag = caso["tag"]
            total_monto = caso["total"]
            is_rfce = (tag == "rfce")

            case_result = {
                "encf": encf, "tipo": tipo, "total": total_monto,
                "grupo": int(g), "tag": tag, "success": False,
            }

            try:
                e32_signed_path = os.path.join(xml_dir, f"{encf}_e32_firmado.xml")
                
                # 1. Generate full raw XML
                full_raw_xml = DgiiTestDataLoader.build_xml_from_row(row_dict, headers)
                
                # 2. Get or create full signed E32 (crucial to share exact signature between Group 3 and 4)
                #    La DGII exige que CodigoSeguridadeCF (RFCE) == SignatureValue[:6]
                #    del E32 completo que se sube al portal. Reutilizar el cache cuando
                #    el contenido coincide (ignorando FechaHoraFirma y firma) mantiene
                #    la firma estable entre clics y entre corridas.
                full_signed_xml = None
                if os.path.exists(e32_signed_path):
                    with open(e32_signed_path, "rb") as f:
                        cached_content = f.read()
                        if (b"<ECF>" in cached_content and b"Signature" in cached_content
                                and b"SIMULATION_SIGNATURE" not in cached_content
                                and cls._cached_e32_matches(cached_content, full_raw_xml)):
                            full_signed_xml = cached_content

                if not full_signed_xml:
                    full_signed_xml = DgiiSigner.sign_xml(full_raw_xml, company_profile)
                    with open(e32_signed_path, "wb") as f:
                        f.write(full_signed_xml)
                    case_result["firma_actualizada"] = True
                
                import hashlib
                sv = DgiiSigner.extract_signature_value(full_signed_xml) or ""
                codigo_seg = sv[:6] if len(sv) >= 6 else hashlib.sha256(full_signed_xml).hexdigest()[:6]

                # 3. Prepare final XML
                if is_rfce:
                    raw_xml = DgiiTestDataLoader.build_rfce_xml_from_row(rfce_row_dict, rfce_headers, codigo_seg)
                    signed_xml = DgiiSigner.sign_xml(raw_xml, company_profile)
                    case_result["codigo_seguridad"] = codigo_seg
                else:
                    raw_xml = full_raw_xml
                    signed_xml = full_signed_xml

                raw_path = os.path.join(xml_dir, f"{encf}_raw.xml")
                with open(raw_path, "wb") as f:
                    f.write(raw_xml)
                case_result["raw_xml_path"] = raw_path

                # Separate file paths for RFCE (Group 3) vs e-CF 32 (Group 4) to avoid collisions
                if is_rfce:
                    signed_path = os.path.join(xml_dir, f"{encf}_rfce_signed.xml")
                else:
                    signed_path = os.path.join(xml_dir, f"{encf}_signed.xml")
                with open(signed_path, "wb") as f:
                    f.write(signed_xml)
                case_result["signed_xml_path"] = signed_path

                if dry_run:
                    case_result["success"] = True
                    case_result["dry_run"] = True
                    case_result["dgii_status"] = "DRY_RUN"
                    accepted += 1
                    results.append(case_result)
                    continue

                if g == "4":
                    manual_path = os.path.join(xml_dir, f"{encf}_manual_signed.xml")
                    with open(manual_path, "wb") as f:
                        f.write(full_signed_xml)
                    case_result["signed_xml_path"] = manual_path
                    case_result["success"] = True
                    case_result["codigo_seguridad"] = codigo_seg
                    nota = "Subir manualmente en portal DGII > Facturas de consumo < 250Mil"
                    if case_result.get("firma_actualizada"):
                        nota += (" | ATENCION: la firma cambio - reenvia el grupo 3 (RFCE) con el "
                                 "CodigoSeguridadeCF actualizado ANTES de subir este XML al portal")
                    case_result["nota"] = nota
                    accepted += 1
                    results.append(case_result)
                    continue

                if token is None and g in ("1", "2", "3"):
                    token, err = cls._get_cert_token(company_profile)
                    if err:
                        case_result["error_message"] = f"Error de autenticacion DGII: {err}"
                        rejected += 1
                        results.append(case_result)
                        continue

                if is_rfce and tag == "rfce":
                    result = cls._send_rfce(company_profile, signed_xml, token, caso)
                else:
                    result = cls._send_ecf(company_profile, signed_xml, token, caso)

                case_result.update({
                    "success": result.get("success", False),
                    "track_id": result.get("track_id"),
                    "dgii_status": result.get("dgii_status"),
                    "response_data": result.get("response_data", {}),
                    "error_message": result.get("error_message"),
                })

                if case_result["success"]:
                    accepted += 1
                else:
                    rejected += 1

            except Exception as e:
                case_result["error_message"] = str(e)
                rejected += 1

            idx_in_results = next((i for i, r in enumerate(results) if r["encf"] == encf and str(r.get("grupo", "")) == str(g)), -1)
            if idx_in_results >= 0:
                results[idx_in_results] = case_result
            else:
                results.append(case_result)
            time.sleep(0.3)

        total_accepted = sum(1 for c in results if c.get("success"))
        total_rejected = sum(1 for c in results if not c.get("success"))
        is_fully_completed = (total_accepted == total_in_excel)

        run_dict = {
            "run_number": run_number,
            "step": 2,
            "status": "completed" if is_fully_completed else ("failed" if total_rejected > 0 else "in_progress"),
            "started_at": _now(),
            "total_cases": total_in_excel,
            "accepted": total_accepted,
            "rejected": total_rejected,
            "pending": total_in_excel - (total_accepted + total_rejected),
            "manual_uploaded": 0,
            "cases": results,
            "evidencias_dir": evidence_dir,
            "dry_run": dry_run,
        }

        if total_rejected > 0:
            cls.fail_step(company_id, 2, run_number, run_dict)
            run_dict["status"] = "failed"
        elif is_fully_completed:
            cls.complete_step(company_id, 2, run_number, run_dict)
            run_dict["status"] = "completed"
        else:
            cls.save_run_progress(company_id, 2, run_number, run_dict)

        return {
            "success": total_rejected == 0,
            "total": total_in_excel,
            "accepted": total_accepted,
            "rejected": total_rejected,
            "results": results,
            "run_number": run_number,
            "evidence_dir": evidence_dir,
        }

    @classmethod
    def check_case_status(cls, company_profile, track_id):
        try:
            endpoints = cls._cert_endpoints()
            consulta_url = endpoints.get("consulta_estado")
            cert_path = DgiiDirectService._prepare_tls_cert(company_profile)

            result = DgiiDirectService.check_status(
                company_profile, track_id, sandbox=None
            )
            result["success"] = result.get("success", False)
            return {
                "success": result.get("success", False),
                "dgii_status": result.get("dgiiStatus", "UNKNOWN"),
                "track_id": track_id,
                "mensajes": result.get("mensajes", []),
                "response": result,
            }
        except Exception as e:
            return {"success": False, "dgii_status": "ERROR", "track_id": track_id, "error": str(e)}

    # ═══════════════════════════════════════════════════════════════
    # Paso 3: Pruebas de Datos Aprobación Comercial (ACECF)
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def parse_step3_excel(cls, excel_path):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        def col_letter(n):
            r = ""
            while n > 0:
                n -= 1
                r = chr(n % 26 + ord("A")) + r
                n //= 26
            return r

        headers = {}
        for cell in ws[1]:
            if cell.value is not None:
                headers[col_letter(cell.column)] = str(cell.value).strip()

        rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            row_dict = {}
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "#e":
                    row_dict[col_letter(cell.column)] = str(cell.value).strip()
            if row_dict:
                rows.append((row_dict, headers))
        wb.close()

        preview = []
        grupos_raw = {"1": []}
        for row_dict, headers in rows:
            encf = row_dict.get("D", row_dict.get("E", "?"))
            tipo = "ACECF"
            preview.append({"encf": encf, "tipo": tipo, "grupo": 1})
            grupos_raw["1"].append({"row_dict": row_dict, "headers": headers, "encf": encf})

        return {"total_cases": len(preview), "casos": preview, "_grupos_raw": grupos_raw}

    @classmethod
    def process_step3_generate(cls, company_id, company_profile, parsed_data,
                               dry_run=False, run_number=1):
        grupos = parsed_data.get("_grupos_raw", {})
        evidence_dir = _get_evidence_dir(company_id, 3, run_number)
        xml_dir = os.path.join(evidence_dir, "xml")
        _ensure_dir(xml_dir)

        token, err = cls._get_cert_token(company_profile)
        if err:
            return {"success": False, "error": f"Error de autenticacion: {err}"}

        endpoints = cls._cert_endpoints()
        acecf_url = endpoints.get("aprobacion_comercial")

        results = []
        all_cases = grupos.get("1", [])
        total = len(all_cases)
        accepted = rejected = 0

        for idx, caso in enumerate(all_cases, 1):
            encf = caso["encf"]
            row_dict = caso["row_dict"]
            headers = caso["headers"]

            case_result = {"encf": encf, "tipo": "ACECF", "success": False, "grupo": 1}

            try:
                raw_xml = DgiiTestDataLoader.build_acecf_xml_from_row(row_dict, headers)
                raw_path = os.path.join(xml_dir, f"ACECF_{encf}_raw.xml")
                with open(raw_path, "wb") as f:
                    f.write(raw_xml)
                case_result["raw_xml_path"] = raw_path

                signed_xml = DgiiSigner.sign_xml(raw_xml, company_profile)
                signed_path = os.path.join(xml_dir, f"ACECF_{encf}_signed.xml")
                with open(signed_path, "wb") as f:
                    f.write(signed_xml)
                case_result["signed_xml_path"] = signed_path

                if dry_run:
                    case_result["success"] = True
                    case_result["dry_run"] = True
                    case_result["dgii_status"] = "DRY_RUN"
                    accepted += 1
                    results.append(case_result)
                    continue

                cert_path = DgiiDirectService._prepare_tls_cert(company_profile)
                response = DgiiDirectService._multipart_post(
                    acecf_url, signed_xml, token=token, filename=f"ACECF_{encf}.xml", cert_path=cert_path
                )
                text = response.text if response else ""
                data = DgiiDirectService._safe_json(response) if response else None

                success = response is not None and response.status_code == 200
                case_result["success"] = success
                case_result["response_data"] = data or {}
                case_result["track_id"] = DgiiDirectService._extract_track_id(data, text)

                if success:
                    accepted += 1
                else:
                    rejected += 1

            except Exception as e:
                case_result["error_message"] = str(e)
                rejected += 1

            results.append(case_result)
            time.sleep(0.3)

        run_dict = {
            "run_number": run_number,
            "step": 3,
            "status": "in_progress",
            "started_at": _now(),
            "total_cases": total,
            "accepted": accepted,
            "rejected": rejected,
            "pending": 0,
            "cases": results,
            "evidencias_dir": evidence_dir,
            "dry_run": dry_run,
        }

        if rejected > 0:
            cls.fail_step(company_id, 3, run_number, run_dict)
            run_dict["status"] = "failed"
        else:
            cls.complete_step(company_id, 3, run_number, run_dict)
            run_dict["status"] = "completed"

        cls.save_run_progress(company_id, 3, run_number, run_dict)

        return {
            "success": rejected == 0,
            "total": total,
            "accepted": accepted,
            "rejected": rejected,
            "results": results,
            "run_number": run_number,
            "evidence_dir": evidence_dir,
        }

    # ═══════════════════════════════════════════════════════════════
    # Descarga de evidencias
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def download_xml(cls, xml_path):
        if not os.path.exists(xml_path):
            return None
        with open(xml_path, "rb") as f:
            return f.read()

    @classmethod
    def download_all_evidence_zip(cls, company_id, step, run_number):
        evidence_dir = _get_evidence_dir(company_id, step, run_number)
        if not os.path.exists(evidence_dir):
            return None

        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(evidence_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, evidence_dir)
                    zf.write(file_path, arcname)

        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def mark_manual_uploaded(cls, company_id, step_num, run_number, encf):
        run_path = _get_run_doc_path(company_id, step_num, run_number)
        run_data = cls._get_firestore_doc(run_path) or {}
        cases = run_data.get("cases", [])

        manual_count = 0
        for case in cases:
            if case.get("encf") == encf:
                case["manual_uploaded"] = True
            if case.get("manual_uploaded"):
                manual_count += 1

        run_data["cases"] = cases
        run_data["manual_uploaded"] = manual_count
        cls.save_run_progress(company_id, step_num, run_number, run_data)
        return {"success": True, "manual_uploaded": manual_count}

    # ═══════════════════════════════════════════════════════════════
    # Paso 4: Pruebas de Simulación e-CF (emisión real desde el sistema)
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def emit_for_certification(cls, company_profile, invoice_dict):
        token, err = cls._get_cert_token(company_profile)
        if err:
            return {"success": False, "error": f"Error de autenticacion: {err}"}

        endpoints = cls._cert_endpoints()
        recepcion_url = endpoints.get("recepcion")
        rfce_url = endpoints.get("rfce_recepcion")

        try:
            raw_xml = DgiiXmlBuilder.build_invoice_xml(company_profile, invoice_dict)
            signed_xml = DgiiSigner.sign_xml(raw_xml, company_profile)
            xml_signature = DgiiSigner.extract_signature_value(signed_xml) or hashlib.sha256(signed_xml).hexdigest()
            codigo_seguridad = xml_signature[:6]
            # Fecha/hora de firma REAL (Informe Técnico §18.2.3: FechaFirma del QR
            # debe ser dd-MM-aaaa HH:mm:ss de la firma digital, no una fecha fabricada).
            fhf = DgiiSigner.extract_fecha_hora_firma(signed_xml)
            if fhf:
                invoice_dict["fechaHoraFirma"] = fhf
            try:
                qr_url = DgiiDirectService.build_qr_url(company_profile, invoice_dict, codigo_seguridad)
            except Exception:
                qr_url = ""

            encf = invoice_dict.get("encf", "")
            company_rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
            total = float(invoice_dict.get("total", 0.0))
            ecf_type = invoice_dict.get("ecfType", "")
            is_rfce = "E32" in ecf_type and total < RFCE_THRESHOLD

            if is_rfce and rfce_url:
                rfce_xml = DgiiXmlBuilder.build_rfce_summary_xml(company_profile, invoice_dict, codigo_seguridad)
                rfce_signed = DgiiSigner.sign_xml(rfce_xml, company_profile)
                cert_path = DgiiDirectService._prepare_tls_cert(company_profile)
                # DGII valida el nombre del archivo: {RNC}{eNCF}.xml (longitud 26).
                response = DgiiDirectService._multipart_post(
                    rfce_url, rfce_signed, token=token, filename=f"{company_rnc}{encf}.xml", cert_path=cert_path
                )
            else:
                rfce_signed = None
                cert_path = DgiiDirectService._prepare_tls_cert(company_profile)
                filename = f"{company_rnc}{encf}.xml"
                response = DgiiDirectService._multipart_post(
                    recepcion_url, signed_xml, token=token, filename=filename, cert_path=cert_path
                )

            text = response.text if response else ""
            data = DgiiDirectService._safe_json(response) if response else None
            status_code = response.status_code if response else 0
            dgii_status = DgiiDirectService._extract_status(data, text)
            track_id = DgiiDirectService._extract_track_id(data, text)

            success = status_code >= 200 and status_code < 300
            rejection_error = None

            # La DGII devuelve HTTP 200 incluso cuando rechaza el contenido.
            # Verificar error/mensaje/mensajes del JSON de respuesta.
            if isinstance(data, dict):
                rejection_error = data.get("error") or data.get("mensaje") or ""
                if not rejection_error:
                    mensajes = data.get("mensajes")
                    if isinstance(mensajes, list):
                        msgs = [str(m.get("valor", "")).strip()
                                for m in mensajes if isinstance(m, dict) and m.get("valor")]
                        rejection_error = "; ".join(msgs) or None
                    elif isinstance(mensajes, str) and mensajes.strip():
                        rejection_error = mensajes.strip()
                if isinstance(rejection_error, str):
                    rejection_error = rejection_error.strip() or None

            if not rejection_error and dgii_status == "REJECTED":
                accepted = cls._check_dgii_acceptance(
                    company_profile, encf, track_id=track_id, attempts=1, delay=2
                )
                if accepted is True:
                    dgii_status = "ACCEPTED"
                elif accepted is False:
                    rejection_error = "La DGII rechazó el comprobante (consulta de estado: RECHAZADO)"
                else:
                    dgii_status = dgii_status or "PENDING"

            if rejection_error:
                success = False
                dgii_status = "REJECTED"

            return {
                "success": success,
                "track_id": track_id,
                "dgii_status": dgii_status or "UNKNOWN",
                "codigo_seguridad": codigo_seguridad,
                "xml_signature": xml_signature,
                "qrCodeURL": qr_url,
                "status_code": status_code,
                "response_data": data or {},
                "error": rejection_error,
                # Bytes EXACTOS enviados a la DGII (no re-firmar al persistir):
                # el vínculo CodigoSeguridadeCF == SignatureValue[:6] del E32
                # completo debe preservarse entre RFCE (API) y XML (portal).
                "signed_xml": signed_xml,
                "rfce_signed_xml": rfce_signed if is_rfce else None,
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


    # ═══════════════════════════════════════════════════════════════
    # Paso 4: Set de pruebas automático (25 comprobantes por bloques DGII)
    # ═══════════════════════════════════════════════════════════════

    STEP4_TIPOS = ["E31", "E32", "E33", "E34", "E41", "E43", "E44", "E45", "E46", "E47"]

    # Números a omitir por tipo al regenerar el set (evita "ya enviado" en DGII).
    STEP4_SEQUENCE_SKIP = 10

    STEP4_SET_TEMPLATE = [
        {"index": 1, "tipo": "E31", "kind": "invoice", "count": 4, "label": "Comprobantes tipo 31"},
        {"index": 2, "tipo": "E32", "kind": "invoice", "count": 2, "label": "Comprobantes tipo 32 >= 250Mil"},
        {"index": 3, "tipo": "E33", "kind": "nota", "count": 1, "label": "Comprobantes tipo 33"},
        {"index": 4, "tipo": "E34", "kind": "nota", "count": 2, "label": "Comprobantes tipo 34"},
        {"index": 5, "tipo": "E41", "kind": "expense", "count": 2, "label": "Comprobantes tipo 41"},
        {"index": 6, "tipo": "E43", "kind": "expense", "count": 2, "label": "Comprobantes tipo 43"},
        {"index": 7, "tipo": "E44", "kind": "invoice", "count": 2, "label": "Comprobantes tipo 44"},
        {"index": 8, "tipo": "E45", "kind": "invoice", "count": 2, "label": "Comprobantes tipo 45"},
        {"index": 9, "tipo": "E46", "kind": "invoice", "count": 2, "label": "Comprobantes tipo 46"},
        {"index": 10, "tipo": "E47", "kind": "supplier_invoice", "count": 2, "label": "Comprobantes tipo 47"},
        {"index": 11, "tipo": "E32", "kind": "invoice", "count": 4, "label": "Comprobantes tipo 32 RFCE", "rfce": True},
    ]

    STEP4_BASE_PRICES = {
        "E31": [45000.0, 62000.0, 78500.0, 95000.0],
        "E32_GE": [220000.0, 235000.0],
        "E32_RFCE": [12500.0, 18750.0, 24300.0, 31400.0],
        "E33": [5000.0],
        "E34": [2000.0, 3500.0],
        "E41": [15000.0, 22500.0],
        "E43": [8000.0, 12000.0],
        "E44": [20000.0, 25000.0],
        "E45": [30000.0, 40000.0],
        "E46": [50000.0, 65000.0],
        "E47": [10000.0, 15000.0],
    }

    STEP4_ECF_LABELS = {
        "E31": "Factura de Crédito Fiscal (E31)",
        "E32": "Factura de Consumo (E32)",
        "E33": "Nota de Débito (E33)",
        "E34": "Nota de Crédito (E34)",
        "E41": "Comprobante de Compras (E41)",
        "E43": "Comprobante para Gastos Menores (E43)",
        "E44": "Comprobante de Regímenes Especiales (E44)",
        "E45": "Comprobante Gubernamental (E45)",
        "E46": "Comprobante de Exportación (E46)",
        "E47": "Pagos al Exterior (E47)",
    }

    @classmethod
    def ensure_cert_sequences(cls, owner_uid, company_id, sandbox=True):
        """Garantiza que exista una secuencia ACTIVA para cada tipo e-CF del set.
        Si falta, la crea automáticamente (solo para certificación)."""
        created, existing = [], []
        sequences = DatabaseService.get_sequences(owner_uid, sandbox=sandbox, company_id=company_id) or []
        today = datetime.now(timezone.utc)
        for tipo in cls.STEP4_TIPOS:
            active = [s for s in sequences
                      if s.get("tipoComprobante") == tipo
                      and s.get("estado", "").upper() == "ACTIVA"
                      and not s.get("bloqueadaManualmente", False)]
            if active:
                existing.append(tipo)
                continue
            seq_id = str(uuid.uuid4())
            seq_dict = {
                "tipoComprobante": tipo,
                "prefijo": tipo,
                "secuenciaInicial": 1,
                "secuenciaFinal": 1000000,
                "ultimoConsecutivoUsado": 0,
                "alertaMinimoDisponible": 100,
                "fechaAutorizacion": today.strftime("%Y-%m-%d"),
                "fechaExpiracion": "2028-12-31",
                "numeroAutorizacionDgii": "CERTIFICACION-AUTO",
                "estado": "ACTIVA",
                "ambiente": "SANDBOX" if sandbox else "PRODUCCION",
                "bloqueadaManualmente": False,
            }
            DatabaseService.save_sequence(owner_uid, seq_id, seq_dict, sandbox=sandbox, company_id=company_id)
            created.append(tipo)
        return {"created": created, "existing": existing}

    @classmethod
    def skip_step4_sequences(cls, owner_uid, company_id, sandbox=True, skip=None):
        """Avanza las secuencias ACTIVAS del set para que el próximo eNCF quede
        `skip` números por encima del máximo consecutivo YA GENERADO (los logs
        de secuencia incluyen emisiones manuales y sets anteriores).
        Evita que la DGII marque 'ya enviado' comprobantes de sets nuevos o
        regenerados. Sin historial (logs vacíos) no avanza: la primera
        generación usa los números desde la posición actual."""
        skip = int(skip or cls.STEP4_SEQUENCE_SKIP)
        sequences = DatabaseService.get_sequences(owner_uid, sandbox=sandbox, company_id=company_id) or []
        logs = DatabaseService.get_sequence_logs(owner_uid, sandbox=sandbox, company_id=company_id) or []

        max_used = {}
        for log in logs:
            tipo = log.get("tipoComprobante", "")
            try:
                consecutivo = int(log.get("consecutivo", 0))
            except (TypeError, ValueError):
                consecutivo = 0
            if consecutivo > max_used.get(tipo, 0):
                max_used[tipo] = consecutivo

        skipped = {}
        for tipo in cls.STEP4_TIPOS:
            prev_max = max_used.get(tipo, 0)
            if prev_max <= 0:
                continue
            active = [s for s in sequences
                      if s.get("tipoComprobante") == tipo
                      and s.get("estado", "").upper() == "ACTIVA"
                      and not s.get("bloqueadaManualmente", False)]
            if not active:
                continue
            for seq in active:
                seq = dict(seq)
                inicial = int(seq.get("secuenciaInicial", 1))
                final = int(seq.get("secuenciaFinal", 1))
                usado = int(seq.get("ultimoConsecutivoUsado", inicial - 1))
                base = max(usado, prev_max)
                nuevo = min(final, base + skip)
                if nuevo > usado:
                    seq["ultimoConsecutivoUsado"] = nuevo
                    DatabaseService.save_sequence(owner_uid, seq.get("id"), seq,
                                                  sandbox=sandbox, company_id=company_id)
                    skipped.setdefault(tipo, {"desde": base, "hasta": nuevo, "max_usado": prev_max})
        return skipped

    @classmethod
    def _step4_price(cls, block, case_idx):
        if block.get("rfce"):
            key = "E32_RFCE"
        elif block["tipo"] == "E32":
            key = "E32_GE"
        else:
            key = block["tipo"]
        prices = cls.STEP4_BASE_PRICES.get(key, [])
        return float(prices[case_idx] if case_idx < len(prices) else prices[0])

    @classmethod
    def _step4_itbis_rate(cls, block):
        if block.get("rfce") or block["tipo"] in ("E31", "E32", "E33", "E34", "E41", "E45"):
            return 0.18
        return 0.0

    @classmethod
    def _step4_items(cls, block, case_idx, name):
        return [{
            "id": str(uuid.uuid4()),
            "code": "",
            "type": "Servicio",
            "name": name,
            "unit": "Servicio",
            "price": cls._step4_price(block, case_idx),
            "quantity": 1,
            "itbisRate": cls._step4_itbis_rate(block),
            "discountRate": 0.0,
        }]

    @classmethod
    def _step4_build_payload(cls, doc_dict, kind):
        """Adapta un documento (invoice/nota/expense/supplier_invoice) al formato
        invoice_dict que consumen DgiiXmlBuilder/EcfEmissionService.
        Réplica local de _build_expense_ecf_payload y _build_supplier_invoice_ecf_payload
        para evitar dependencias service→web."""
        if kind in ("invoice", "nota"):
            return doc_dict
        if kind == "expense":
            is_e43 = doc_dict.get("ecfType") == "E43"
            amount = float(doc_dict.get("amount", 0.0))
            itbis = 0.0 if is_e43 else float(doc_dict.get("itbisAmount", 0.0))
            subtotal = round(amount - itbis, 2)
            if subtotal < 0:
                subtotal = amount
            date_str = doc_dict.get("date", "") or datetime.now(timezone.utc).isoformat()
            due_str = doc_dict.get("dueDate") or date_str
            return {
                "id": doc_dict.get("id", ""),
                "ecfType": cls.STEP4_ECF_LABELS.get(doc_dict.get("ecfType"), "Comprobante de Compras (E41)"),
                "encf": doc_dict.get("encf", ""),
                "date": date_str,
                "dueDate": due_str,
                "clientName": doc_dict.get("providerName") or "Proveedor Genérico",
                "clientRNC": doc_dict.get("rncEmisor", ""),
                "paymentType": doc_dict.get("paymentType", "Contado"),
                "paymentMethod": "Efectivo",
                "subtotal": subtotal,
                "totalITBIS": itbis,
                "total": amount,
                "netPayable": amount,
                "retainedITBIS": float(doc_dict.get("retainedITBIS", 0.0)),
                "retainedISR": float(doc_dict.get("retainedISR", 0.0)),
                "notes": doc_dict.get("notes", ""),
                "invoiceNumber": doc_dict.get("ecfNumber") or doc_dict.get("ncf", ""),
                "items": [{
                    "id": doc_dict.get("id", "item-gasto-1"),
                    "code": "GASTO-01",
                    "name": doc_dict.get("concept", "Gasto Operativo"),
                    "type": "Servicio",
                    "quantity": 1,
                    "price": subtotal,
                    "subtotal": subtotal,
                    "itbisRate": 0.0 if is_e43 else (round(itbis / subtotal, 4) if subtotal > 0 else 0.0),
                    "total": amount,
                }],
            }
        if kind == "supplier_invoice":
            amount = float(doc_dict.get("total", 0.0))
            itbis = float(doc_dict.get("itbis", 0.0))
            subtotal = float(doc_dict.get("subtotal", 0.0))
            date_str = doc_dict.get("date", "")
            due_str = doc_dict.get("dueDate") or date_str
            items = []
            for item in doc_dict.get("items", []):
                unit_price = float(item.get("unitPrice", 0.0))
                qty = float(item.get("quantity", 0.0))
                item_subtotal = float(item.get("subtotal", 0.0))
                item_data = {
                    "name": item.get("name", "Item"),
                    "quantity": qty,
                    "price": unit_price,
                    "subtotal": item_subtotal,
                    "itbisRate": 0.0,
                    "unit": item.get("unit", "Unidad"),
                    "type": "Servicio",
                }
                item_isr_rate = float(doc_dict.get("retainedISR", 0.27))
                item_data["retainedISR"] = round(item_subtotal * item_isr_rate, 2)
                items.append(item_data)
            if not items:
                items = [{
                    "name": doc_dict.get("supplierName", "Compra"),
                    "quantity": 1,
                    "price": subtotal,
                    "subtotal": subtotal,
                    "itbisRate": 0.0,
                    "unit": "Unidad",
                    "type": "Servicio",
                    "retainedISR": round(subtotal * float(doc_dict.get("retainedISR", 0.27)), 2),
                }]
            payment_map = {"Contado": "Efectivo", "Crédito": "Crédito", "credito_30d": "Crédito"}
            ret_isr_rate = float(doc_dict.get("retainedISR", 0.0))
            ret_itbis_rate = float(doc_dict.get("retainedITBIS", 0.0))
            payload = {
                "id": doc_dict.get("id", ""),
                "ecfType": cls.STEP4_ECF_LABELS.get(doc_dict.get("ecfType"), "Pagos al Exterior (E47)"),
                "encf": doc_dict.get("encf", ""),
                "date": date_str,
                "dueDate": due_str,
                "clientRNC": doc_dict.get("supplierRnc", ""),
                "razonSocial": doc_dict.get("supplierName", "Proveedor"),
                "clientName": doc_dict.get("supplierName", "Proveedor"),
                "paymentMethod": payment_map.get(doc_dict.get("paymentType", "Contado"), "Efectivo"),
                "paymentType": doc_dict.get("paymentType", "Contado"),
                "subtotal": subtotal,
                "totalITBIS": itbis,
                "total": amount,
                "montoExento": amount,
                "retainedITBIS": round(itbis * ret_itbis_rate, 2),
                "retainedISR": round(amount * ret_isr_rate, 2),
                "items": items,
                "invoiceNumber": doc_dict.get("supplierInvoiceNumber", ""),
                "internalInvoiceNumber": doc_dict.get("invoiceNumber", ""),
                "paisDestino": doc_dict.get("paisDestino", "US"),
                "isrAsumido": doc_dict.get("isrAsumido", False),
            }
            return payload
        return doc_dict

    @classmethod
    def _load_step4_payload(cls, owner_uid, case, sandbox=True, company_id=None):
        kind = case.get("kind", "invoice")
        doc_id = case.get("doc_id", "")
        if kind in ("invoice", "nota"):
            doc = DatabaseService.get_invoice(owner_uid, doc_id, sandbox=sandbox, company_id=company_id)
        elif kind == "expense":
            doc = DatabaseService.get_expense(owner_uid, doc_id, sandbox=sandbox, company_id=company_id)
        elif kind == "supplier_invoice":
            doc = SupplierInvoiceService.get(owner_uid=owner_uid, invoice_id=doc_id, sandbox=sandbox, company_id=company_id)
        else:
            doc = None
        if not doc:
            return None
        return cls._step4_build_payload(doc, kind)

    @classmethod
    def _step4_case_pdf(cls, payload, company_profile, pdf_dir, encf):
        """Genera el PDF de evidencia usando la plantilla y el pipeline actuales."""
        try:
            from flask import render_template, request, current_app
            from app.utils.pdf import pdf_write_options
            from weasyprint import HTML as WeasyprintHTML
        except Exception:
            return None
        try:
            invoice = dict(payload or {})
            invoice["isQuotation"] = False
            invoice.setdefault("invoiceNumber", invoice.get("encf", "CERT"))
            invoice.setdefault("clientName", invoice.get("clientName", "Consumidor Final"))
            invoice.setdefault("clientRNC", "")
            invoice.setdefault("currency", "DOP")
            invoice.setdefault("exchangeRate", 1.0)
            invoice.setdefault("paymentMethod", "Efectivo")
            invoice.setdefault("paymentType", "Contado")
            invoice.setdefault("subtotal", invoice.get("total", 0.0))
            invoice.setdefault("totalITBIS", 0.0)
            invoice["status"] = "Emitida"
            invoice.setdefault("notes", "")
            invoice.setdefault("comentario", "")
            invoice.setdefault("footer", "")
            invoice.setdefault("dueDate", "")
            invoice.setdefault("emisionMode", "")
            invoice.setdefault("xmlSignature", "")
            invoice.setdefault("qrCodeURL", "")
            invoice.setdefault("discountAmount", 0.0)
            invoice.setdefault("foreignTaxId", "")
            invoice.setdefault("retainedISR", 0.0)
            invoice.setdefault("retainedITBIS", 0.0)
            invoice.setdefault("netPayable", float(invoice.get("total", 0.0)))
            invoice.setdefault("ecfType", "Factura de Consumo (E32)")
            invoice.setdefault("encf", "")
            # RI DGII (§18.2.1/18.2.3): Fecha Vencimiento = vencimiento de la
            # secuencia e-NCF (31/12/2028), excepto E32/E34 (la plantilla los oculta).
            fvs = payload.get("fechaVencimientoSecuencia") or ""
            if fvs and len(fvs) >= 10:
                invoice["dueDate"] = f"{fvs[-4:]}-{fvs[3:5]}-{fvs[:2]}"  # dd-MM-aaaa → ISO
            else:
                invoice["dueDate"] = "2028-12-31"
            fecha_firma_real = payload.get("fechaHoraFirma") or ""
            items = []
            for it in invoice.get("items", []):
                item = dict(it)
                item.setdefault("code", "")
                item.setdefault("type", it.get("type", "Servicio"))
                item.setdefault("quantity", float(it.get("quantity", 1.0)))
                item.setdefault("price", float(it.get("price", 0.0)))
                item.setdefault("subtotal", float(it.get("subtotal", 0.0)))
                item.setdefault("itbisRate", float(it.get("itbisRate", 0.0)))
                item.setdefault("discountRate", float(it.get("discountRate", 0.0)))
                item.setdefault("total", float(it.get("total", item["subtotal"])))
                item.setdefault("itbis_amount", float(it.get("itbis_amount", 0.0)))
                item.setdefault("unit", it.get("unit", "Unidad"))
                item["isc_especifico_amount"] = float(item.get("isc_especifico_amount", 0.0) or 0.0)
                item["isc_advalorem_amount"] = float(item.get("isc_advalorem_amount", 0.0) or 0.0)
                item["otros_impuestos_amount"] = float(item.get("otros_impuestos_amount", 0.0) or 0.0)
                items.append(item)
            invoice["items"] = items
            invoice["totalISCEspecifico"] = 0.0
            invoice["totalISCAdValorem"] = 0.0
            invoice["totalOtrosSelectivos"] = 0.0
            invoice["totalCDT"] = 0.0
            invoice["totalPropina"] = 0.0

            # Self-heal: QR con FechaFirma placeholder → recalcular con la real
            # antes de imprimir el PDF (la DGII valida FechaFirma en consulta).
            try:
                invoice["qrCodeURL"] = DgiiDirectService.qr_url_valido(
                    company_profile, payload
                ) or invoice.get("qrCodeURL", "")
            except Exception:
                pass

            import qrcode as _qrcode
            qr = _qrcode.QRCode(version=1, box_size=10, border=0)
            qr.add_data(invoice.get("qrCodeURL") or "https://dgii.gov.do/validaecf")
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            stream = BytesIO()
            img.save(stream, format="PNG")
            qr_base64 = base64.b64encode(stream.getvalue()).decode("utf-8")

            base_url = request.host_url if request else None
            rendered_html = render_template(
                "invoices/pdf.html",
                invoice=invoice,
                company=company_profile or {},
                branch={},
                auto_print=False,
                qr_base64=qr_base64,
                fecha_firma_str=fecha_firma_real,
                sandbox=True,
            )
            pdf_bytes = WeasyprintHTML(string=rendered_html, base_url=base_url).write_pdf(**pdf_write_options())
            rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
            pdf_name = f"{rnc}{encf}.pdf" if rnc else f"{encf}.pdf"
            pdf_path = os.path.join(pdf_dir, pdf_name)
            with open(pdf_path, "wb") as f:
                f.write(pdf_bytes)
            return pdf_path
        except Exception as e:
            print(f"⚠️ Error generando PDF para {encf}: {e}")
            return None

    @classmethod
    def _step4_case_artifacts(cls, company_profile, payload, case, xml_dir, pdf_dir):
        """Genera XML raw/firmado (+RFCE si aplica) y PDF para un caso.
        Usa los servicios de generación y firma actuales: cualquier error aquí
        es señal de que el flujo manual de facturación fallaría igual.
        Nomenclatura DGII: {rnc}{encf}.xml / {rnc}{encf}_rfce.xml / {rnc}{encf}.pdf"""
        encf = case.get("encf", "SIN-ENCF")
        rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
        prefix = f"{rnc}{encf}" if rnc else encf
        raw_xml = DgiiXmlBuilder.build_invoice_xml(company_profile, payload)
        signed_xml = DgiiSigner.sign_xml(raw_xml, company_profile)
        with open(os.path.join(xml_dir, f"{prefix}_raw.xml"), "wb") as f:
            f.write(raw_xml)
        with open(os.path.join(xml_dir, f"{prefix}.xml"), "wb") as f:
            f.write(signed_xml)
        case["xml_path"] = os.path.join(xml_dir, f"{prefix}.xml")
        payload["xmlContent"] = signed_xml.decode("utf-8", errors="replace")
        sig = DgiiSigner.extract_signature_value(signed_xml) or hashlib.sha256(signed_xml).hexdigest()
        payload["xmlSignature"] = sig
        fhf = DgiiSigner.extract_fecha_hora_firma(signed_xml)
        if fhf:
            payload["fechaHoraFirma"] = fhf
        fvs = DgiiSigner.extract_fecha_vencimiento_secuencia(signed_xml)
        if fvs:
            payload["fechaVencimientoSecuencia"] = fvs
        # Auto-reparación: si el QR guardado trae FechaFirma placeholder ('12:00:00'
        # fabricada), la DGII lo rechaza en consulta — se recalcula con la
        # FechaHoraFirma real del XML firmado.
        try:
            payload["qrCodeURL"] = DgiiDirectService.qr_url_valido(company_profile, payload) or payload.get("qrCodeURL", "")
        except Exception:
            pass
        if case.get("rfce"):
            rfce_raw = DgiiXmlBuilder.build_rfce_summary_xml(company_profile, payload, sig[:6])
            rfce_signed = DgiiSigner.sign_xml(rfce_raw, company_profile)
            with open(os.path.join(xml_dir, f"{prefix}_rfce.xml"), "wb") as f:
                f.write(rfce_signed)
            case["rfce_xml_path"] = os.path.join(xml_dir, f"{prefix}_rfce.xml")
        pdf_path = cls._step4_case_pdf(payload, company_profile, pdf_dir, encf)
        if pdf_path:
            case["pdf_path"] = pdf_path

    @classmethod
    def _step4_case_artifacts_from_emission(cls, company_profile, payload, case, result, xml_dir, pdf_dir):
        """Persiste en disco EXACTAMENTE los XML firmados que se enviaron a la DGII
        (sin re-firmar). Regla DGII: CodigoSeguridadeCF del RFCE (API) debe ser igual
        a los 6 primeros dígitos del SignatureValue del E32 completo (portal);
        re-firmar en cada paso rompería ese vínculo (FechaHoraFirma cambia)."""
        encf = case.get("encf", "SIN-ENCF")
        rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
        prefix = f"{rnc}{encf}" if rnc else encf
        signed_bytes = result.get("signed_xml")
        if isinstance(signed_bytes, bytes) and signed_bytes:
            with open(os.path.join(xml_dir, f"{prefix}.xml"), "wb") as f:
                f.write(signed_bytes)
            case["xml_path"] = os.path.join(xml_dir, f"{prefix}.xml")
            payload["xmlContent"] = signed_bytes.decode("utf-8", errors="replace")
            payload["xmlSignature"] = result.get("xml_signature") or payload.get("xmlSignature", "")
            fhf = DgiiSigner.extract_fecha_hora_firma(signed_bytes)
            if fhf:
                payload["fechaHoraFirma"] = fhf
            fvs = DgiiSigner.extract_fecha_vencimiento_secuencia(signed_bytes)
            if fvs:
                payload["fechaVencimientoSecuencia"] = fvs
        # Auto-reparación de QR con FechaFirma placeholder (emisiones previas al
        # fix): recalcular usando el xmlContent real recién guardado.
        try:
            payload["qrCodeURL"] = DgiiDirectService.qr_url_valido(company_profile, payload) or payload.get("qrCodeURL", "")
        except Exception:
            pass
        rfce_bytes = result.get("rfce_signed_xml")
        if isinstance(rfce_bytes, bytes) and rfce_bytes:
            with open(os.path.join(xml_dir, f"{prefix}_rfce.xml"), "wb") as f:
                f.write(rfce_bytes)
            case["rfce_xml_path"] = os.path.join(xml_dir, f"{prefix}_rfce.xml")
        try:
            raw_xml = DgiiXmlBuilder.build_invoice_xml(company_profile, payload)
            with open(os.path.join(xml_dir, f"{prefix}_raw.xml"), "wb") as f:
                f.write(raw_xml)
        except Exception as raw_err:
            print(f"⚠️ No se regeneró raw.xml para {encf}: {raw_err}")
        pdf_path = cls._step4_case_pdf(payload, company_profile, pdf_dir, encf)
        if pdf_path:
            case["pdf_path"] = pdf_path

    @classmethod
    def refresh_step4_qr_and_pdfs(cls, company_id, company_profile, owner_uid, run_number,
                                  sandbox=True):
        """Repara los QR con FechaFirma placeholder y regenera los PDFs de una
        corrida existente SIN re-emitir a la DGII (solo lee el XML firmado en
        disco, recalcula el qrCodeURL y lo persiste en Firestore)."""
        run_path = _get_run_doc_path(company_id, 4, run_number)
        run_data = cls._get_firestore_doc(run_path) or {}
        test_set = run_data.get("test_set") or {}

        evidence_dir = _get_evidence_dir(company_id, 4, run_number)
        xml_dir = os.path.join(evidence_dir, "xml")
        pdf_dir = os.path.join(evidence_dir, "pdf")
        _ensure_dir(pdf_dir)

        rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
        updated = 0
        pdfs = 0
        errors = []
        for block in test_set.get("blocks", []):
            for case in block.get("cases", []):
                try:
                    changed, pdf_ok = cls._repair_case_qr_and_pdf(
                        company_profile, owner_uid, case, xml_dir, pdf_dir, rnc,
                        sandbox=sandbox, company_id=company_id)
                    if changed:
                        updated += 1
                    if pdf_ok:
                        pdfs += 1
                except Exception as e:
                    encf = case.get("encf", "?")
                    errors.append(f"{encf}: {e}")

        run_data["test_set"] = test_set
        cls.save_run_progress(company_id, 4, run_number, run_data)
        return {"success": True, "qr_repaired": updated, "pdfs_regenerated": pdfs,
                "errors": errors}

    @classmethod
    def _repair_case_qr_and_pdf(cls, company_profile, owner_uid, case, xml_dir, pdf_dir, rnc,
                                sandbox=True, company_id=None):
        """Repara el QR (FechaFirma placeholder) y regenera el PDF de UN caso.
        Retorna (qr_cambiado, pdf_ok). No re-emite a la DGII."""
        encf = case.get("encf", "")
        doc_id = case.get("doc_id", "")
        if not encf or not doc_id or case.get("status") == "error":
            return False, False
        prefix = f"{rnc}{encf}" if rnc else encf
        xml_path = case.get("xml_path") or os.path.join(xml_dir, f"{prefix}.xml")
        if not os.path.exists(xml_path):
            return False, False
        payload = cls._load_step4_payload(owner_uid, case, sandbox=sandbox,
                                          company_id=company_id)
        if not payload:
            return False, False
        signed_bytes = open(xml_path, "rb").read()
        payload["xmlContent"] = signed_bytes.decode("utf-8", errors="replace")
        if not payload.get("xmlSignature"):
            sig = DgiiSigner.extract_signature_value(signed_bytes) or ""
            if sig:
                payload["xmlSignature"] = sig
        fhf = DgiiSigner.extract_fecha_hora_firma(signed_bytes)
        if fhf:
            payload["fechaHoraFirma"] = fhf
        before = payload.get("qrCodeURL", "")
        payload["qrCodeURL"] = DgiiDirectService.qr_url_valido(company_profile, payload) or before
        changed = payload.get("qrCodeURL") != before
        pdf_path = cls._step4_case_pdf(payload, company_profile, pdf_dir, encf)
        if pdf_path:
            case["pdf_path"] = pdf_path
        if changed:
            cls._save_step4_case_qr(owner_uid, case, payload,
                                    sandbox=sandbox, company_id=company_id)
        return changed, bool(pdf_path)

    @classmethod
    def refresh_step4_case(cls, company_id, company_profile, owner_uid, run_number, encf,
                           sandbox=True):
        """Repara QR y regenera el PDF de UN caso de una corrida (sin re-emitir).
        Usado por la descarga de PDF del paso 4 para auto-reparar el QR impreso."""
        run_data = cls._get_firestore_doc(_get_run_doc_path(company_id, 4, run_number)) or {}
        test_set = run_data.get("test_set") or {}
        evidence_dir = _get_evidence_dir(company_id, 4, run_number)
        xml_dir = os.path.join(evidence_dir, "xml")
        pdf_dir = os.path.join(evidence_dir, "pdf")
        _ensure_dir(pdf_dir)
        rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
        for block in test_set.get("blocks", []):
            for case in block.get("cases", []):
                if case.get("encf") != encf:
                    continue
                changed, pdf_ok = cls._repair_case_qr_and_pdf(
                    company_profile, owner_uid, case, xml_dir, pdf_dir, rnc,
                    sandbox=sandbox, company_id=company_id)
                run_data["test_set"] = test_set
                cls.save_run_progress(company_id, 4, run_number, run_data)
                return {"success": True, "qr_repaired": 1 if changed else 0,
                        "pdfs_regenerated": 1 if pdf_ok else 0, "errors": []}
        return {"success": False, "error": f"Caso {encf} no encontrado en la corrida {run_number}"}

    @classmethod
    def find_case_xml_on_disk(cls, company_id, rnc, encf):
        """Busca el XML firmado de un caso del paso 4 en los directorios de
        evidencias (última corrida primero). Usado como fuente de FechaHoraFirma
        para descargas de PDF de gastos/facturas de proveedor emitidos durante
        la certificación cuyo doc Firestore no guarda xmlContent."""
        base = f"uploads/certificacion/{company_id}/step4"
        if not os.path.isdir(base):
            return ""
        runs = sorted(os.listdir(base), reverse=True)
        for run in runs:
            xml_path = os.path.join(base, run, "xml", f"{rnc}{encf}.xml")
            if os.path.exists(xml_path):
                try:
                    return open(xml_path, encoding="utf-8", errors="ignore").read()
                except Exception:
                    return ""
        return ""

    @classmethod
    def qr_reparado_con_disco(cls, company_profile, doc, company_id):
        """Devuelve el QR válido de un documento: auto-repara el placeholder de
        FechaFirma con la FechaHoraFirma real, usando el xmlContent del doc o,
        si falta (docs de certificación), el XML firmado en disco del paso 4.
        Retorna el qrCodeURL guardado si no hay forma de repararlo."""
        xml_content = doc.get("xmlContent") or ""
        if not xml_content and doc.get("encf"):
            rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
            xml_content = cls.find_case_xml_on_disk(company_id, rnc, doc.get("encf", ""))
        if not xml_content:
            try:
                return DgiiDirectService.qr_url_valido(company_profile, doc) or doc.get("qrCodeURL", "")
            except Exception:
                return doc.get("qrCodeURL", "")
        fixed = dict(doc)
        fixed["xmlContent"] = xml_content
        try:
            return DgiiDirectService.qr_url_valido(company_profile, fixed) or doc.get("qrCodeURL", "")
        except Exception:
            return doc.get("qrCodeURL", "")

    @classmethod
    def _save_step4_case_qr(cls, owner_uid, case, payload, sandbox=True, company_id=None):
        """Persiste el qrCodeURL corregido (y el xmlContent si estaba ausente)
        en el doc Firestore del caso."""
        doc_id = case.get("doc_id", "")
        kind = case.get("kind", "invoice")
        qr = payload.get("qrCodeURL", "")
        if not doc_id or not qr:
            return
        xml_content = payload.get("xmlContent", "")
        try:
            if kind in ("invoice", "nota"):
                inv = DatabaseService.get_invoice(owner_uid, doc_id, sandbox=sandbox, company_id=company_id)
                if inv:
                    inv["qrCodeURL"] = qr
                    if xml_content and not inv.get("xmlContent"):
                        inv["xmlContent"] = xml_content
                    DatabaseService.save_invoice(owner_uid, doc_id, inv, sandbox=sandbox, company_id=company_id)
            elif kind == "expense":
                exp = DatabaseService.get_expense(owner_uid, doc_id, sandbox=sandbox, company_id=company_id)
                if exp:
                    exp["qrCodeURL"] = qr
                    if xml_content and not exp.get("xmlContent"):
                        exp["xmlContent"] = xml_content
                    DatabaseService.save_expense(owner_uid, doc_id, exp, sandbox=sandbox, company_id=company_id)
            elif kind == "supplier_invoice":
                update_fields = {"qrCodeURL": qr}
                if xml_content:
                    update_fields["xmlContent"] = xml_content
                SupplierInvoiceService.update(owner_uid, doc_id, update_fields,
                                              sandbox=sandbox, company_id=company_id)
        except Exception as e:
            print(f"⚠️ No se persistió qrCodeURL de {doc_id}: {e}")

    @classmethod
    def generate_step4_test_set(cls, company_id, company_profile, owner_uid, user_email,
                                sandbox=True, run_number=1, force_rerun=False):
        run_path = _get_run_doc_path(company_id, 4, run_number)
        run_data = cls._get_firestore_doc(run_path) or {}
        if run_data.get("test_set") and not force_rerun:
            return {"success": True, "reused": True, "run_number": run_number,
                    "set": run_data["test_set"], "errors": run_data["test_set"].get("set_errors", [])}

        deleted_docs = 0
        seq_skip_info = {}
        if force_rerun and run_data.get("test_set"):
            for block in run_data["test_set"].get("blocks", []):
                for case in block.get("cases", []):
                    doc_id = case.get("doc_id", "")
                    kind = case.get("kind", "invoice")
                    if not doc_id:
                        continue
                    try:
                        if kind in ("invoice", "nota"):
                            DatabaseService.delete_invoice(owner_uid, doc_id, sandbox=sandbox, company_id=company_id)
                        elif kind == "expense":
                            DatabaseService.delete_expense(owner_uid, doc_id, sandbox=sandbox, company_id=company_id)
                        elif kind == "supplier_invoice":
                            SupplierInvoiceService.delete(owner_uid=owner_uid, invoice_id=doc_id, sandbox=sandbox, company_id=company_id)
                        deleted_docs += 1
                    except Exception as del_err:
                        print(f"⚠️ No se pudo eliminar doc previo del set {doc_id}: {del_err}")

        seq_info = cls.ensure_cert_sequences(owner_uid, company_id, sandbox=sandbox)

        # Omitir números ya generados/enviados en TODA generación (nueva corrida
        # o regeneración): valida contra los logs de secuencia (incluye emisiones
        # manuales) y deja un hueco de STEP4_SEQUENCE_SKIP números.
        try:
            seq_skip_info = cls.skip_step4_sequences(owner_uid, company_id, sandbox=sandbox)
        except Exception as skip_err:
            print(f"⚠️ No se pudo omitir secuencias al generar el set: {skip_err}")

        evidence_dir = _get_evidence_dir(company_id, 4, run_number)
        xml_dir = os.path.join(evidence_dir, "xml")
        pdf_dir = os.path.join(evidence_dir, "pdf")
        _ensure_dir(xml_dir)
        _ensure_dir(pdf_dir)

        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        due = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")

        set_errors = []
        total = 0
        blocks = []
        for block in cls.STEP4_SET_TEMPLATE:
            # E33/E34 se emiten manualmente desde el detalle de la factura
            # (la DGII rechazaba su emisión desde el wizard). Solo marcadores.
            if block["tipo"] in ("E33", "E34"):
                cases = [{
                    "tipo": block["tipo"],
                    "kind": "nota_manual",
                    "manual": True,
                    "status": "pending",
                    "invoiceNumber": f"{block['tipo']} manual #{i + 1}",
                } for i in range(block["count"])]
                blocks.append({**block, "manual_required": True, "status": "pending", "cases": cases})
                total += len(cases)
                continue

            cases = []
            for i in range(block["count"]):
                case = {"tipo": block["tipo"], "kind": block["kind"],
                        "rfce": bool(block.get("rfce")), "status": "pending"}
                try:
                    encf, _log_id = DatabaseService.consume_next_sequence(
                        owner_uid, block["tipo"], user_email, sandbox=sandbox, company_id=company_id
                    )
                    case["encf"] = encf
                    case["date"] = today
                    label = cls.STEP4_ECF_LABELS[block["tipo"]]

                    if block["kind"] == "invoice":
                        items = cls._step4_items(block, i, f"Servicio de certificación DGII {block['tipo']} #{i + 1}")
                        calcs = DGIIService.calculate_invoice_totals(items)
                        inv_id = str(uuid.uuid4())
                        inv = {
                            "id": inv_id,
                            "invoiceNumber": f"CERT-{block['tipo']}-{i + 1:02d}",
                            "date": today,
                            "clientId": "",
                            "clientName": "CERTIFICACION DGII SRL",
                            "clientRNC": "131880681",
                            "status": "Borrador",
                            "ecfType": label,
                            "encf": encf,
                            "xmlSignature": "",
                            "qrCodeURL": "",
                            "isSyncedWithDGII": False,
                            "subtotal": calcs["subtotal"],
                            "totalITBIS": calcs["total_itbis"],
                            "montoExento": calcs["monto_exento"],
                            "total": calcs["total"],
                            "retainedISR": calcs["retained_isr"],
                            "retainedITBIS": calcs["retained_itbis"],
                            "netPayable": calcs["net_payable"],
                            "isQuotation": False,
                            "notes": "Set certificación DGII paso 4",
                            "currency": "DOP",
                            "exchangeRate": 1.0,
                            "paymentType": "Contado",
                            "paymentMethod": "Efectivo",
                            "incomeType": "01 - Ingresos por operaciones",
                            "items": calcs["items"],
                            "totalPaid": 0.0,
                            "remainingBalance": calcs["net_payable"],
                            "paymentAgreement": {"enabled": False},
                            "branchId": "default-sucursal-principal",
                            "createdAt": datetime.now(timezone.utc).isoformat(),
                        }
                        if block["tipo"] == "E46":
                            inv["identificadorExtranjero"] = "US123456789"
                            inv["foreignTaxId"] = "US123456789"
                            inv["clientCountry"] = "US"
                            inv["clientRNC"] = ""
                        inv["dueDate"] = due
                        DatabaseService.save_invoice(owner_uid, inv_id, inv, sandbox=sandbox, company_id=company_id)
                        case["doc_id"] = inv_id
                        case["total"] = round(calcs["total"], 2)
                        case["invoiceNumber"] = inv["invoiceNumber"]

                    elif block["kind"] == "expense":
                        base = cls._step4_price(block, i)
                        is_e43 = block["tipo"] == "E43"
                        itbis = 0.0 if is_e43 else round(base * 0.18, 2)
                        amount = round(base + itbis, 2)
                        exp_id = str(uuid.uuid4())
                        exp_dict = {
                            "id": exp_id,
                            "concept": f"Compra certificación DGII {block['tipo']} #{i + 1}",
                            "category": "Operativos",
                            "amount": amount,
                            "itbisAmount": itbis,
                            "date": today,
                            "dueDate": due,
                            "providerName": "PROVEEDOR FORMAL CERT SRL" if not is_e43 else "PROVEEDOR INFORMAL CERT",
                            "rncEmisor": "" if is_e43 else "131880681",
                            "paymentType": "Contado",
                            "paymentMethod": "Efectivo",
                            "status": "Pendiente",
                            "ecfType": block["tipo"],
                            "isMinorExpense": is_e43,
                            "encf": encf,
                            "ecfNumber": encf,
                            "ncf": encf,
                            "isSyncedWithDGII": False,
                            "retainedISR": 0.0,
                            "retainedITBIS": 0.0,
                            "notes": "Set certificación DGII paso 4",
                            "branchId": "default-sucursal-principal",
                            "createdAt": datetime.now(timezone.utc).isoformat(),
                        }
                        DatabaseService.save_expense(owner_uid, exp_id, exp_dict, sandbox=sandbox, company_id=company_id)
                        case["doc_id"] = exp_id
                        case["total"] = amount
                        case["invoiceNumber"] = f"EXP-{block['tipo']}-{i + 1:02d}"

                    elif block["kind"] == "supplier_invoice":
                        base = cls._step4_price(block, i)
                        sinv = {
                            "invoiceNumber": f"FI-CERT-E47-{i + 1:02d}",
                            "supplierName": "FOREIGN SERVICES INC",
                            "supplierRnc": "350555123",
                            "supplierType": "formal",
                            "ecfType": "E47",
                            "encf": encf,
                            "date": today,
                            "dueDate": due,
                            "subtotal": base,
                            "itbis": 0.0,
                            "total": base,
                            "currency": "DOP",
                            "exchangeRate": 1.0,
                            "paymentTerms": "contado",
                            "paymentType": "Contado",
                            "status": "registrada",
                            "cxpStatus": "Pendiente",
                            "retainedISR": 0.27,
                            "retainedITBIS": 0.0,
                            "items": [{
                                "name": f"Servicio exterior certificación E47 #{i + 1}",
                                "unitPrice": base,
                                "quantity": 1,
                                "subtotal": base,
                                "itbisRate": 0.0,
                            }],
                            "notes": "Set certificación DGII paso 4 — E47",
                            "branchId": "default-sucursal-principal",
                        }
                        SupplierInvoiceService.create(owner_uid=owner_uid, data=sinv, sandbox=sandbox, company_id=company_id)
                        case["doc_id"] = sinv["id"]
                        case["total"] = base
                        case["invoiceNumber"] = sinv["invoiceNumber"]
                except Exception as e:
                    case["status"] = "error"
                    case["error_message"] = str(e)
                    set_errors.append(f"{block['tipo']} #{i + 1}: {e}")
                cases.append(case)
                total += 1
            blocks.append({**block, "status": "pending", "cases": cases})

        for block in blocks:
            for case in block.get("cases", []):
                if case.get("status") == "error" or not case.get("doc_id"):
                    continue
                try:
                    payload = cls._load_step4_payload(owner_uid, case, sandbox=sandbox, company_id=company_id)
                    if not payload:
                        raise ValueError("Documento no encontrado en Firestore")
                    cls._step4_case_artifacts(company_profile, payload, case, xml_dir, pdf_dir)
                    case["validation"] = "ok"
                except Exception as e:
                    case["validation"] = f"ERROR: {e}"
                    set_errors.append(f"{case.get('encf', '?')}: {e}")

        set_warnings = []
        if seq_info.get("created"):
            set_warnings.append(
                f"Secuencias auto-creadas para {', '.join(seq_info['created'])} (rango 1–1,000,000). "
                "Verifica que la DGII haya autorizado estos rangos para la certificación: "
                "eNCF fuera de rango autorizado son rechazados."
            )
        if seq_skip_info:
            det = "; ".join(f"{t}: {v['desde']}→{v['hasta']} (usado máx {v['max_usado']})"
                            for t, v in sorted(seq_skip_info.items()))
            set_warnings.append(
                f"Se omitieron {cls.STEP4_SEQUENCE_SKIP} números por tipo para evitar colisiones "
                f"con e-CF ya generados/enviados ({det})."
            )

        test_set = {
            "created_at": _now(),
            "total": total,
            "blocks": blocks,
            "sequence_info": seq_info,
            "sequence_skip": seq_skip_info,
            "set_errors": set_errors,
            "warnings": set_warnings,
        }
        run_data["test_set"] = test_set
        run_data["total_cases"] = total
        run_data["run_number"] = run_number
        run_data["status"] = "in_progress"
        cls.save_run_progress(company_id, 4, run_number, run_data)

        return {
            "success": True,
            "reused": False,
            "run_number": run_number,
            "set": test_set,
            "errors": set_errors,
            "deleted_docs": deleted_docs,
        }

    @classmethod
    def get_step4_set(cls, company_id, run_number=None):
        if not run_number:
            process = cls.get_process(company_id)
            run_number = process.get("steps", {}).get("4", {}).get("current_run", 0)
        if not run_number:
            return None
        run_data = cls._get_firestore_doc(_get_run_doc_path(company_id, 4, run_number)) or {}
        return run_data.get("test_set")

    @classmethod
    def mark_step4_block_sent(cls, company_id, run_number, block_index, marked_by=""):
        """Marca un bloque como enviado manualmente (el usuario ya envió esos
        comprobantes a la DGII desde sus módulos). Permite proseguir al siguiente
        bloque/paso sin re-enviar desde el wizard."""
        run_path = _get_run_doc_path(company_id, 4, run_number)
        run_data = cls._get_firestore_doc(run_path) or {}
        test_set = run_data.get("test_set")
        if not test_set:
            return {"success": False, "error": "No existe un set de pruebas para esta corrida. Genéralo primero."}

        blocks = test_set.get("blocks", [])
        if block_index < 1 or block_index > len(blocks):
            return {"success": False, "error": "Bloque inválido."}

        block = blocks[block_index - 1]
        if block.get("status") == "sent":
            return {"success": True, "reused": True, "block": block, "all_blocks_sent": all(b.get("status") == "sent" for b in blocks)}

        for prev in blocks[:block_index - 1]:
            if prev.get("status") != "sent":
                return {"success": False,
                        "error": f"El bloque {prev['index']} ({prev['label']}) debe enviarse primero."}

        for case in block.get("cases", []):
            if case.get("status") not in ("accepted", "rejected", "error"):
                case["status"] = "manual_sent"
        block["status"] = "sent"
        block["manual_sent"] = True
        block["marked_by"] = marked_by
        block["sent_count"] = len([c for c in block.get("cases", []) if c.get("status") in ("accepted", "manual_sent")])
        block["failed_count"] = len([c for c in block.get("cases", []) if c.get("status") in ("rejected", "error")])

        test_set["blocks"] = blocks
        run_data["test_set"] = test_set
        run_data["status"] = "in_progress"
        cls.save_run_progress(company_id, 4, run_number, run_data)

        all_sent = all(b.get("status") == "sent" for b in blocks)
        if all_sent:
            cls.complete_step(company_id, 4, run_number, run_data)

        return {
            "success": True,
            "block": block,
            "run_number": run_number,
            "all_blocks_sent": all_sent,
        }

    @classmethod
    def _mark_step4_case_emitted(cls, owner_uid, case, payload, result, sandbox=True, company_id=None):
        xml_signature = result.get("xml_signature", "")
        track_id = result.get("track_id", "")
        updates = {
            "encf": payload.get("encf", case.get("encf", "")),
            "xmlSignature": xml_signature,
            "qrCodeURL": payload.get("qrCodeURL", ""),
            "trackId": track_id,
            "isSyncedWithDGII": True,
            "dgiiStatus": result.get("dgii_status") or "ACCEPTED",
            "emisionMode": "API",
        }
        kind = case.get("kind", "invoice")
        doc_id = case.get("doc_id", "")
        if kind in ("invoice", "nota"):
            inv = DatabaseService.get_invoice(owner_uid, doc_id, sandbox=sandbox, company_id=company_id)
            if inv:
                inv.update(updates)
                inv["status"] = "Emitida"
                try:
                    inv["xmlContent"] = open(case.get("xml_path", ""), "rb").read().decode("utf-8") if case.get("xml_path") and os.path.exists(case.get("xml_path", "")) else ""
                except Exception:
                    inv["xmlContent"] = ""
                DatabaseService.save_invoice(owner_uid, doc_id, inv, sandbox=sandbox, company_id=company_id)
        elif kind == "expense":
            exp = DatabaseService.get_expense(owner_uid, doc_id, sandbox=sandbox, company_id=company_id)
            if exp:
                exp.update(updates)
                exp["ecfNumber"] = payload.get("encf", exp.get("encf", ""))
                exp["ncf"] = payload.get("encf", exp.get("encf", ""))
                DatabaseService.save_expense(owner_uid, doc_id, exp, sandbox=sandbox, company_id=company_id)
        elif kind == "supplier_invoice":
            sinv_updates = dict(updates)
            sinv_updates.update({
                "ecfNumber": payload.get("encf", case.get("encf", "")),
                "ncf": payload.get("encf", case.get("encf", "")),
                "status": "emitida",
            })
            SupplierInvoiceService.update(owner_uid, doc_id, sinv_updates, sandbox=sandbox, company_id=company_id)

    @classmethod
    def _classify_consulta_status(cls, data, text):
        """Clasifica la respuesta de consulta de estado de CerteCF.
        Distingue ACCEPTED / REJECTED / NOT_FOUND (no confundir 'no aparece' con rechazo)."""
        raw = (text or "") + " " + json.dumps(data or {}, default=str)
        r = raw.upper()
        if any(t in r for t in ["RECHAZADO", "REJECTED", "ANULADO", "CANCELADO"]):
            return "REJECTED"
        if any(t in r for t in ["ACEPTADO", "APROBADO", "ACCEPTED"]):
            return "ACCEPTED"
        if any(t in r for t in ["NO EXISTE", "NO ENCONTRADO", "NO ENCONTRADA",
                                "NO SE ENCONTRO", "NO SE ENCONTRÓ", "INEXISTENTE",
                                "SIN RESULTADOS", "NO HAY RESULTADOS"]):
            return "NOT_FOUND"
        return None

    @classmethod
    def _check_dgii_acceptance(cls, company_profile, encf, track_id=None, attempts=4, delay=12):
        """Consulta el estado del e-CF en CerteCF. Retorna True (ACCEPTED),
        False (REJECTED) o None (pendiente/desconocido/no aparece).
        Prefiere consultaresultado por trackid (más confiable); fallback a
        consultaestado por eNCF con parámetro 'encf' (no 'ncfelectronico')."""
        endpoints = cls._cert_endpoints()
        company_rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
        urls = []
        if track_id:
            url = endpoints.get("consulta_resultado")
            if url:
                urls.append((url, {"rncemisor": company_rnc, "trackid": track_id}))
        url_estado = endpoints.get("consulta_estado")
        if url_estado:
            urls.append((url_estado, {"rncemisor": company_rnc, "encf": encf}))
        if not urls:
            return None
        token, err = cls._get_cert_token(company_profile)
        if err:
            return None
        cert_path = DgiiDirectService._prepare_tls_cert(company_profile)
        try:
            for attempt in range(attempts):
                for url, params in urls:
                    try:
                        response = DgiiDirectService._get_with_params(
                            url, params, token=token, cert_path=cert_path
                        )
                        text = response.text if response is not None else ""
                        data = DgiiDirectService._safe_json(response) if response is not None else None
                        estado = cls._classify_consulta_status(data, text)
                        if estado is None:
                            # Solo aceptamos ACCEPTED del extractor genérico; un "ERROR"
                            # genérico no debe tratarse como rechazo (evita falsos REJECTED).
                            estado = DgiiDirectService._extract_status(data, text)
                            if estado != "ACCEPTED":
                                estado = None
                        if estado == "ACCEPTED":
                            return True
                        if estado == "REJECTED":
                            return False
                    except Exception:
                        pass
                if attempt < attempts - 1:
                    time.sleep(delay)
            return None
        finally:
            DgiiDirectService._cleanup_tls_cert(cert_path)

    @classmethod
    def _gate_nota_references(cls, company_profile, blocks, block_index):
        """Regla DGII: una nota (E33/E34) no puede enviarse si el comprobante
        modificado fue RECHAZADO. Solo bloquea ante rechazo explícito confirmado;
        pendiente/desconocido no bloquea (la DGII rechazará explícitamente y el
        mensaje quedará visible en el caso)."""
        block = blocks[block_index - 1]
        ref_encfs = {c.get("ncfModified") for c in block.get("cases", []) if c.get("ncfModified")}
        if not ref_encfs:
            return None
        for prev in blocks[:block_index - 1]:
            for pc in prev.get("cases", []):
                if pc.get("encf") in ref_encfs:
                    if pc.get("status") != "accepted":
                        continue
                    if pc.get("track_id"):
                        accepted = cls._check_dgii_acceptance(
                            company_profile, pc.get("encf"), track_id=pc.get("track_id")
                        )
                        if accepted is False:
                            return (f"El comprobante referenciado {pc.get('encf')} fue RECHAZADO por la DGII. "
                                    "Corrige y reenvía antes de enviar la nota.")
        return None

    @classmethod
    def send_step4_block(cls, company_id, company_profile, owner_uid, user_email,
                         sandbox=True, run_number=1, block_index=1, resend=False):
        run_path = _get_run_doc_path(company_id, 4, run_number)
        run_data = cls._get_firestore_doc(run_path) or {}
        test_set = run_data.get("test_set")
        if not test_set:
            return {"success": False, "error": "No existe un set de pruebas para esta corrida. Genéralo primero."}

        blocks = test_set.get("blocks", [])
        if block_index < 1 or block_index > len(blocks):
            return {"success": False, "error": "Bloque inválido."}

        block = blocks[block_index - 1]
        if block.get("manual_required"):
            return {"success": False,
                    "error": (f"Los comprobantes {block.get('tipo')} se emiten manualmente desde el detalle de "
                              "la factura (generar nota de débito/crédito). Luego usa 'Marcar enviado (manual)' "
                              "en este bloque para continuar.")}
        if block.get("status") == "sent" and not resend:
            return {"success": True, "reused": True, "block": block, "sent": block.get("sent_count", 0), "failed": 0}

        for prev in blocks[:block_index - 1]:
            if prev.get("status") != "sent":
                return {"success": False,
                        "error": f"El bloque {prev['index']} ({prev['label']}) debe enviarse primero."}

        if resend:
            # Reenvío intencional con los MISMOS eNCF (la DGII los rechazará por
            # duplicado y reiniciará esa prueba específica).
            for case in block.get("cases", []):
                if case.get("doc_id") and case.get("status") != "error":
                    case["status"] = "pending"
                    case.pop("track_id", None)
                    case.pop("dgii_status", None)
                    case.pop("dgii_message", None)
                    case.pop("error_message", None)
            block["status"] = "pending"
            block["resend_count"] = block.get("resend_count", 0) + 1

        if not resend and block.get("tipo") in ("E33", "E34"):
            gate_err = cls._gate_nota_references(company_profile, blocks, block_index)
            if gate_err:
                return {"success": False, "error": gate_err, "block": block}

        xml_dir = os.path.join(_get_evidence_dir(company_id, 4, run_number), "xml")
        _ensure_dir(xml_dir)
        pdf_dir = os.path.join(_get_evidence_dir(company_id, 4, run_number), "pdf")
        _ensure_dir(pdf_dir)

        sent = failed = 0
        stop = False
        for case in block.get("cases", []):
            if case.get("status") == "accepted":
                sent += 1
                continue
            if case.get("status") == "error" or not case.get("doc_id"):
                case["status"] = "rejected"
                failed += 1
                continue
            payload = cls._load_step4_payload(owner_uid, case, sandbox=sandbox, company_id=company_id)
            if not payload:
                case["status"] = "rejected"
                case["error_message"] = "Documento no encontrado en Firestore"
                failed += 1
                continue
            result = cls.emit_for_certification(company_profile, payload)
            case["track_id"] = result.get("track_id")
            case["dgii_status"] = result.get("dgii_status")
            rd = result.get("response_data") or {}
            if isinstance(rd, dict):
                dgii_msg = str(rd.get("message") or rd.get("mensajes") or rd.get("estado") or "").strip()
                if dgii_msg:
                    case["dgii_message"] = dgii_msg[:300]
                case["response_data"] = {k: (str(v)[:300] if isinstance(v, str) else v)
                                         for k, v in list(rd.items())[:8]}
            elif isinstance(rd, str) and rd:
                case["dgii_message"] = rd[:300]
                case["response_data"] = rd[:300]
            else:
                case["response_data"] = {}
            if result.get("success"):
                case["status"] = "accepted"
                sent += 1
                try:
                    if result.get("qrCodeURL"):
                        payload["qrCodeURL"] = result["qrCodeURL"]
                    cls._step4_case_artifacts_from_emission(company_profile, payload, case, result, xml_dir, pdf_dir)
                    cls._mark_step4_case_emitted(owner_uid, case, payload, result,
                                                 sandbox=sandbox, company_id=company_id)
                except Exception as art_err:
                    case["artifact_error"] = str(art_err)
            else:
                case["status"] = "rejected"
                case["error_message"] = case.get("dgii_message") or result.get("error") or "Rechazado por DGII"
                failed += 1
                stop = True
                break

        if stop:
            block["status"] = "failed"
        elif failed > 0:
            block["status"] = "failed"
        else:
            block["status"] = "sent"
        block["sent_count"] = sent
        block["failed_count"] = failed

        test_set["blocks"] = blocks
        run_data["test_set"] = test_set
        run_data["accepted"] = run_data.get("accepted", 0) + sent
        run_data["rejected"] = run_data.get("rejected", 0) + failed
        run_data["status"] = "in_progress"
        cls.save_run_progress(company_id, 4, run_number, run_data)

        all_sent = all(b.get("status") == "sent" for b in blocks)
        if all_sent and failed == 0:
            cls.complete_step(company_id, 4, run_number, run_data)
        elif failed > 0:
            cls.fail_step(company_id, 4, run_number, run_data)

        return {
            "success": failed == 0 and not stop,
            "block": block,
            "sent": sent,
            "failed": failed,
            "run_number": run_number,
            "all_blocks_sent": all_sent and failed == 0,
        }

    # ═══════════════════════════════════════════════════════════════
    # Paso 1: Firma de XML de postulacion
    # ═══════════════════════════════════════════════════════════════

    @classmethod
    def sign_postulacion_xml(cls, company_profile, raw_xml_bytes):
        signed_xml = DgiiSigner.sign_xml(raw_xml_bytes, company_profile)
        rnc = company_profile.get("companyRNC", "").replace("-", "")
        filename = f"{rnc}_postulacion_firmada.xml" if rnc else "postulacion_firmada.xml"
        return signed_xml, filename
