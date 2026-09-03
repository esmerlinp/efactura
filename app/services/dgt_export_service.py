"""DGTExportService — Exportación de formularios DGT en .txt (SIRLA fixed-width), .xlsx y .pdf.

Formato principal: .txt fixed-width SIRLA con registros E (Encabezado), D (Detalle), S (Sumario).
Formatos secundarios: .xlsx (revisión) y .pdf (visualización/impresión).
"""

import csv
import io
from datetime import datetime


def _format_date(d: str) -> str:
    """Retorna la fecha tal cual (ya debe venir en DD/MM/AAAA)."""
    return d


def _ljust(value, length):
    """Alinea a la izquierda y rellena con espacios a la derecha."""
    s = str(value) if value is not None else ""
    return s[:length].ljust(length)

def _rjust_zero(value, length):
    """Alinea a la derecha y rellena con ceros a la izquierda."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return "0".rjust(length, "0")
    try:
        s = str(int(value))
    except (ValueError, TypeError):
        s = "0"
    return s[-length:].rjust(length, "0")


def _decimal(value, length, decimals=2):
    """Formatea un número con punto decimal fijo, alineado a la derecha con ceros.

    Ejemplos (SIRLA DGT-2): _decimal(100.0, 8) -> "00100.00";
    _decimal(35.0, 6) -> "035.00"; _decimal(0.0, 5) -> "00.00".
    """
    try:
        v = float(value)
    except (TypeError, ValueError):
        v = 0.0
    return f"{v:0{length}.{decimals}f}"


def _sanitize_cause(cause: str) -> str:
    """Causa de prolongación válida según art. 153 Código de Trabajo: a-e."""
    c = (cause or "").strip().lower()
    return c if c in ("a", "b", "c", "d", "e") else ""


def _build_days_block(overtime_days: dict) -> str:
    """Construye los 31 bloques diarios (5 chars horas + 6 chars porcentaje).

    Cada día ocupa 11 caracteres (5 + 6), 31 días -> 341 caracteres.
    Los días sin horas extras se llenan con "00.00" y "000.00".
    """
    block = ""
    for day in range(1, 32):
        data = overtime_days.get(day) or overtime_days.get(str(day)) or {}
        hours = float(data.get("hours", 0) or 0)
        pct = float(data.get("percentage", 0) or 0)
        block += _decimal(hours, 5, 2) + _decimal(pct, 6, 2)
    return block


class DGTExportService:

    @staticmethod
    def to_txt(lines: list[dict], company_info: dict = None, year: int = None,
               month: int = None) -> str:
        """Genera archivo fixed-width para SIRLA con registros E (Header), D (Detalle), S (Sumario).

        Especificación oficial SIRLA DGT-3 (carga de trabajadores):
          E: ET3<RNC_11><MMYYYY>                                                          → 20 chars
          D: D<NI_3><doc_type><doc_25><names_50><ap1_40><ap2_40><birth_8><sex_1><salary_16>
             <hire_8><occup_6><cargo_150><vac_ini_8><vac_fin_8><turno_6><loc_6>
             <reserved_11><educ_5><disc_50>                                              → 443 chars
          S: S<total_6>                                                                    → 7 chars

        Posiciones por campo (1-based):
          1        Tipo de registro "D"
          2-4      Tipo de novedad "NI" (padded a 3 con espacio)
          5        Tipo de documento (C/P/N/M/I)
          6-30     Número de documento (25, AN left)
          31-80    Nombres (50, AN left)
          81-120   Primer apellido (40, AN left)
          121-160  Segundo apellido (40, AN left)
          161-168  Fecha nacimiento (8, DDMMYYYY)
          169      Sexo (1)
          170-185  Salario (16, N right-zero)
          186-193  Fecha de ingreso (8, DDMMYYYY)
          194-199  Ocupación (6, N right-zero)
          200-349  Cargo (150, AN left)
          350-357  Inicio vacaciones (8, DDMMYYYY)
          358-365  Fin vacaciones (8, DDMMYYYY)
          366-371  Turno (6, N right-zero)
          372-377  Localidad (6, N right-zero)
          378-388  Reservado (11 espacios)
          389-393  Nivel educación (5, N right-zero)
          394-443  Discapacidad (50, AN left)
        """
        ci = company_info or {}
        rnc = (ci.get("companyRNC", "") or "").replace("-", "").replace(" ", "")[:11]
        rnl = (ci.get("rnlNumber", "") or "").replace("-", "").replace(" ", "")
        localidad = rnl[-4:].rjust(6, "0") if rnl else "000000"

        if year and month:
            periodo = f"{int(month):02d}{int(year):04d}"
        else:
            periodo = datetime.now().strftime("%m%Y")

        output = io.StringIO()

        # Registro E — Encabezado (20 chars)
        header = f"ET3{rnc.rjust(11, '0')}{periodo}"
        output.write(header.ljust(20) + "\n")

        # Registros D — Detalle
        num_registros = 2  # E + S (se suma un D por cada empleado)
        for emp in lines:
            doc_type = emp.get("docTypeSirla", "C")
            doc_num = (emp.get("documento", "") or "")[:25]
            nombres = emp.get("primerNombre", "")[:50]
            ap1 = emp.get("primerApellido", "")[:40]
            ap2 = emp.get("segundoApellido", "")[:40]
            fnac = emp.get("fechaNacimientoSirla", "")[:8]
            sexo = emp.get("sexo", "")[:1] or " "
            salario = int(float(emp.get("salario", 0)))
            fing = emp.get("fechaIngresoSirla", "")[:8]
            oc_cod = (emp.get("ocupacionCodigo", "") or "").replace(" ", "")[:6]
            cargo = (emp.get("cargo", "") or "")[:150]
            vac_ini = emp.get("inicioVacaciones", "")[:8]
            vac_fin = emp.get("finVacaciones", "")[:8]
            turno = emp.get("turnoSirla", 1) or 1
            educ = emp.get("gradoInstruccion", 0) or 0
            disc = (emp.get("discapacidad", "") or "")[:50]

            linea = (
                f"D"
                f"{_ljust('NI', 3)}"
                f"{doc_type}"
                f"{_ljust(doc_num, 25)}"
                f"{_ljust(nombres, 50)}"
                f"{_ljust(ap1, 40)}"
                f"{_ljust(ap2, 40)}"
                f"{_ljust(fnac, 8)}"
                f"{sexo}"
                f"{_rjust_zero(salario, 16)}"
                f"{_ljust(fing, 8)}"
                f"{_rjust_zero(oc_cod, 6)}"
                f"{_ljust(cargo, 150)}"
                f"{_ljust(vac_ini, 8)}"
                f"{_ljust(vac_fin, 8)}"
                f"{_rjust_zero(turno, 6)}"
                f"{localidad}"
                f"{' ' * 11}"
                f"{_rjust_zero(educ, 5)}"
                f"{_ljust(disc, 50)}"
            )
            output.write(linea + "\n")
            num_registros += 1

        # Registro S — Sumario
        output.write(f"S{num_registros:06d}\n")

        return output.getvalue()

    @staticmethod
    def to_sirla_txt_dgt2(lines: list[dict], company_info: dict = None,
                          establishment_id: str = "000000", year: int = None,
                          month: int = None) -> str:
        """Genera archivo fixed-width SIRLA DGT-2 (Cartel de Horas y Vacaciones).

        Especificación oficial DGT-2 (SIRLA — carga de trabajadores):
          E: ET2<RNC_11><MMYYYY>                                               → 20 chars
          D: D<NC_3><doc_type><doc_25><estab_6><valor_hora_8><31d×11><causa_15> → 400 chars
          S: S<total_6>                                                        → 7 chars

        Posiciones por campo (1-based):
          1        Tipo de registro "D"
          2-4      Tipo de novedad "NC" (padded a 3 con espacio)
          5        Tipo de documento (C/P/N/M/I)
          6-30     Número de documento (25, AN left)
          31-36    ID establecimiento (6, N right-zero)
          37-44    Valor de la hora normal (8, con punto decimal: "00100.00")
          45-385   Días 1..31: 5 chars horas + 6 chars porcentaje (341)
          386-400  Causa de prolongación (15, AN left)
        """
        ci = company_info or {}
        rnc = (ci.get("companyRNC", "") or "").replace("-", "").replace(" ", "")[:11]
        if year and month:
            periodo = f"{int(month):02d}{int(year):04d}"
        else:
            periodo = datetime.now().strftime("%m%Y")

        output = io.StringIO()

        # Registro E — Encabezado (20 chars)
        output.write(f"ET2{rnc.rjust(11, '0')}{periodo}".ljust(20) + "\n")

        num_registros = 2  # E + S
        for emp in lines:
            doc_type = emp.get("docTypeSirla", "C")
            doc_num = (emp.get("documento", "") or "")[:25]
            hourly_rate = float(emp.get("hourlyRate", 0) or 0)
            cause = _sanitize_cause(emp.get("overtimeCause", ""))

            # 31 días de horas extras (5h + 6%) desde el detalle por día
            dias_he = _build_days_block(emp.get("overtimeDays", {}) or {})

            linea = (
                f"D"
                f"{_ljust('NC', 3)}"
                f"{doc_type}"
                f"{_ljust(doc_num, 25)}"
                f"{_rjust_zero(establishment_id, 6)}"
                f"{_decimal(hourly_rate, 8, 2)}"
                f"{dias_he}"
                f"{_ljust(cause, 15)}"
            )
            output.write(linea + "\n")
            num_registros += 1

        # Registro S — Sumario (7 chars)
        output.write(f"S{num_registros:06d}\n")

        return output.getvalue()

    @staticmethod
    def to_sirla_txt_dgt4(lines: list[dict], company_info: dict = None,
                          year: int = None, month: int = None) -> str:
        """Genera archivo fixed-width SIRLA DGT-4 (Cambios en Personal Fijo).

        Especificación oficial SIRLA DGT-4 (carga de trabajadores):
          E: ET4<RNC_11><MMYYYY>                                                          → 20 chars
          D: D<NI|NS|NC_3><doc_type><doc_25><nombres_50><ap1_40><ap2_40><fnac_8><sexo_1>
             <salario_16><fing_8><fsal_8><ocup_cod_6><ocup_desc_150><vac_ini_8><vac_fin_8>
             <turno_6><localidad_6><nacion_3><fcambio_8><educ_5><disc_50>                → 451 chars
          S: S<total_6>                                                                    → 7 chars

        Posiciones por campo (1-based):
          1        Tipo de registro "D"
          2-4      Tipo de novedad NI/NS/NC (padded a 3 con espacio)
          5        Tipo de documento (C/P/N/M/I)
          6-30     Número de documento (25, AN left)
          31-80    Nombres (50, AN left)
          81-120   Primer apellido (40, AN left)
          121-160  Segundo apellido (40, AN left)
          161-168  Fecha nacimiento (8, DDMMYYYY)
          169      Sexo (1)
          170-185  Salario (16, N right-zero)
          186-193  Fecha de ingreso (8, DDMMYYYY)
          194-201  Fecha de salida (8, DDMMYYYY)
          202-207  Ocupación (6, N right-zero)
          208-357  Descripción ocupación (150, AN left)
          358-365  Inicio vacaciones (8, DDMMYYYY)
          366-373  Fin vacaciones (8, DDMMYYYY)
          374-379  Turno (6, N right-zero)
          380-385  Localidad (6, N right-zero)
          386-388  Nacionalidad (3, AN left, solo extranjeros)
          389-396  Fecha del cambio (8, DDMMYYYY)
          397-401  Nivel educación (5, N right-zero)
          402-451  Discapacidad (50, AN left)
        """
        ci = company_info or {}
        rnc = (ci.get("companyRNC", "") or "").replace("-", "").replace(" ", "")[:11]
        rnl = (ci.get("rnlNumber", "") or "").replace("-", "").replace(" ", "")
        localidad = rnl[-4:].rjust(6, "0") if rnl else "000000"
        if year and month:
            periodo = f"{int(month):02d}{int(year):04d}"
        else:
            periodo = datetime.now().strftime("%m%Y")

        output = io.StringIO()

        output.write(f"ET4{rnc.rjust(11, '0')}{periodo}".ljust(20) + "\n")

        num_registros = 2
        for emp in lines:
            linea = (
                f"D"
                f"{_ljust(emp.get('novedadSirla', 'NI') or 'NI', 3)}"
                f"{emp.get('docTypeSirla', 'C')}"
                f"{_ljust((emp.get('documento', '') or '')[:25], 25)}"
                f"{_ljust(emp.get('primerNombre', '')[:50], 50)}"
                f"{_ljust(emp.get('primerApellido', '')[:40], 40)}"
                f"{_ljust(emp.get('segundoApellido', '')[:40], 40)}"
                f"{_ljust(emp.get('fechaNacimientoSirla', '')[:8], 8)}"
                f"{(emp.get('sexo', '') or ' ')[:1]}"
                f"{_rjust_zero(int(float(emp.get('salario', 0))), 16)}"
                f"{_ljust(emp.get('fechaIngresoSirla', '')[:8], 8)}"
                f"{_ljust(emp.get('fechaSalidaSirla', '')[:8], 8)}"
                f"{_rjust_zero((emp.get('ocupacionCodigo', '') or '').replace(' ', '')[:6], 6)}"
                f"{_ljust((emp.get('cargo', '') or '')[:150], 150)}"
                f"{_ljust(emp.get('inicioVacaciones', '')[:8], 8)}"
                f"{_ljust(emp.get('finVacaciones', '')[:8], 8)}"
                f"{_rjust_zero(emp.get('turnoSirla', 1) or 1, 6)}"
                f"{localidad}"
                f"{_ljust(emp.get('nacionalidadSirla', '')[:3], 3)}"
                f"{_ljust(emp.get('fechaCambioSirla', '')[:8], 8)}"
                f"{_rjust_zero(emp.get('gradoInstruccion', 0) or 0, 5)}"
                f"{_ljust((emp.get('discapacidad', '') or '')[:50], 50)}"
            )
            output.write(linea + "\n")
            num_registros += 1

        output.write(f"S{num_registros:06d}\n")
        return output.getvalue()

    @staticmethod
    def to_sirla_txt_dgt9(sirla_data: dict) -> str:
        """Genera archivo fixed-width SIRLA DGT-9 (Suspensión de Contratos).

        Un archivo por cada suspensión activa.
          E: ET9<RNC_11><MMYYYY><fechaInicio_DDMMYYYY><duracion_2><causa_15>  → 45 chars
          D: D<doc_type><doc_25><estab_6><prov_mun_4><dir_300><tel_10>        → 347 chars
          S: S<total_lineas>
        """
        ci = sirla_data.get("company", {})
        rnc = (ci.get("companyRNC", "") or "").replace("-", "").replace(" ", "")[:11]
        output = io.StringIO()

        for susp in sirla_data.get("suspensions", []):
            inicio = susp.get("fechaInicio", "")
            periodo = inicio[4:8] + inicio[2:4] if len(inicio) >= 8 else "00000000"  # MMYYYY from DDMMYYYY
            duracion = susp.get("duracion", 0) or 0
            causa = susp.get("causaCodigo", "0".rjust(15, "0"))
            est_id = (susp.get("establishmentId", "") or "")[:6]

            # Registro E — Encabezado
            output.write(
                f"E"
                f"T9"
                f"{rnc.rjust(11, '0')}"
                f"{periodo}"
                f"{inicio}"
                f"{_rjust_zero(duracion, 2)}"
                f"{causa}".ljust(45) + "\n"
            )

            num_registros = 2  # E + S
            for w in susp.get("trabajadores", []):
                linea = (
                    f"D"
                    f"{w.get('docTypeSirla', 'C')}"
                    f"{_ljust((w.get('documento', '') or '')[:25], 25)}"
                    f"{est_id}"
                    f"{_ljust((w.get('provinciaMunicipio', '') or '')[:4], 4)}"
                    f"{_ljust((w.get('direccion', '') or '')[:300], 300)}"
                    f"{_ljust((w.get('telefono', '') or '')[:10], 10)}"
                )
                output.write(linea + "\n")
                num_registros += 1

            output.write(f"S{num_registros:06d}\n\n")

        return output.getvalue()

    @staticmethod
    def to_excel(lines: list[dict], title: str = "DGT") -> io.BytesIO:
        """Genera archivo .xlsx para revisión (con cabecera y estilos)."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        headers = [
            "Tipo Doc", "Documento", "Nombres", "Apellidos", "Nacionalidad",
            "Sexo", "Fecha Nac.", "Estado Civil", "Salario", "Moneda",
            "Frec. Pago", "Cód. Ocup.", "Ocupación", "Fecha Ingreso",
            "Tipo Contrato", "Horas Sem.", "Turno", "Estado",
            "Tipo Novedad", "Fecha Novedad", "Instrucción", "Vacaciones",
        ]

        # Header style
        h_font = Font(bold=True, size=10, color="FFFFFF")
        h_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        h_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            bottom=Side(style="hair", color="E2E8F0"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = h_font
            cell.fill = h_fill
            cell.alignment = h_align

        for row_idx, emp in enumerate(lines, 2):
            data = [
                emp.get("tipoDocumento", 1),
                emp.get("documento", ""),
                emp.get("nombres", ""),
                emp.get("apellidos", ""),
                emp.get("nacionalidad", 1),
                emp.get("sexo", ""),
                _format_date(emp.get("fechaNacimiento", "")),
                emp.get("estadoCivil", ""),
                float(emp.get("salario", 0)),
                emp.get("tipoMoneda", 1),
                emp.get("frecuenciaPago", 1),
                emp.get("ocupacionCodigo", ""),
                emp.get("ocupacionTexto", ""),
                _format_date(emp.get("fechaIngreso", "")),
                emp.get("tipoContrato", 1),
                emp.get("horasSemanales", 44),
                emp.get("turnoTrabajo", 1),
                emp.get("estadoTrabajador", 1),
                emp.get("tipoNovedad", 0),
                _format_date(emp.get("fechaNovedad", "")),
                emp.get("gradoInstruccion", 0),
                emp.get("concesionVacaciones", 1),
            ]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(size=9)
                cell.border = thin_border
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx == 2:
                    cell.number_format = "@"
                    cell.alignment = Alignment(horizontal="center")

        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    val_len = len(str(cell.value))
                    max_len = max(max_len, val_len)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def to_pdf(lines: list[dict], form_type: str, title: str,
               owner_info: dict = None, data: dict = None) -> io.BytesIO:
        """Genera PDF usando WeasyPrint (debe existir el template)."""
        from flask import render_template
        import weasyprint
        from app.utils.pdf import pdf_write_options

        html = render_template(
            f"rrhh/dgt/{form_type}_pdf.html",
            lines=lines,
            title=title,
            owner_info=owner_info or {},
            data=data or {},
            now=datetime.now(),
        )
        pdf = weasyprint.HTML(string=html).write_pdf(**pdf_write_options())
        output = io.BytesIO(pdf)
        output.seek(0)
        return output
