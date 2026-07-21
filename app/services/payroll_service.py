"""
PayrollService — Cálculo de nómina dominicana con ISR, TSS, prestaciones y asientos contables.

Basado en:
- Ley 16-92 (Código de Trabajo RD)
- Ley 87-01 (Seguridad Social)
- Tablas ISR DGII 2026 (escala anual para personas físicas)
- Norma 08-04 DGII sobre retenciones asalariados
"""

import calendar
from datetime import date, datetime, timedelta
from typing import Optional
import unicodedata

from app.countries.do.payroll_rules import (
    AFP_EMPLOYEE_RATE, AFP_EMPLOYER_RATE,
    SFS_EMPLOYEE_RATE, SFS_EMPLOYER_RATE,
    SRL_EMPLOYER_RATE, INFOTEP_RATE,
    AFP_SALARY_CAP, SFS_SALARY_CAP,
    ISR_ANNUAL_TABLE,
    ANNUAL_EDUCATION_DEDUCTION, MIN_SALARY,
    DEFAULT_OVERTIME_RATE,
    DEFAULT_WORKING_DAYS_PER_MONTH, DEFAULT_WORKING_HOURS_PER_DAY,
    DEFAULT_INFOTEP_THRESHOLD_MULTIPLIER,
    DEFAULT_ACCOUNT_SALARIES_PAYABLE,
    DEFAULT_ACCOUNT_AFP_EMPLOYEE, DEFAULT_ACCOUNT_SFS_EMPLOYEE,
    DEFAULT_ACCOUNT_ISR_EMPLOYEE,
    DEFAULT_ACCOUNT_AFP_EMPLOYER, DEFAULT_ACCOUNT_SFS_EMPLOYER,
    DEFAULT_ACCOUNT_SRL_EMPLOYER,
    DEFAULT_ACCOUNT_INFOTEP_EMPLOYER, DEFAULT_ACCOUNT_INFOTEP_EMPLOYEE,
    DEFAULT_ACCOUNT_OTHER_DEDUCTIONS,
    DEFAULT_COST_CENTER_ACCOUNTS,
)

INFOTEP_EMPLOYEE_THRESHOLD = MIN_SALARY * 5

# Alias privados para compatibilidad interna (usados en el servicio)
_AFP_EMPLOYEE_RATE = AFP_EMPLOYEE_RATE
_AFP_EMPLOYER_RATE = AFP_EMPLOYER_RATE
_SFS_EMPLOYEE_RATE = SFS_EMPLOYEE_RATE
_SFS_EMPLOYER_RATE = SFS_EMPLOYER_RATE
_SRL_EMPLOYER_RATE = SRL_EMPLOYER_RATE
_INFOTEP_RATE = INFOTEP_RATE
_AFP_SALARY_CAP = AFP_SALARY_CAP
_SFS_SALARY_CAP = SFS_SALARY_CAP
_ISR_ANNUAL_TABLE = ISR_ANNUAL_TABLE
_ANNUAL_EDUCATION_DEDUCTION = ANNUAL_EDUCATION_DEDUCTION
_MIN_SALARY = MIN_SALARY
_DEFAULT_OVERTIME_RATE = DEFAULT_OVERTIME_RATE
_DEFAULT_WORKING_DAYS_PER_MONTH = DEFAULT_WORKING_DAYS_PER_MONTH
_DEFAULT_WORKING_HOURS_PER_DAY = DEFAULT_WORKING_HOURS_PER_DAY
_DEFAULT_INFOTEP_THRESHOLD_MULTIPLIER = DEFAULT_INFOTEP_THRESHOLD_MULTIPLIER
_DEFAULT_ACCOUNT_SALARIES_PAYABLE = DEFAULT_ACCOUNT_SALARIES_PAYABLE
_DEFAULT_ACCOUNT_AFP_EMPLOYEE = DEFAULT_ACCOUNT_AFP_EMPLOYEE
_DEFAULT_ACCOUNT_SFS_EMPLOYEE = DEFAULT_ACCOUNT_SFS_EMPLOYEE
_DEFAULT_ACCOUNT_ISR_EMPLOYEE = DEFAULT_ACCOUNT_ISR_EMPLOYEE
_DEFAULT_ACCOUNT_AFP_EMPLOYER = DEFAULT_ACCOUNT_AFP_EMPLOYER
_DEFAULT_ACCOUNT_SFS_EMPLOYER = DEFAULT_ACCOUNT_SFS_EMPLOYER
_DEFAULT_ACCOUNT_SRL_EMPLOYER = DEFAULT_ACCOUNT_SRL_EMPLOYER
_DEFAULT_ACCOUNT_INFOTEP_EMPLOYER = DEFAULT_ACCOUNT_INFOTEP_EMPLOYER
_DEFAULT_ACCOUNT_INFOTEP_EMPLOYEE = DEFAULT_ACCOUNT_INFOTEP_EMPLOYEE
_DEFAULT_ACCOUNT_OTHER_DEDUCTIONS = DEFAULT_ACCOUNT_OTHER_DEDUCTIONS
_DEFAULT_COST_CENTER_ACCOUNTS = DEFAULT_COST_CENTER_ACCOUNTS


class PeriodImmutableError(Exception):
    """El período de nómina está en un estado que no permite modificaciones."""
    pass


class PayrollService:
    """Servicio de cálculo de nómina dominicana."""

    @staticmethod
    def get_rates(tax_rates: dict = None) -> dict:
        """Obtiene tasas desde dict de Firestore o usa valores por defecto."""
        if not tax_rates:
            tax_rates = {}
        return {
            "afp_employee_rate": tax_rates.get("afpEmployeeRate", _AFP_EMPLOYEE_RATE),
            "afp_employer_rate": tax_rates.get("afpEmployerRate", _AFP_EMPLOYER_RATE),
            "sfs_employee_rate": tax_rates.get("sfsEmployeeRate", _SFS_EMPLOYEE_RATE),
            "sfs_employer_rate": tax_rates.get("sfsEmployerRate", _SFS_EMPLOYER_RATE),
            "srl_employer_rate": tax_rates.get("srlEmployerRate", _SRL_EMPLOYER_RATE),
            "infotep_rate": tax_rates.get("infotepRate", _INFOTEP_RATE),
            "afp_salary_cap": tax_rates.get("afpSalaryCap", _AFP_SALARY_CAP),
            "sfs_salary_cap": tax_rates.get("sfsSalaryCap", _SFS_SALARY_CAP),
            "min_salary": tax_rates.get("minSalary", _MIN_SALARY),
            "education_deduction": tax_rates.get("educationDeduction", _ANNUAL_EDUCATION_DEDUCTION),
            "isr_table": tax_rates.get("isrAnnualTable", _ISR_ANNUAL_TABLE),
            "overtime_rate": tax_rates.get("overtimeRate", _DEFAULT_OVERTIME_RATE),
            "working_days_per_month": tax_rates.get("workingDaysPerMonth", _DEFAULT_WORKING_DAYS_PER_MONTH),
            "working_hours_per_day": tax_rates.get("workingHoursPerDay", _DEFAULT_WORKING_HOURS_PER_DAY),
            "infotep_threshold_multiplier": tax_rates.get("infotepThresholdMultiplier", _DEFAULT_INFOTEP_THRESHOLD_MULTIPLIER),
            # ── Cuentas contables ──
            "account_salaries_payable": tax_rates.get("accountSalariesPayable", _DEFAULT_ACCOUNT_SALARIES_PAYABLE),
            "account_afp_employee": tax_rates.get("accountAfpEmployee", _DEFAULT_ACCOUNT_AFP_EMPLOYEE),
            "account_sfs_employee": tax_rates.get("accountSfsEmployee", _DEFAULT_ACCOUNT_SFS_EMPLOYEE),
            "account_isr_employee": tax_rates.get("accountIsrEmployee", _DEFAULT_ACCOUNT_ISR_EMPLOYEE),
            "account_afp_employer": tax_rates.get("accountAfpEmployer", _DEFAULT_ACCOUNT_AFP_EMPLOYER),
            "account_sfs_employer": tax_rates.get("accountSfsEmployer", _DEFAULT_ACCOUNT_SFS_EMPLOYER),
            "account_srl_employer": tax_rates.get("accountSrlEmployer", _DEFAULT_ACCOUNT_SRL_EMPLOYER),
            "account_infotep_employer": tax_rates.get("accountInfotepEmployer", _DEFAULT_ACCOUNT_INFOTEP_EMPLOYER),
            "account_infotep_employee": tax_rates.get("accountInfotepEmployee", _DEFAULT_ACCOUNT_INFOTEP_EMPLOYEE),
            "account_other_deductions": tax_rates.get("accountOtherDeductions", _DEFAULT_ACCOUNT_OTHER_DEDUCTIONS),
            "cost_center_accounts": tax_rates.get("costCenterAccounts", _DEFAULT_COST_CENTER_ACCOUNTS),
        }

    # ═══════════════════════════════════════════════════════════════════════
    # GUARDAS DE ESTADO
    # ═══════════════════════════════════════════════════════════════════════

    IMMUTABLE_STATUSES = ("cerrada", "cancelled")

    @classmethod
    def assert_period_mutable(cls, period: dict):
        status = period.get("status", "")
        if status in cls.IMMUTABLE_STATUSES:
            from app.services.payroll_audit_service import STATUS_LABELS as _s
            raise PeriodImmutableError(
                f"El período está en estado «{_s.get(status, status)}» y no puede modificarse. "
                f"Revierta el período primero."
            )

    @classmethod
    def is_period_mutable(cls, period: dict) -> bool:
        return period.get("status", "") not in cls.IMMUTABLE_STATUSES

    @classmethod
    def resolve_rates(cls, owner_uid: str, group_id: str = "", employee_id: str = "",
                      sandbox: bool = True) -> dict:
        """Resuelve tasas en cascada: Política → Grupo → Empleado → Global.

        Orden de precedencia:
        1. Override del empleado (si existe)
        2. Override del grupo (si tiene policyOverrides)
        3. Política asignada al grupo (si tiene policyId)
        4. Política default del tenant
        5. Tasas globales (tax_rates en hr_config) — fallback final
        """
        from app.services import hr_data_service as hr
        from app.models.payroll_policy import PolicyOverride

        base_rates = cls.get_rates(hr.get_tax_rates(owner_uid, sandbox=sandbox))

        # Nivel 4: Política default
        default_policy = hr.get_default_payroll_policy(owner_uid, sandbox=sandbox)
        if default_policy:
            policy_rates = cls.get_rates({
                "afpEmployeeRate": default_policy.get("afpEmployeeRate"),
                "afpEmployerRate": default_policy.get("afpEmployerRate"),
                "sfsEmployeeRate": default_policy.get("sfsEmployeeRate"),
                "sfsEmployerRate": default_policy.get("sfsEmployerRate"),
                "srlEmployerRate": default_policy.get("srlEmployerRate"),
                "infotepRate": default_policy.get("infotepRate"),
                "afpSalaryCap": default_policy.get("afpSalaryCap"),
                "sfsSalaryCap": default_policy.get("sfsSalaryCap"),
                "minSalary": default_policy.get("minSalary"),
                "educationDeduction": default_policy.get("educationDeduction"),
                "isrAnnualTable": default_policy.get("isrAnnualTable"),
                "overtimeRate": default_policy.get("overtimeRate"),
                "workingDaysPerMonth": default_policy.get("workingDaysPerMonth"),
                "workingHoursPerDay": default_policy.get("workingHoursPerDay"),
                "infotepThresholdMultiplier": default_policy.get("infotepThresholdMultiplier"),
                "accountSalariesPayable": default_policy.get("accountSalariesPayable"),
                "accountAfpEmployee": default_policy.get("accountAfpEmployee"),
                "accountSfsEmployee": default_policy.get("accountSfsEmployee"),
                "accountIsrEmployee": default_policy.get("accountIsrEmployee"),
                "accountAfpEmployer": default_policy.get("accountAfpEmployer"),
                "accountSfsEmployer": default_policy.get("accountSfsEmployer"),
                "accountSrlEmployer": default_policy.get("accountSrlEmployer"),
                "accountInfotepEmployer": default_policy.get("accountInfotepEmployer"),
                "accountInfotepEmployee": default_policy.get("accountInfotepEmployee"),
                "accountOtherDeductions": default_policy.get("accountOtherDeductions"),
                "costCenterAccounts": default_policy.get("costCenterAccounts"),
            })
            for key in policy_rates:
                if policy_rates[key] is not None:
                    base_rates[key] = policy_rates[key]

        # Nivel 3: Política asignada al grupo
        if group_id:
            group = hr.get_payroll_group(owner_uid, group_id, sandbox=sandbox)
            if group:
                policy_id = group.get("policyId", "")
                if policy_id:
                    policy = hr.get_payroll_policy(owner_uid, policy_id, sandbox=sandbox)
                    if policy:
                        group_policy_rates = cls.get_rates({
                            "afpEmployeeRate": policy.get("afpEmployeeRate"),
                            "afpEmployerRate": policy.get("afpEmployerRate"),
                            "sfsEmployeeRate": policy.get("sfsEmployeeRate"),
                            "sfsEmployerRate": policy.get("sfsEmployerRate"),
                            "srlEmployerRate": policy.get("srlEmployerRate"),
                            "infotepRate": policy.get("infotepRate"),
                            "afpSalaryCap": policy.get("afpSalaryCap"),
                            "sfsSalaryCap": policy.get("sfsSalaryCap"),
                            "minSalary": policy.get("minSalary"),
                            "educationDeduction": policy.get("educationDeduction"),
                            "isrAnnualTable": policy.get("isrAnnualTable"),
                            "overtimeRate": policy.get("overtimeRate"),
                            "workingDaysPerMonth": policy.get("workingDaysPerMonth"),
                            "workingHoursPerDay": policy.get("workingHoursPerDay"),
                            "infotepThresholdMultiplier": policy.get("infotepThresholdMultiplier"),
                            "accountSalariesPayable": policy.get("accountSalariesPayable"),
                            "accountAfpEmployee": policy.get("accountAfpEmployee"),
                            "accountSfsEmployee": policy.get("accountSfsEmployee"),
                            "accountIsrEmployee": policy.get("accountIsrEmployee"),
                            "accountAfpEmployer": policy.get("accountAfpEmployer"),
                            "accountSfsEmployer": policy.get("accountSfsEmployer"),
                            "accountSrlEmployer": policy.get("accountSrlEmployer"),
                            "accountInfotepEmployer": policy.get("accountInfotepEmployer"),
                            "accountInfotepEmployee": policy.get("accountInfotepEmployee"),
                            "accountOtherDeductions": policy.get("accountOtherDeductions"),
                            "costCenterAccounts": policy.get("costCenterAccounts"),
                        })
                        for key in group_policy_rates:
                            if group_policy_rates[key] is not None:
                                base_rates[key] = group_policy_rates[key]

                # Nivel 2: Overrides del grupo
                group_overrides = group.get("policyOverrides")
                if group_overrides:
                    override = PolicyOverride(**group_overrides)
                    base_rates = override.apply_to(base_rates)

        # Nivel 1: Override del empleado (futuro: leer de EmploymentContract o Employee)
        # Actualmente no implementado — reservado para fase futura

        return base_rates

    @staticmethod
    def get_period_lines(period: dict, owner_uid: str = "", sandbox: bool = True) -> list:
        """Obtiene las líneas de un período desde subcolección, con fallback a embebidas."""
        from app.services import hr_data_service as _hr
        return _hr.get_payroll_lines_unified(period, owner_uid=owner_uid, sandbox=sandbox)

    @staticmethod
    def merge_group_overrides(global_rates: dict, overrides: dict) -> dict:
        """Mergea groupOverrides del grupo sobre tasas globales.

        NOTA: Solo se mergean campos NO legales (cuentas contables,
        centros de costo). Los parámetros legales (tasas TSS, ISR,
        topes) son globales por ley y NO deben variar por grupo.
        """
        merged = dict(global_rates)
        overrideable_fields = [
            "accountSalariesPayable", "accountAfpEmployee",
            "accountSfsEmployee", "accountIsrEmployee",
            "accountAfpEmployer", "accountSfsEmployer",
            "accountSrlEmployer", "accountInfotepEmployer",
            "infotepEmployeeRate", "accountInfotepEmployee",
            "accountOtherDeductions", "costCenterAccounts",
        ]
        for field in overrideable_fields:
            if field in overrides and overrides[field] is not None:
                merged[field] = overrides[field]
        return merged

    @classmethod
    def resolve_overtime_rates(cls, owner_uid: str, group_id: str = "",
                                sandbox: bool = True) -> dict:
        """Resuelve tasas de horas extra desde la configuración del grupo o defaults."""
        defaults = {
            "default_rate": _DEFAULT_OVERTIME_RATE,
            "night_rate": 2.0,
            "holiday_rate": 2.5,
        }
        if not group_id:
            return defaults
        try:
            from app.services import hr_data_service as _hr
            group = _hr.get_payroll_group(owner_uid, group_id, sandbox=sandbox)
            if group and group.get("overtimeRules"):
                rules = group["overtimeRules"]
                defaults["default_rate"] = rules.get("default_rate", defaults["default_rate"])
                defaults["night_rate"] = rules.get("night_rate", defaults["night_rate"])
                defaults["holiday_rate"] = rules.get("holiday_rate", defaults["holiday_rate"])
        except Exception:
            pass
        return defaults

    @classmethod
    def calculate_overtime_pay(cls, base_salary: float, hours: float,
                                overtime_type: str = "default",
                                overtime_rules: dict = None) -> float:
        """Calcula pago de horas extra según tipo (default, night, holiday)."""
        rules = overtime_rules or {
            "default_rate": _DEFAULT_OVERTIME_RATE,
            "night_rate": 2.0,
            "holiday_rate": 2.5,
        }
        rate_key = f"{overtime_type}_rate" if overtime_type != "default" else "default_rate"
        rate = rules.get(rate_key, rules.get("default_rate", _DEFAULT_OVERTIME_RATE))
        return cls._calculate_overtime(base_salary, hours, rate)

    @classmethod
    def detect_anomalies(cls, lines: list, employees: dict,
                         previous_lines: list = None,
                         owner_uid: str = "", sandbox: bool = True) -> dict:
        """Detecta anomalías en líneas de nómina recién calculadas.

        Returns:
            Dict con {warnings: [...], errors: [...]}
        """
        warnings = []
        errors = []

        prev_map = {}
        if previous_lines:
            prev_map = {pl.get("employeeId", ""): pl for pl in previous_lines}

        for line in lines:
            emp_id = line.get("employeeId", "")
            emp_name = line.get("employeeName", emp_id)
            emp = employees.get(emp_id, {})

            if not emp.get("afpProvider"):
                errors.append(
                    f"{emp_name}: no tiene AFP asignada. La nómina no puede calcularse sin AFP."
                )

            net_salary = line.get("netSalary", 0)
            if net_salary <= 0 and line.get("totalIncome", 0) > 0:
                warnings.append(
                    f"{emp_name}: salario neto RD$ {net_salary:,.2f} — "
                    f"verifica que las deducciones no excedan el ingreso."
                )

            base_salary = emp.get("baseSalary", emp.get("salary", 0))
            if base_salary <= 0:
                errors.append(
                    f"{emp_name}: salario base en cero o no definido."
                )

            total_income = line.get("totalIncome", 0)
            if base_salary > 0 and total_income > 0:
                variation = abs(total_income - base_salary) / base_salary
                if variation > 0.5:
                    warnings.append(
                        f"{emp_name}: ingreso total (RD$ {total_income:,.2f}) "
                        f"varía más del 50% respecto al salario base (RD$ {base_salary:,.2f})."
                    )

            if prev_map and emp_id in prev_map:
                prev = prev_map[emp_id]
                prev_net = prev.get("netSalary", 0)
                if prev_net > 0:
                    pct_change = abs(net_salary - prev_net) / prev_net
                    if pct_change > 0.20:
                        direction = "aumentó" if net_salary > prev_net else "disminuyó"
                        warnings.append(
                            f"{emp_name}: salario neto {direction} "
                            f"{pct_change * 100:.0f}% respecto al período anterior "
                            f"(RD$ {prev_net:,.2f} → RD$ {net_salary:,.2f})."
                        )

            if line.get("overtimeHours", 0) > 40:
                warnings.append(
                    f"{emp_name}: {line['overtimeHours']} horas extra — "
                    f"excede el límite legal de 40 horas/semana equivalentes."
                )

        return {"warnings": warnings, "errors": errors}

    @classmethod
    def validate_employees_before_payroll(cls, employees: list) -> dict:
        """Pre-validación de empleados antes de calcular nómina.
        Detecta errores y advertencias que impedirían o afectarían el cálculo.

        Returns:
            Dict con {errors: [...], warnings: [...]}
            Cada item: {"employeeId": id, "employeeName": name, "issue": str, "field": str}
        """
        errors = []
        warnings = []

        for emp in employees:
            emp_id = emp.get("id", "")
            emp_name = emp.get("fullName") or f"{emp.get('firstName', '')} {emp.get('firstLastName', emp.get('lastName', ''))}".strip()
            if not emp_name:
                emp_name = emp_id

            if not emp.get("afpProvider"):
                errors.append({
                    "employeeId": emp_id,
                    "employeeName": emp_name,
                    "issue": "No tiene AFP asignada. La nómina no puede calcularse sin AFP.",
                    "field": "afpProvider",
                })

            base = float(emp.get("baseSalary") or emp.get("salary", 0) or 0)
            if base <= 0:
                errors.append({
                    "employeeId": emp_id,
                    "employeeName": emp_name,
                    "issue": "Salario base en cero o no definido.",
                    "field": "baseSalary",
                })

            if not emp.get("cedula"):
                warnings.append({
                    "employeeId": emp_id,
                    "employeeName": emp_name,
                    "issue": "Sin cédula registrada. Requerida para reportes TSS y DGT-3.",
                    "field": "cedula",
                })

        return {"errors": errors, "warnings": warnings}

    # ═══════════════════════════════════════════════════════════════════════
    # CÁLCULO DE NÓMINA INDIVIDUAL
    # ═══════════════════════════════════════════════════════════════════════

    @classmethod
    def calculate_payroll_line(
        cls,
        base_salary: float,
        tax_rates: dict = None,
        overtime_hours: float = 0.0,
        overtime_rate: float = None,
        commission: float = 0.0,
        bonus: float = 0.0,
        other_income: float = 0.0,
        other_deductions: float = 0.0,
        education_deduction: float = 0.0,
        period_type: str = "mensual",
        prorated_salary: float = None,
    ) -> dict:
        """
        Calcula una línea de nómina completa.

        Args:
            base_salary: Salario base mensual bruto.
            tax_rates: Dict con tasas, topes, tabla ISR y parámetros de cálculo configurables.
            overtime_hours: Horas extras trabajadas.
            overtime_rate: Multiplicador de hora extra (si es None, se usa el de tax_rates o default 1.35).
            commission: Comisiones del período.
            bonus: Bonificación.
            other_income: Otros ingresos.
            other_deductions: Otros descuentos (préstamos, anticipos).
            education_deduction: Deducción mensual por educación (anual/12).
            period_type: "mensual" o "quincenal".
            prorated_salary: Salario ya prorrateado para el período (entrada a mitad, salida, cambio salarial).
                             Si es None, se calcula normalmente.
        """
        r = cls.get_rates(tax_rates)
        if overtime_rate is None:
            overtime_rate = r["overtime_rate"]
        # ── 1. Ingresos ──────────────────────────────────────────────────
        if prorated_salary is not None:
            period_salary = prorated_salary
            period_factor = 24 if period_type == "quincenal" else 12
        elif period_type == "quincenal":
            period_salary = round(base_salary / 2, 2)
            period_factor = 24
        else:
            period_salary = base_salary
            period_factor = 12

        overtime_pay = cls._calculate_overtime(
            base_salary, overtime_hours, overtime_rate,
            working_days=r["working_days_per_month"],
            working_hours=r["working_hours_per_day"],
        )
        gross_salary = period_salary
        total_income = round(period_salary + overtime_pay + commission + bonus + other_income, 2)

        # ── 2. Salario cotizable TSS (topado) ────────────────────────────
        afp_cap = r["afp_salary_cap"] / 2 if period_type == "quincenal" else r["afp_salary_cap"]
        sfs_cap = r["sfs_salary_cap"] / 2 if period_type == "quincenal" else r["sfs_salary_cap"]
        afp_cotizable = min(total_income, afp_cap)
        sfs_cotizable = min(total_income, sfs_cap)

        # ── 3. Descuentos al empleado ───────────────────────────────────
        afp_employee = round(afp_cotizable * r["afp_employee_rate"], 2)
        sfs_employee = round(sfs_cotizable * r["sfs_employee_rate"], 2)

        # INFOTEP empleado: aplica si total_income > umbral * salario mínimo
        infotep_threshold = r["min_salary"] * r["infotep_threshold_multiplier"]
        if total_income > infotep_threshold:
            infotep_employee = round(total_income * r["infotep_rate"], 2)
        else:
            infotep_employee = 0.0

        # ISR: se calcula sobre renta neta anualizable con el factor correcto
        isr_monthly = cls._calculate_isr_monthly(
            total_income, afp_employee, sfs_employee, education_deduction, period_factor,
            tax_rates=r,
        )

        total_deductions = round(afp_employee + sfs_employee + infotep_employee + isr_monthly + other_deductions, 2)
        net_salary = round(total_income - total_deductions, 2)

        # ── 4. Aportes empleador ────────────────────────────────────────
        afp_employer = round(afp_cotizable * r["afp_employer_rate"], 2)
        sfs_employer = round(sfs_cotizable * r["sfs_employer_rate"], 2)
        srl_employer = round(sfs_cotizable * r["srl_employer_rate"], 2)
        infotep_employer = round(total_income * r["infotep_rate"], 2)
        total_employer = round(afp_employer + sfs_employer + srl_employer + infotep_employer, 2)

        return {
            "baseSalary": base_salary,
            "grossSalary": gross_salary,
            "overtimeHours": overtime_hours,
            "overtimePay": overtime_pay,
            "commission": commission,
            "bonus": bonus,
            "otherIncome": other_income,
            "totalIncome": total_income,
            "afpEmployee": afp_employee,
            "sfsEmployee": sfs_employee,
            "infotepEmployee": infotep_employee,
            "isrRetention": isr_monthly,
            "otherDeductions": other_deductions,
            "totalDeductions": total_deductions,
            "netSalary": net_salary,
            "afpEmployer": afp_employer,
            "sfsEmployer": sfs_employer,
            "srlEmployer": srl_employer,
            "infotepEmployer": infotep_employer,
            "totalEmployerContrib": total_employer,
        }

    @classmethod
    def _calculate_overtime(cls, base_salary: float, hours: float, rate: float,
                            working_days: float = 23.83, working_hours: float = 8.0) -> float:
        """Calcula pago de horas extras. Base: (salario / días / horas) * rate."""
        if hours <= 0 or base_salary <= 0:
            return 0.0
        hourly = base_salary / working_days / working_hours
        return round(hourly * rate * hours, 2)

    @classmethod
    def _calculate_isr_monthly(
        cls,
        monthly_income: float,
        afp_deduction: float,
        sfs_deduction: float,
        education_deduction: float = 0.0,
        period_factor: int = 12,
        tax_rates: dict = None,
    ) -> float:
        """
        Calcula ISR del período según tabla DGII para asalariados.

        Método: anualiza el ingreso usando el factor correcto según frecuencia
        (12 para mensual, 24 para quincenal), resta deducciones, aplica tabla
        progresiva anual, divide entre period_factor para obtener el ISR del período.

        Args:
            period_factor: 12 para mensual, 24 para quincenal, 52 para semanal.
        """
        r = cls.get_rates(tax_rates)
        annual_gross = monthly_income * period_factor
        annual_afp = afp_deduction * period_factor
        annual_sfs = sfs_deduction * period_factor
        annual_edu = min(education_deduction * period_factor, r["education_deduction"])

        # Renta neta anual
        annual_taxable = max(0.0, annual_gross - annual_afp - annual_sfs - annual_edu)

        # Aplicar tabla ISR — cada tramo ya incluye la cuota fija acumulada de los anteriores
        annual_isr = 0.0
        for floor, ceiling, rate, fixed in r["isr_table"]:
            if annual_taxable <= floor:
                break
            if annual_taxable > ceiling:
                continue
            annual_isr = round((annual_taxable - floor) * rate + fixed, 2)
            break

        return round(annual_isr / period_factor, 2)

    # ═══════════════════════════════════════════════════════════════════════
    # PRESTACIONES LABORALES (Ley 16-92)
    # ═══════════════════════════════════════════════════════════════════════

    @classmethod
    def calculate_christmas_bonus(cls, base_salary: float, months_worked: int = 12) -> float:
        """
        Regalía pascual (salario de Navidad).
        = 1/12 del salario anual, proporcional a meses trabajados.
        Se paga antes del 20 de diciembre.
        """
        if months_worked >= 12:
            return round(base_salary, 2)
        return round((base_salary / 12.0) * months_worked, 2)

    @classmethod
    def calculate_vacation_days(cls, hire_date_str: str, today: Optional[date] = None,
                                taken_days: int = 0) -> int:
        """
        Calcula días de vacaciones acumulados según Ley 16-92.

        - 1 a 5 años: 14 días hábiles por año
        - Más de 5 años: 18 días hábiles por año
        - Proporcional si no ha cumplido el año

        Args:
            hire_date_str: Fecha de contratación (YYYY-MM-DD).
            today: Fecha de referencia (por defecto hoy).
            taken_days: Días ya tomados y aprobados a descontar.

        Returns:
            Días hábiles disponibles (acumulados - tomados).
        """
        if today is None:
            today = date.today()

        try:
            hire_date = datetime.strptime(hire_date_str[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return 0

        # Años completos trabajados
        years = today.year - hire_date.year
        if today.month < hire_date.month or (today.month == hire_date.month and today.day < hire_date.day):
            years -= 1

        if years < 0:
            years = 0

        # Días proporcionales del año en curso
        if years < 1:
            days_since_hire = (today - hire_date).days
            return max(0, round((days_since_hire / 365.0) * 14))

        # Base: 14 o 18 días por año según antigüedad
        days_per_year = 18 if years >= 5 else 14
        total = years * days_per_year

        # Días proporcionales del año actual
        last_anniversary = hire_date.replace(year=today.year)
        if last_anniversary > today:
            last_anniversary = last_anniversary.replace(year=today.year - 1)
        days_proportional = max(0, round(((today - last_anniversary).days / 365.0) * days_per_year))
        total += days_proportional

        total -= taken_days
        return max(0, total)

    @classmethod
    def calculate_severance(cls, base_salary: float, hire_date_str: str, termination_date_str: str = "") -> dict:
        """
        Calcula prestaciones por cesantía/desahucio según Ley 16-92.

        - Preaviso: Art. 76 (7-28 días de salario según antigüedad)
        - Cesantía: Art. 80 (6-21 días por año trabajado, máx 10 años en base legal base)

        Returns:
            Dict con preaviso, cesantía, total, y vacaciones pendientes.
        """
        today = date.today()
        if termination_date_str:
            try:
                today = datetime.strptime(termination_date_str[:10], "%Y-%m-%d").date()
            except ValueError:
                pass

        try:
            hire_date = datetime.strptime(hire_date_str[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return {"preaviso": 0.0, "cesantia": 0.0, "vacaciones_pendientes": 0.0, "total": 0.0}

        months = (today.year - hire_date.year) * 12 + (today.month - hire_date.month)
        if today.day < hire_date.day:
            months -= 1
        years = max(0, months // 12)
        daily_salary = base_salary / 23.83

        # Preaviso (Art. 76)
        if months < 3:
            preaviso_days = 0
        elif months < 6:
            preaviso_days = 7
        elif months < 12:
            preaviso_days = 14
        else:
            preaviso_days = 28

        # Cesantía (Art. 80)
        # 6 días por cada año (o fracción > 6 meses), topado a los años de la escala
        full_years = years
        remaining_months = months - (full_years * 12)
        if remaining_months >= 6:
            full_years += 1
        cesantia_days = min(full_years, 10) * 21 + (max(0, full_years - 10) * 0)
        # Simplificado: 21 días por año, máx 10 años
        cesantia_days = min(years * 21, 10 * 21)
        if cesantia_days == 0 and months >= 3:
            cesantia_days = 6  # Fracción mínima

        vacaciones_pendientes = cls.calculate_vacation_days(hire_date_str, today)
        vacaciones_pago = daily_salary * vacaciones_pendientes

        preaviso_amount = round(daily_salary * preaviso_days, 2)
        cesantia_amount = round(daily_salary * cesantia_days, 2)
        total = round(preaviso_amount + cesantia_amount + vacaciones_pago, 2)

        return {
            "preaviso": preaviso_amount,
            "preaviso_days": preaviso_days,
            "cesantia": cesantia_amount,
            "cesantia_days": cesantia_days,
            "vacaciones_pendientes": vacaciones_pago,
            "vacaciones_days": vacaciones_pendientes,
            "total": total,
        }

    @classmethod
    def prorate_salary(
        cls,
        monthly_salary: float,
        period_start: str,
        period_end: str,
        hire_date: str = "",
        termination_date: str = "",
        salary_history: list = None,
    ) -> float | None:
        try:
            ps = datetime.strptime(period_start, "%Y-%m-%d").date()
            pe = datetime.strptime(period_end, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return monthly_salary
        period_days = (pe - ps).days + 1
        if period_days <= 0:
            return 0.0
        emp_start = ps
        emp_end = pe
        if hire_date:
            try:
                hd = datetime.strptime(hire_date[:10], "%Y-%m-%d").date()
                emp_start = max(ps, hd)
            except (ValueError, TypeError):
                pass
        if termination_date:
            try:
                td = datetime.strptime(termination_date[:10], "%Y-%m-%d").date()
                emp_end = min(pe, td)
            except (ValueError, TypeError):
                pass
        if emp_start > emp_end:
            return 0.0
        emp_days = (emp_end - emp_start).days + 1

        # Si el empleado trabajó el período completo y no hay cambios salariales,
        # retorna None para que calculate_payroll_line use la fórmula estándar
        if emp_start == ps and emp_end == pe:
            has_changes = False
            if salary_history:
                for h in salary_history:
                    try:
                        eff = datetime.strptime(h.get("effectiveDate", ""), "%Y-%m-%d").date()
                        if ps <= eff <= pe:
                            has_changes = True
                            break
                    except (ValueError, TypeError):
                        continue
            if not has_changes:
                return None

        daily_rate = monthly_salary / 23.83

        if salary_history:
            total = 0.0
            for h in sorted(salary_history, key=lambda x: x.get("effectiveDate", "")):
                try:
                    eff = datetime.strptime(h["effectiveDate"], "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    continue
                end_date_str = h.get("endDate", "")
                try:
                    end_d = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else pe
                except (ValueError, TypeError):
                    end_d = pe
                seg_start = max(eff, emp_start)
                seg_end = min(end_d, emp_end)
                if seg_start <= seg_end:
                    seg_days = (seg_end - seg_start).days + 1
                    seg_amount = h.get("amount", monthly_salary)
                    seg_daily = seg_amount / 23.83
                    total += round(seg_daily * seg_days, 2)
            if total > 0:
                return round(total, 2)

        return round(daily_rate * emp_days, 2)

    @classmethod
    def calculate_business_days(cls, start_date_str: str, end_date_str: str) -> int:
        """Cuenta días hábiles (lunes-viernes) entre dos fechas."""
        try:
            start = datetime.strptime(start_date_str[:10], "%Y-%m-%d").date()
            end = datetime.strptime(end_date_str[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return 0

        if start > end:
            return 0

        days = 0
        current = start
        while current <= end:
            if current.weekday() < 5:  # L-V
                days += 1
            current += timedelta(days=1)
        return days

    # ═══════════════════════════════════════════════════════════════════════
    # RETROACTIVIDAD SALARIAL
    # ═══════════════════════════════════════════════════════════════════════

    @classmethod
    def calculate_retroactive_pay(cls, employee: dict, salary_history: list,
                                   new_salary: float, effective_date: str,
                                   tax_rates: dict = None) -> dict:
        """Calcula la retroactividad salarial cuando un aumento se aplica con efecto retroactivo.

        Args:
            employee: Dict del empleado con sus datos actuales.
            salary_history: Lista de cambios salariales previos.
            new_salary: Nuevo salario mensual.
            effective_date: Fecha desde la cual aplica el nuevo salario (YYYY-MM-DD).
            tax_rates: Tasas impositivas vigentes.

        Returns:
            Dict con {retroactiveMonths, retroactivePay, adjustedNet, details}
        """
        from datetime import date as dt_date, timedelta

        r = cls.get_rates(tax_rates)
        today = dt_date.today()
        try:
            eff_date = datetime.strptime(effective_date[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return {"retroactiveMonths": 0, "retroactivePay": 0, "adjustedNet": 0,
                    "details": [], "error": "Fecha de vigencia inválida"}

        if eff_date >= today:
            return {"retroactiveMonths": 0, "retroactivePay": 0, "adjustedNet": 0,
                    "details": [], "note": "La fecha de vigencia es futura, no hay retroactividad"}

        # Determinar cuántos meses completos hay entre la fecha efectiva y hoy
        months_retroactive = (today.year - eff_date.year) * 12 + (today.month - eff_date.month)
        if today.day < eff_date.day:
            months_retroactive -= 1
        months_retroactive = max(0, months_retroactive)

        if months_retroactive == 0:
            return {"retroactiveMonths": 0, "retroactivePay": 0, "adjustedNet": 0,
                    "details": [], "note": "Menos de un mes desde la fecha efectiva"}

        base_salary = float(employee.get("baseSalary", employee.get("salary", 0)))
        if base_salary <= 0:
            return {"retroactiveMonths": months_retroactive, "retroactivePay": 0,
                    "adjustedNet": 0, "details": [], "error": "Salario base no definido"}

        salary_diff = new_salary - base_salary
        if salary_diff <= 0:
            return {"retroactiveMonths": 0, "retroactivePay": 0, "adjustedNet": 0,
                    "details": [], "note": "No hay diferencia positiva de salario"}

        retroactive_gross = round(salary_diff * months_retroactive, 2)

        from app.services.concept_engine import ConceptEngine, TSSContext, ISRContext
        params = {
            "afp_salary_cap": r.get("afp_salary_cap", 464460.0),
            "sfs_salary_cap": r.get("sfs_salary_cap", 232230.0),
            "afp_employee_rate": r.get("afp_employee_rate", 0.0287),
            "sfs_employee_rate": r.get("sfs_employee_rate", 0.0304),
            "afp_employer_rate": r.get("afp_employer_rate", 0.0710),
            "sfs_employer_rate": r.get("sfs_employer_rate", 0.0709),
            "srl_employer_rate": r.get("srl_employer_rate", 0.0120),
            "infotep_rate": r.get("infotep_rate", 0.01),
            "min_salary": r.get("min_salary", 23223.0),
            "education_deduction": r.get("education_deduction", 50000.0),
            "isr_table": r.get("isr_table", []),
        }
        is_q = employee.get("paymentFrequency", "mensual") in ("quincenal", "semanal")

        def _calc_net_for_salary(sal):
            tss_ctx = TSSContext(base_salary=sal, gross_income=sal, is_quincenal=is_q)
            afp_emp = ConceptEngine.resolve_employee_tss("AFP_EMPLEADO", tss_ctx, params)
            sfs_emp = ConceptEngine.resolve_employee_tss("SFS_EMPLEADO", tss_ctx, params)
            gross_after_tss = max(0, sal - afp_emp - sfs_emp)
            isr_ctx = ISRContext(gross_income=gross_after_tss, is_quincenal=is_q,
                                 afp_deduction=afp_emp, sfs_deduction=sfs_emp)
            isr = ConceptEngine.resolve_isr(isr_ctx, params)
            return round(sal - afp_emp - sfs_emp - isr, 2), round(afp_emp + sfs_emp + isr, 2)

        old_net, _ = _calc_net_for_salary(base_salary)
        new_net, _ = _calc_net_for_salary(new_salary)
        monthly_net_diff = round(new_net - old_net, 2)
        retroactive_net = round(monthly_net_diff * months_retroactive, 2)

        details = []
        for i in range(months_retroactive):
            m = eff_date.month + i
            y = eff_date.year + (m - 1) // 12
            m = ((m - 1) % 12) + 1
            details.append({
                "month": f"{y}-{m:02d}",
                "oldSalary": base_salary,
                "newSalary": new_salary,
                "grossDiff": round(salary_diff, 2),
                "netDiff": round(monthly_net_diff, 2),
            })

        return {
            "retroactiveMonths": months_retroactive,
            "retroactiveGross": retroactive_gross,
            "retroactiveNet": retroactive_net,
            "monthlyNetDiff": monthly_net_diff,
            "salaryDiff": round(salary_diff, 2),
            "oldSalary": base_salary,
            "newSalary": new_salary,
            "effectiveDate": effective_date,
            "details": details,
        }

    # ═══════════════════════════════════════════════════════════════════════
    # WHAT-IF ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════

    @classmethod
    def what_if_analysis(cls, employees: list, scenario: dict,
                         tax_rates: dict = None) -> dict:
        """Simula escenarios de cambio salarial y proyecta impacto.

        Args:
            employees: Lista de empleados (dicts) a analizar.
            scenario: Dict con la simulación:
                - type: "pct_increase", "fixed_increase", "department_increase", "executive_bonus"
                - value: float (porcentaje o monto)
                - filter_department: str (opcional)
                - filter_area: str (opcional)
                - include_employer_contrib: bool
            tax_rates: Tasas impositivas.

        Returns:
            Dict con {scenario, currentMonthly, projectedMonthly, impact, perEmployee}
        """
        r = cls.get_rates(tax_rates)
        scenario_type = scenario.get("type", "pct_increase")
        value = float(scenario.get("value", 0))
        filter_dept = scenario.get("filter_department", "")
        filter_area = scenario.get("filter_area", "")
        include_employer = scenario.get("include_employer_contrib", True)

        current_total_gross = 0.0
        current_total_net = 0.0
        current_total_employer = 0.0
        projected_total_gross = 0.0
        projected_total_net = 0.0
        projected_total_employer = 0.0
        affected = 0
        per_employee = []

        for emp in employees:
            if emp.get("status") != "activo":
                continue
            dept = emp.get("department", emp.get("area", ""))
            area = emp.get("area", "")
            if filter_dept and filter_dept != dept:
                continue
            if filter_area and filter_area != area:
                continue

            base = float(emp.get("baseSalary", emp.get("salary", 0)))
            if base <= 0:
                continue

            current_line = cls.calculate_payroll_line(base_salary=base, tax_rates=tax_rates)
            current_total_gross += current_line["totalIncome"]
            current_total_net += current_line["netSalary"]
            current_total_employer += current_line["totalEmployerContrib"]

            new_base = base
            if scenario_type == "pct_increase":
                new_base = round(base * (1 + value / 100), 2)
            elif scenario_type == "fixed_increase":
                new_base = round(base + value, 2)
            elif scenario_type == "department_increase":
                new_base = round(base * (1 + value / 100), 2)
            elif scenario_type == "executive_bonus":
                new_base = base

            projected_line = cls.calculate_payroll_line(
                base_salary=new_base,
                bonus=value if scenario_type == "executive_bonus" else 0,
                tax_rates=tax_rates,
            )
            projected_total_gross += projected_line["totalIncome"]
            projected_total_net += projected_line["netSalary"]
            projected_total_employer += projected_line["totalEmployerContrib"]

            affected += 1
            per_employee.append({
                "employeeId": emp.get("id", ""),
                "employeeName": emp.get("fullName", ""),
                "currentBase": base,
                "newBase": new_base,
                "currentNet": current_line["netSalary"],
                "projectedNet": projected_line["netSalary"],
                "netDiff": round(projected_line["netSalary"] - current_line["netSalary"], 2),
                "employerDiff": round(projected_line["totalEmployerContrib"] - current_line["totalEmployerContrib"], 2),
            })

        net_impact = round(projected_total_net - current_total_net, 2)
        employer_impact = round(projected_total_employer - current_total_employer, 2) if include_employer else 0
        total_impact = round(net_impact + employer_impact, 2)

        return {
            "scenario": {
                "type": scenario_type,
                "value": value,
                "label": cls._what_if_label(scenario_type, value),
            },
            "affectedEmployees": affected,
            "currentMonthly": {
                "gross": round(current_total_gross, 2),
                "net": round(current_total_net, 2),
                "employer": round(current_total_employer, 2),
            },
            "projectedMonthly": {
                "gross": round(projected_total_gross, 2),
                "net": round(projected_total_net, 2),
                "employer": round(projected_total_employer, 2),
            },
            "impact": {
                "net": net_impact,
                "employer": employer_impact,
                "total": total_impact,
                "annualTotal": round(total_impact * 12, 2),
            },
            "perEmployee": per_employee,
        }

    @staticmethod
    def _what_if_label(scenario_type: str, value: float) -> str:
        labels = {
            "pct_increase": f"Aumento del {value}% en salario base",
            "fixed_increase": f"Aumento fijo de RD$ {value:,.2f}",
            "department_increase": f"Aumento del {value}% por departamento",
            "executive_bonus": f"Bono ejecutivo de RD$ {value:,.2f}",
        }
        return labels.get(scenario_type, f"Escenario: {scenario_type}")

    # ═══════════════════════════════════════════════════════════════════════
    # ASIENTO CONTABLE DE NÓMINA
    # ═══════════════════════════════════════════════════════════════════════

    DEFAULT_COST_CENTER_ACCOUNTS = {
        "General": "6.2.1.01",
        "Ventas": "6.2.1.01.01",
        "Produccion": "6.2.1.01.02",
        "Administrativa": "6.2.1.01.03",
    }

    @classmethod
    def build_payroll_accounting_lines(cls, payroll_period: dict, employees: dict = None, tax_rates: dict = None,
                                       owner_uid: str = "", sandbox: bool = True) -> list:
        lines = []
        period_label = payroll_period.get("periodKey", "")
        plines = cls.get_period_lines(payroll_period, owner_uid=owner_uid, sandbox=sandbox)
        employees = employees or {}
        from app.services.hr_data_service import get_tax_rates_snapshot
        snapshot = get_tax_rates_snapshot(payroll_period)
        effective_rates = snapshot if snapshot else tax_rates
        r = cls.get_rates(effective_rates)

        total_gross = 0.0
        total_net = 0.0
        total_afp_emp = 0.0
        total_sfs_emp = 0.0
        total_isr = 0.0
        total_afp_empl = 0.0
        total_sfs_empl = 0.0
        total_srl_empl = 0.0
        total_infotep = 0.0
        total_infotep_emp = 0.0
        total_other_ded = 0.0

        cc_accounts = r.get("cost_center_accounts", cls.DEFAULT_COST_CENTER_ACCOUNTS)

        by_cc = {}
        for pl in plines:
            total_net += pl.get("netSalary", 0)
            total_afp_emp += pl.get("afpEmployee", 0)
            total_sfs_emp += pl.get("sfsEmployee", 0)
            total_isr += pl.get("isrRetention", 0)
            total_afp_empl += pl.get("afpEmployer", 0)
            total_sfs_empl += pl.get("sfsEmployer", 0)
            total_srl_empl += pl.get("srlEmployer", 0)
            total_infotep += pl.get("infotepEmployer", 0)
            total_infotep_emp += pl.get("infotepEmployee", 0)
            total_other_ded += pl.get("otherDeductions", 0)

            emp_id = pl.get("employeeId", "")
            emp = employees.get(emp_id, {})
            cc = emp.get("costCenter", "") or emp.get("area", "") or emp.get("department", "") or "General"
            if cc not in by_cc:
                by_cc[cc] = {"gross": 0.0, "employer": 0.0}
            by_cc[cc]["gross"] += pl.get("totalIncome", 0)
            by_cc[cc]["employer"] += pl.get("totalEmployerContrib", 0)
            total_gross += pl.get("totalIncome", 0)

        total_employer = total_afp_empl + total_sfs_empl + total_srl_empl + total_infotep

        # DEBE: Líneas de gasto por centro de costo
        for cc, totals in by_cc.items():
            cc_gasto = round(totals["gross"] + totals["employer"], 2)
            acct_code = cc_accounts.get(cc, "6.2.1.01")
            lines.append({
                "accountCode": acct_code,
                "accountName": f"Sueldos y salarios - {cc}",
                "debit": cc_gasto,
                "credit": 0.00,
                "description": f"Nómina período {period_label} - {cc}",
            })

        # HABER: Salarios por pagar (neto)
        if total_net > 0:
            lines.append({
                "accountCode": r["account_salaries_payable"],
                "accountName": "Salarios por pagar",
                "debit": 0.00, "credit": round(total_net, 2),
                "description": f"Salario neto período {period_label}",
            })
        if total_afp_emp > 0:
            lines.append({
                "accountCode": r["account_afp_employee"],
                "accountName": "Retenciones empleado AFP",
                "debit": 0.00, "credit": round(total_afp_emp, 2),
                "description": f"AFP empleado {period_label}",
            })
        if total_sfs_emp > 0:
            lines.append({
                "accountCode": r["account_sfs_employee"],
                "accountName": "Retenciones a empleado SFS",
                "debit": 0.00, "credit": round(total_sfs_emp, 2),
                "description": f"SFS empleado {period_label}",
            })
        if total_isr > 0:
            lines.append({
                "accountCode": r["account_isr_employee"],
                "accountName": "Retención ISR empleados",
                "debit": 0.00, "credit": round(total_isr, 2),
                "description": f"ISR empleados {period_label}",
            })
        if total_afp_empl > 0:
            lines.append({
                "accountCode": r["account_afp_employer"],
                "accountName": "Acumulaciones AFP",
                "debit": 0.00, "credit": round(total_afp_empl, 2),
                "description": f"AFP empleador {period_label}",
            })
        if total_sfs_empl > 0:
            lines.append({
                "accountCode": r["account_sfs_employer"],
                "accountName": "Acumulaciones SFS",
                "debit": 0.00, "credit": round(total_sfs_empl, 2),
                "description": f"SFS empleador {period_label}",
            })
        if total_srl_empl > 0:
            lines.append({
                "accountCode": r["account_srl_employer"],
                "accountName": "Acumulaciones SRL",
                "debit": 0.00, "credit": round(total_srl_empl, 2),
                "description": f"SRL empleador {period_label}",
            })
        if total_infotep > 0:
            lines.append({
                "accountCode": r["account_infotep_employer"],
                "accountName": "Acumulaciones INFOTEP",
                "debit": 0.00, "credit": round(total_infotep, 2),
                "description": f"INFOTEP {period_label}",
            })
        if total_infotep_emp > 0:
            lines.append({
                "accountCode": r["account_infotep_employee"],
                "accountName": "Retención INFOTEP empleados",
                "debit": 0.00, "credit": round(total_infotep_emp, 2),
                "description": f"INFOTEP empleado {period_label}",
            })
        if total_other_ded > 0:
            lines.append({
                "accountCode": r["account_other_deductions"],
                "accountName": "Deducciones varias por pagar",
                "debit": 0.00, "credit": round(total_other_ded, 2),
                "description": f"Otras deducciones {period_label}",
            })

        return lines

    # ═══════════════════════════════════════════════════════════════════════
    # PROVISIONES CONTABLES DE NÓMINA (DEVENGADO)
    # ═══════════════════════════════════════════════════════════════════════

    @classmethod
    def build_vacation_provision_lines(cls, employees: list, tax_rates: dict = None,
                                       month_label: str = "") -> list:
        lines = []
        r = cls.get_rates(tax_rates)
        total_provision = 0.0
        for emp in employees:
            base = float(emp.get("baseSalary", emp.get("salary", 0)))
            if base <= 0:
                continue
            monthly_provision = round(base / 12.0, 2)
            total_provision += monthly_provision
        if total_provision <= 0:
            return lines
        lines.append({
            "accountCode": r.get("account_vacation_expense", "6.1.1.05"),
            "accountName": "Provisión vacaciones",
            "debit": round(total_provision, 2), "credit": 0.00,
            "description": f"Provisión mensual de vacaciones {month_label}",
        })
        lines.append({
            "accountCode": r.get("account_vacation_payable", "2.1.2.1.03"),
            "accountName": "Vacaciones por pagar (provisión)",
            "debit": 0.00, "credit": round(total_provision, 2),
            "description": f"Provisión mensual de vacaciones {month_label}",
        })
        return lines

    @classmethod
    def build_christmas_bonus_provision_lines(cls, employees: list, tax_rates: dict = None,
                                              month_label: str = "") -> list:
        lines = []
        r = cls.get_rates(tax_rates)
        total_provision = 0.0
        for emp in employees:
            base = float(emp.get("baseSalary", emp.get("salary", 0)))
            if base <= 0:
                continue
            monthly_provision = round(base / 12.0, 2)
            total_provision += monthly_provision
        if total_provision <= 0:
            return lines
        lines.append({
            "accountCode": r.get("account_christmas_expense", "6.1.1.02"),
            "accountName": "Provisión salario de navidad",
            "debit": round(total_provision, 2), "credit": 0.00,
            "description": f"Provisión mensual de regalía pascual {month_label}",
        })
        lines.append({
            "accountCode": r.get("account_christmas_payable", "2.1.2.1.01"),
            "accountName": "Regalía pascual por pagar (provisión)",
            "debit": 0.00, "credit": round(total_provision, 2),
            "description": f"Provisión mensual de regalía pascual {month_label}",
        })
        return lines

    @classmethod
    def generate_monthly_provisions(cls, owner_uid: str, month_label: str = "",
                                     sandbox: bool = True) -> dict:
        from app.services.accounting_service import AccountingService
        from app.services import hr_data_service as hr
        from app.services.db_service import DatabaseService

        employees = [e for e in hr.get_employees(owner_uid, sandbox=sandbox) if e.get("status") == "activo"]
        if not employees:
            return {"vacationEntry": None, "christmasEntry": None, "note": "Sin empleados activos"}

        tax_rates = hr.get_tax_rates(owner_uid, sandbox=sandbox)
        AccountingService.seed_default_accounts(owner_uid)
        accounts = DatabaseService.get_chart_of_accounts(owner_uid)

        results = {}
        vac_lines = cls.build_vacation_provision_lines(employees, tax_rates, month_label)
        if vac_lines:
            vac_full = []
            for al in vac_lines:
                acc = next((a for a in accounts if a.get("code") == al["accountCode"]), None)
                if acc:
                    vac_full.append({
                        "accountId": acc["id"], "accountCode": al["accountCode"],
                        "accountName": al["accountName"], "debit": al["debit"],
                        "credit": al["credit"], "description": al["description"],
                    })
            if vac_full:
                results["vacationEntry"] = AccountingService.generate_entry(owner_uid, {
                    "entryType": "provision",
                    "date": "", "concept": f"Provisión mensual de vacaciones {month_label}",
                    "referenceType": "payroll_provision", "referenceId": f"vac-{month_label}",
                    "referenceNumber": month_label, "lines": vac_full,
                    "createdBy": "system",
                }, sandbox=sandbox)

        chr_lines = cls.build_christmas_bonus_provision_lines(employees, tax_rates, month_label)
        if chr_lines:
            chr_full = []
            for al in chr_lines:
                acc = next((a for a in accounts if a.get("code") == al["accountCode"]), None)
                if acc:
                    chr_full.append({
                        "accountId": acc["id"], "accountCode": al["accountCode"],
                        "accountName": al["accountName"], "debit": al["debit"],
                        "credit": al["credit"], "description": al["description"],
                    })
            if chr_full:
                results["christmasEntry"] = AccountingService.generate_entry(owner_uid, {
                    "entryType": "provision",
                    "date": "", "concept": f"Provisión mensual de regalía pascual {month_label}",
                    "referenceType": "payroll_provision", "referenceId": f"chr-{month_label}",
                    "referenceNumber": month_label, "lines": chr_full,
                    "createdBy": "system",
                }, sandbox=sandbox)

        return results

    # ═══════════════════════════════════════════════════════════════════════
    # VALIDACIONES PRE-CIERRE FISCAL
    # ═══════════════════════════════════════════════════════════════════════

    @classmethod
    def validate_payroll_for_fiscal_closing(cls, owner_uid: str, year: int = None,
                                            sandbox: bool = True) -> dict:
        """Valida que todos los períodos de nómina del año estén listos para el cierre fiscal.

        Returns:
            Dict con {is_valid, warnings, errors, details}
        """
        from datetime import date as dt_date
        from app.services import hr_data_service as hr

        if year is None:
            year = dt_date.today().year

        periods = hr.get_payroll_periods(owner_uid, sandbox=sandbox)
        year_periods = [p for p in periods if p.get("year") == year]

        errors = []
        warnings = []
        details = {"totalPeriods": len(year_periods), "closedCount": 0, "openCount": 0,
                   "totalGross": 0.0, "totalNet": 0.0, "totalIsr": 0.0, "totalTss": 0.0}

        if not year_periods:
            warnings.append(f"No se encontraron períodos de nómina para el año {year}.")
            return {"isValid": False, "warnings": warnings, "errors": errors, "details": details}

        employees = {e["id"]: e for e in hr.get_employees(owner_uid, sandbox=sandbox)}

        for p in year_periods:
            status = p.get("status", "")
            if status == "cerrada":
                details["closedCount"] += 1
            else:
                details["openCount"] += 1
                errors.append(
                    f"Período '{p.get('periodKey', '?')}' ({p.get('periodRange', '')}) "
                    f"está en estado «{status}». Debe estar «cerrada» para el cierre fiscal."
                )

            plines = cls.get_period_lines(p, owner_uid=owner_uid, sandbox=sandbox)
            for line in plines:
                details["totalGross"] += line.get("totalIncome", 0)
                details["totalNet"] += line.get("netSalary", 0)
                details["totalIsr"] += line.get("isrRetention", 0)
                details["totalTss"] += (
                    line.get("afpEmployee", 0) + line.get("sfsEmployee", 0) +
                    line.get("afpEmployer", 0) + line.get("sfsEmployer", 0) +
                    line.get("srlEmployer", 0) + line.get("infotepEmployer", 0)
                )

                emp_id = line.get("employeeId", "")
                emp = employees.get(emp_id, {})
                if not emp.get("afpProvider"):
                    warnings.append(
                        f"Empleado {line.get('employeeName', emp_id)} no tiene AFP asignada "
                        f"(período {p.get('periodKey', '')})."
                    )
                if not (emp.get("baseSalary") or emp.get("salary")):
                    warnings.append(
                        f"Empleado {line.get('employeeName', emp_id)} tiene salario en cero "
                        f"(período {p.get('periodKey', '')})."
                    )

        details["totalGross"] = round(details["totalGross"], 2)
        details["totalNet"] = round(details["totalNet"], 2)
        details["totalIsr"] = round(details["totalIsr"], 2)
        details["totalTss"] = round(details["totalTss"], 2)

        is_valid = len(errors) == 0

        return {
            "isValid": is_valid,
            "errors": errors,
            "warnings": warnings,
            "details": details,
        }

    # ═══════════════════════════════════════════════════════════════════════
    # VALIDACIÓN IR-18 / DGII
    # ═══════════════════════════════════════════════════════════════════════

    @classmethod
    def validate_ir18_readiness(cls, owner_uid: str, year: int = None,
                                 sandbox: bool = True) -> dict:
        """Valida que los datos de IR-18 estén listos para reporte DGII.

        Verifica YTD por empleado, conciliación de ISR retenido, y datos requeridos
        para el formulario IR-18 de la DGII.

        Returns:
            Dict con {isReady, warnings, errors, summary}
        """
        from datetime import date as dt_date
        from app.services import hr_data_service as hr

        if year is None:
            year = dt_date.today().year

        errors = []
        warnings = []
        employees = hr.get_employees(owner_uid, sandbox=sandbox)
        active = [e for e in employees if e.get("status") == "activo"]
        periods = hr.get_payroll_periods(owner_uid, sandbox=sandbox)
        year_periods = [p for p in periods if p.get("year") == year]

        total_isr_ytd = 0.0
        total_isr_periods = 0.0
        employees_ready = 0
        employees_missing = []

        for emp in active:
            emp_id = emp["id"]
            name = emp.get("fullName", emp_id)
            issues = []

            if not emp.get("cedula") and not emp.get("idNumber"):
                issues.append("Sin cédula/RNC")
            if not emp.get("nationality"):
                issues.append("Sin nacionalidad")
            if not emp.get("occupationCode"):
                issues.append("Sin código de ocupación CNO-2019")

            ytd = None
            try:
                from app.services.payroll_ytd_service import get_ytd
                ytd = get_ytd(owner_uid, emp_id, year, sandbox=sandbox)
            except Exception:
                pass

            if ytd and ytd.get("totalIsr", 0) > 0:
                total_isr_ytd += ytd.get("totalIsr", 0)
                employees_ready += 1
            else:
                if emp.get("baseSalary", emp.get("salary", 0)) > 0:
                    warnings.append(f"{name}: sin acumulación YTD de ISR para {year}")

            if issues:
                employees_missing.append({"name": name, "issues": issues})

        for p in year_periods:
            plines = cls.get_period_lines(p, owner_uid=owner_uid, sandbox=sandbox)
            for line in plines:
                total_isr_periods += line.get("isrRetention", 0)

        total_isr_ytd = round(total_isr_ytd, 2)
        total_isr_periods = round(total_isr_periods, 2)
        isr_discrepancy = round(abs(total_isr_ytd - total_isr_periods), 2)

        if isr_discrepancy > 1.0:
            warnings.append(
                f"Discrepancia entre ISR acumulado YTD (RD$ {total_isr_ytd:,.2f}) "
                f"e ISR de períodos (RD$ {total_isr_periods:,.2f}). "
                f"Diferencia: RD$ {isr_discrepancy:,.2f}"
            )

        if not year_periods:
            errors.append(f"No hay períodos de nómina registrados para {year}")

        is_ready = len(errors) == 0 and len(employees_missing) == 0

        return {
            "isReady": is_ready,
            "year": year,
            "errors": errors,
            "warnings": warnings,
            "summary": {
                "totalEmployees": len(active),
                "employeesReady": employees_ready,
                "employeesMissing": employees_missing,
                "totalIsrYtd": total_isr_ytd,
                "totalIsrPeriods": total_isr_periods,
                "isrDiscrepancy": isr_discrepancy,
                "totalPeriods": len(year_periods),
            },
        }

    # ═══════════════════════════════════════════════════════════════════════
    # AUTO-GENERACIÓN DE PERÍODOS
    # ═══════════════════════════════════════════════════════════════════════

    @classmethod
    def generate_upcoming_periods(cls, owner_uid: str, months_ahead: int = 6,
                                   sandbox: bool = True) -> dict:
        """Genera automáticamente períodos futuros para todos los grupos activos.

        Útil para programación (APScheduler) o llamado manual desde el dashboard.
        Solo crea períodos en estado 'borrador' si no existen ya.

        Args:
            owner_uid: UID del tenant.
            months_ahead: Cuántos meses hacia adelante generar períodos.
            sandbox: Modo sandbox.

        Returns:
            Dict con {created: int, skipped: int, errors: list}
        """
        import uuid
        import calendar as cal_mod
        from datetime import date as dt_date, timedelta
        from app.services import hr_data_service as hr

        created = 0
        skipped = 0
        errors = []

        groups = hr.get_payroll_groups(owner_uid, sandbox=sandbox)
        active_groups = [g for g in groups if g.get("isActive", True)]
        if not active_groups:
            return {"created": 0, "skipped": 0, "errors": ["No hay grupos de nómina activos."]}

        today = dt_date.today()
        existing_periods = hr.get_payroll_periods(owner_uid, sandbox=sandbox)
        existing_keys = {(p.get("periodKey"), p.get("payrollGroupId", "")) for p in existing_periods}

        for g in active_groups:
            freq = g.get("frequency", "mensual")
            gid = g["id"]

            for offset in range(months_ahead):
                target = today.replace(day=1) + timedelta(days=32 * offset)
                target = target.replace(day=1)
                year = target.year
                month = target.month

                if freq == "mensual":
                    last_day = cal_mod.monthrange(year, month)[1]
                    period_key = f"{year}-{month:02d}-M"
                    period_start = f"{year}-{month:02d}-01"
                    period_end = f"{year}-{month:02d}-{last_day}"
                    period_type = "mensual"
                    month_es = cls._MESES[month]
                    period_range = f"M: {month_es} {year}"

                    if (period_key, gid) in existing_keys:
                        skipped += 1
                        continue

                    period_id = str(uuid.uuid4())
                    hr.save_payroll_period(owner_uid, period_id, {
                        "id": period_id,
                        "payrollGroupId": gid,
                        "periodKey": period_key,
                        "periodType": period_type,
                        "periodRange": period_range,
                        "startDate": period_start,
                        "endDate": period_end,
                        "month": month,
                        "year": year,
                        "status": "borrador",
                        "lines": [],
                        "totalGross": 0,
                        "totalNet": 0,
                        "totalEmployerContrib": 0,
                        "statusHistory": [{
                            "from": "",
                            "to": "borrador",
                            "by": "auto-generator",
                            "at": today.isoformat(),
                            "comment": "Período generado automáticamente",
                        }],
                    }, sandbox=sandbox)
                    existing_keys.add((period_key, gid))
                    created += 1

                elif freq == "quincenal":
                    for q in (1, 2):
                        period_key = f"{year}-{month:02d}-{q}"
                        if (period_key, gid) in existing_keys:
                            skipped += 1
                            continue
                        last_day = cal_mod.monthrange(year, month)[1]
                        month_es = cls._MESES[month]
                        if q == 1:
                            period_start = f"{year}-{month:02d}-01"
                            period_end = f"{year}-{month:02d}-15"
                            period_range = f"Q1: 1 {month_es} - 15 {month_es}"
                        else:
                            period_start = f"{year}-{month:02d}-16"
                            period_end = f"{year}-{month:02d}-{last_day}"
                            period_range = f"Q2: 16 {month_es} - {last_day} {month_es}"

                        period_id = str(uuid.uuid4())
                        hr.save_payroll_period(owner_uid, period_id, {
                            "id": period_id,
                            "payrollGroupId": gid,
                            "periodKey": period_key,
                            "periodType": "quincenal",
                            "periodRange": period_range,
                            "startDate": period_start,
                            "endDate": period_end,
                            "month": month,
                            "year": year,
                            "status": "borrador",
                            "lines": [],
                            "totalGross": 0,
                            "totalNet": 0,
                            "totalEmployerContrib": 0,
                            "statusHistory": [{
                                "from": "",
                                "to": "borrador",
                                "by": "auto-generator",
                                "at": today.isoformat(),
                                "comment": "Período generado automáticamente",
                            }],
                        }, sandbox=sandbox)
                        existing_keys.add((period_key, gid))
                        created += 1

        return {"created": created, "skipped": skipped, "errors": errors}

    # ═══════════════════════════════════════════════════════════════════════
    # EXPORTACIÓN TSS
    # ═══════════════════════════════════════════════════════════════════════

    @classmethod
    def generate_tss_csv(cls, payroll_period: dict, employees: list,
                         owner_uid: str = "", sandbox: bool = True) -> str:
        """
        Genera CSV en formato TSS (Tesorería de la Seguridad Social).

        Columnas requeridas por TSS:
        Cédula, Nombre, Salario Cotizable, AFP Empleado, SFS Empleado,
        AFP Empleador, SFS Empleador, SRL, INFOTEP, Total
        """
        import csv, io

        output = io.StringIO()
        writer = csv.writer(output)

        # Header TSS
        writer.writerow([
            "Cedula", "Nombre", "Salario_Cotizable", "AFP_Empleado",
            "SFS_Empleado", "AFP_Empleador", "SFS_Empleador",
            "SRL", "INFOTEP", "Total_Aportes"
        ])

        period_label = payroll_period.get("periodKey", "")
        lines = cls.get_period_lines(payroll_period, owner_uid=owner_uid, sandbox=sandbox)
        emp_map = {e.get("id", ""): e for e in employees}

        for pl in lines:
            emp = emp_map.get(pl.get("employeeId", ""), {})
            writer.writerow([
                pl.get("cedula", ""),
                pl.get("employeeName", ""),
                f"{pl.get('totalIncome', 0):.2f}",
                f"{pl.get('afpEmployee', 0):.2f}",
                f"{pl.get('sfsEmployee', 0):.2f}",
                f"{pl.get('afpEmployer', 0):.2f}",
                f"{pl.get('sfsEmployer', 0):.2f}",
                f"{pl.get('srlEmployer', 0):.2f}",
                f"{pl.get('infotepEmployer', 0):.2f}",
                f"{pl.get('afpEmployee',0) + pl.get('sfsEmployee',0) + pl.get('afpEmployer',0) + pl.get('sfsEmployer',0) + pl.get('srlEmployer',0) + pl.get('infotepEmployer',0):.2f}",
            ])

        return output.getvalue()

    # ═══════════════════════════════════════════════════════════════════════
    # AUTODETERMINACIÓN TSS — Formato oficial Tesorería de la Seguridad Social
    # ═══════════════════════════════════════════════════════════════════════


    # Meses en español para el período
    _MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
              "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    @staticmethod
    def _ascii_upper(s: str) -> str:
        """Convierte a uppercase ASCII: elimina tildes, ñ→n, etc."""
        nfkd = unicodedata.normalize("NFKD", s)
        return nfkd.encode("ascii", "ignore").decode("ascii").upper()

    @classmethod
    def generate_tss_autodeterminacion(cls, payroll_period: dict, employees: list,
                                        employer_rnc: str = "",
                                        owner_uid: str = "", sandbox: bool = True) -> dict:
        """
        Genera archivo de Autodeterminación TSS en formato OFICIAL SUIRPLUS v6.0.

        Formato: archivo de texto de ancho fijo con 3 tipos de registro:
          E = Encabezado (20 caracteres)
          D = Detalle (366 caracteres por empleado)
          S = Sumario (7 caracteres)

        Especificación: Instructivo Construcción Archivos Autodeterminación y
        Novedades v6.0 — TSS, Junio 2025.

        Args:
            payroll_period: Dict del período de nómina.
            employees: Lista de empleados con datos completos.
            employer_rnc: RNC o Cédula del empleador (sin guiones).

        Returns:
            Dict con {content, filename, periodo, periodo_tss, total_empleados, resumen}.
        """
        from datetime import datetime

        period_key = payroll_period.get("periodKey", "")
        year = payroll_period.get("year", datetime.now().year)
        month = payroll_period.get("month", datetime.now().month)
        periodo_mmaaaa = f"{month:02d}{year}"
        periodo_label = f"{cls._MESES[month]}_{year}"

        lines = cls.get_period_lines(payroll_period, owner_uid=owner_uid, sandbox=sandbox)
        emp_map = {e.get("id", ""): e for e in employees}

        # Limpiar RNC: solo dígitos, quitar guiones
        rnc = "".join(c for c in (employer_rnc or "") if c.isdigit())
        if len(rnc) < 11:
            rnc = rnc.rjust(11)

        output_lines = []

        # ═══════════════════════════════════════════════════════════════
        # ENCABEZADO — 20 caracteres
        # Pos: 1(1) E, 2-3(2) AM, 4-14(11) RNC, 15-20(6) MMAAAA
        # ═══════════════════════════════════════════════════════════════
        header = f"EAM{rnc[-11:].rjust(11)}{periodo_mmaaaa}"
        assert len(header) == 20, f"Encabezado: {len(header)} chars, deben ser 20"
        output_lines.append(header)

        empleados_contados = 0

        for pl in lines:
            emp = emp_map.get(pl.get("employeeId", ""), {})
            if not emp:
                continue

            empleados_contados += 1

            # ── Datos del empleado ──

            # Clave nómina (3, justificada derecha)
            tss_key = str(emp.get("tssKey", "") or "").strip()[:3]
            clave = tss_key.rjust(3)

            # Tipo documento (1)
            id_type = (emp.get("idType", "") or "cedula").lower()
            tipo_doc = "C" if id_type == "cedula" else ("P" if id_type == "pasaporte" else "N")

            # Documento (25, justificado izquierda, sin guiones)
            doc = "".join(c for c in (emp.get("cedula", "") or emp.get("idNumber", "") or "") if c.isalnum())
            documento = doc.ljust(25)[:25]

            # Nombres (50, justificado izquierda) — ASCII uppercase sin tildes
            nombres_raw = f"{emp.get('firstName', '')} {emp.get('middleName', '')}".strip()
            nombres = cls._ascii_upper(nombres_raw).ljust(50)[:50]

            # 1er Apellido (40) — ASCII uppercase sin tildes
            apellido1 = (emp.get("firstLastName", "") or emp.get("lastName", "") or "").strip()
            apellido1 = cls._ascii_upper(apellido1).ljust(40)[:40]

            # 2do Apellido (40) — ASCII uppercase sin tildes
            apellido2 = (emp.get("secondLastName", "") or "").strip()
            apellido2 = cls._ascii_upper(apellido2).ljust(40)[:40]

            # Sexo (1)
            gender = (emp.get("gender", "") or "").lower()
            sexo = "M" if gender in ("masculino", "male", "m") else "F"

            # Fecha nacimiento (8, DDMMAAAA)
            birth = (emp.get("birthDate", "") or "").strip()
            fecha_nac = ""
            if birth:
                try:
                    bd = datetime.strptime(birth[:10], "%Y-%m-%d")
                    fecha_nac = bd.strftime("%d%m%Y")
                except ValueError:
                    fecha_nac = ""
            fecha_nac = fecha_nac.ljust(8)[:8]

            # ── Salarios ──
            # Regla: todo se reporta en TSS. Salario_SS = total devengado.
            # Solo son exentos: regalía (01), preaviso/cesantía (02), pensión alimenticia (03).
            total_income = pl.get("totalIncome", 0) or 0
            afp_cap = emp.get("afpSalaryCap", 0) or _AFP_SALARY_CAP
            salario_ss = min(total_income, afp_cap)

            # Salario_SS (16, ceros izq con 2 decimales) — incluye TODO: base + extras + comisiones + bonos
            salario_ss_str = f"{salario_ss:016.2f}"[:16]

            # Aporte voluntario (16, ceros)
            aporte_vol = "0000000000000.00"

            # Salario_ISR (16): igual a Salario_SS (todo tributa salvo exentos)
            salario_isr_str = salario_ss_str

            # Otras remuneraciones (16): 0, todo va dentro de Salario_SS
            otras_rem_str = "0000000000000.00"

            # RNC agente retención (11, justificado derecha)
            agente_ret = "".rjust(11)

            # Remun. otros empleadores (16, ceros)
            rem_otros = "0000000000000.00"

            # Ingresos exentos ISR (16): siempre en cero (se desglosan en 01/02/03)
            ingresos_exentos = "0000000000000.00"

            # Saldo a favor (16, ceros)
            saldo_favor = "0000000000000.00"

            # Salario INFOTEP (16): igual a Salario_SS
            salario_infotep_str = salario_ss_str

            # Tipo ingreso (4): default 0001 (Normal)
            tipo_ingreso = "0001"

            # ── Ingresos exentos desglosados (18 chars c/u: código 2 + monto 16) ──
            # 01 = Regalía Pascual
            regalia = pl.get("christmasBonus", 0) or 0
            regalia_str = f"01{regalia:016.2f}"[:18]

            # 02 = Preaviso, Cesantía, Viáticos e Indemnizaciones
            preaviso_cesantia = 0.0
            pc_str = f"02{preaviso_cesantia:016.2f}"[:18]

            # 03 = Retención Pensión Alimenticia
            pension = pl.get("pensionAlimenticia", 0) or 0
            pension_str = f"03{pension:016.2f}"[:18]

            # ═══════════════════════════════════════════════════════════
            # DETALLE — 366 caracteres
            # ═══════════════════════════════════════════════════════════
            detalle = (
                "D" + clave + tipo_doc + documento + nombres + apellido1 +
                apellido2 + sexo + fecha_nac + salario_ss_str + aporte_vol +
                salario_isr_str + otras_rem_str + agente_ret + rem_otros +
                ingresos_exentos + saldo_favor + salario_infotep_str +
                tipo_ingreso + regalia_str + pc_str + pension_str
            )
            assert len(detalle) == 366, f"Detalle: {len(detalle)} chars, deben ser 366"
            output_lines.append(detalle)

        # ═══════════════════════════════════════════════════════════════
        # SUMARIO — 7 caracteres
        # Pos: 1(1) S, 2-7(6) total registros (E + D's + S)
        # ═══════════════════════════════════════════════════════════════
        total_registros = 1 + empleados_contados + 1  # header + details + trailer
        trailer = f"S{total_registros:06d}"
        output_lines.append(trailer)

        content = "\n".join(output_lines) + "\n"

        # Clean RNC for filename (just numbers, no spaces)
        rnc_clean = "".join(c for c in (employer_rnc or "000000000") if c.isdigit())
        filename = f"AM_{rnc_clean}_{periodo_mmaaaa}.txt"

        return {
            "content": content,
            "filename": filename,
            "periodo": periodo_label,
            "periodo_tss": periodo_mmaaaa,
            "total_empleados": empleados_contados,
            "resumen": {
                "total_empleados": empleados_contados,
                "total_registros": total_registros,
            },
        }

    @classmethod
    def generate_tss_autodeterminacion_xls(cls, payroll_period: dict, employees: list,
                                            employer_rnc: str = "",
                                            tipo_archivo: str = "AM",
                                            owner_uid: str = "", sandbox: bool = True) -> dict:
        """
        Genera el archivo Excel (.xlsx) poblado con la plantilla oficial de
        Autodeterminación TSS de la Tesorería de la Seguridad Social RD.

        Columnas exactas de la plantilla oficial (21 columnas B-V, col A vacía):
          B: Clave Nómina        C: Tipo Doc.           D: Número Documento
          E: Nombres             F: 1er. Apellido       G: 2do. Apellido
          H: Sexo                I: Fecha Nacimiento
          J: Salario Cotizable   K: Aporte Voluntario
          L: Salario ISR         M: Tipo Ingreso         N: Otras Remuneraciones
          O: RNC/Céd. Agente Ret P: Remun. Otros Agentes Q: Remuneración del período
          R: Saldo a favor       S: Regalía Pascual      T: Preaviso/Cesantía/Indemniz.
          U: Retención Pensión   V: Salario INFOTEP

        Args:
            payroll_period: Dict del período de nómina.
            employees: Lista de empleados con datos completos.
            employer_rnc: RNC o Cédula del empleador.
            tipo_archivo: "AM" (Modificación) o "AR" (Reemplazo).

        Returns:
            Dict con {content (bytes), filename, periodo, total_empleados, resumen}.
        """
        import io
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Autodeterminación"

        period_key = payroll_period.get("periodKey", "")
        year = payroll_period.get("year", datetime.now().year)
        month = payroll_period.get("month", datetime.now().month)
        periodo_mmaaaa = f"{month:02d}{year}"
        periodo_label = f"{cls._MESES[month]}_{year}"

        lines = cls.get_period_lines(payroll_period, owner_uid=owner_uid, sandbox=sandbox)
        emp_map = {e.get("id", ""): e for e in employees}

        # ── Estilos ──
        bold_font = Font(bold=True, size=10)
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True, size=9)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Metadata rows ──
        ws.merge_cells("B1:V1")
        ws["B1"] = "Plantilla de Archivo AutoDeterminación"
        ws["B1"].font = title_font

        ws["B2"] = "Tipo de Archivo:"
        ws["B2"].font = bold_font
        ws["C2"] = tipo_archivo

        ws["B3"] = "RNC o Cédula:"
        ws["B3"].font = bold_font
        ws["C3"] = employer_rnc

        ws["B4"] = "Período:"
        ws["B4"].font = bold_font
        ws["C4"] = periodo_mmaaaa

        ws["B5"] = "# de Empleados:"
        ws["B5"].font = bold_font

        # ── Section headers (row 7) ──
        section_row = 7
        ws.merge_cells(f"B{section_row}:I{section_row}")
        ws[f"B{section_row}"] = "TRABAJADORES"
        ws[f"B{section_row}"].font = Font(bold=True, size=10)
        ws[f"B{section_row}"].fill = section_fill
        ws[f"B{section_row}"].alignment = center_align

        ws.merge_cells(f"J{section_row}:K{section_row}")
        ws[f"J{section_row}"] = "SDSS"
        ws[f"J{section_row}"].font = Font(bold=True, size=10)
        ws[f"J{section_row}"].fill = section_fill
        ws[f"J{section_row}"].alignment = center_align

        ws.merge_cells(f"L{section_row}:T{section_row}")
        ws[f"L{section_row}"] = "DGII"
        ws[f"L{section_row}"].font = Font(bold=True, size=10)
        ws[f"L{section_row}"].fill = section_fill
        ws[f"L{section_row}"].alignment = center_align

        ws.merge_cells(f"U{section_row}:V{section_row}")
        ws[f"U{section_row}"] = "INFOTEP"
        ws[f"U{section_row}"].font = Font(bold=True, size=10)
        ws[f"U{section_row}"].fill = section_fill
        ws[f"U{section_row}"].alignment = center_align

        # ── Column headers (row 8-9, 2-line headers) ──
        headers_r1 = [
            "", "Clave Nómina", "Tipo Doc.", "Número Documento",
            "Nombres", "1er. Apellido", "2do. Apellido", "Sexo",
            "Fecha Nacimiento", "Salario Cotizable", "Aporte Voluntario",
            "Salario ISR", "Tipo Ingreso", "Otras Remuneraciones",
            "RNC/Céd. Agente Ret", "Remun. Otros Agentes",
            "Remuneración del período", "Saldo a favor (Saldo 13)",
            "Regalía Pascual", "Preaviso, Cesantía, Viático e Indemnizaciones",
            "Retención Pensión Alimenticia", "Salario INFOTEP",
        ]

        header_row = 8
        for col_idx, h in enumerate(headers_r1):
            cell = ws.cell(row=header_row, column=col_idx + 1, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # ── Widths ──
        widths = [3, 12, 10, 16, 18, 16, 16, 6, 14, 14, 14,
                  14, 12, 16, 18, 16, 16, 14, 14, 22, 16, 14]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        # ── Data rows ──
        data_start = 9
        total_afp_emp = 0.0
        total_sfs_emp = 0.0
        total_afp_empl = 0.0
        total_sfs_empl = 0.0
        total_srl = 0.0
        total_infotep = 0.0
        emp_count = 0

        for pl in lines:
            emp = emp_map.get(pl.get("employeeId", ""), {})
            if not emp:
                continue

            emp_count += 1
            row = data_start + emp_count

            # Tipo doc
            id_type = emp.get("idType", "cedula") or "cedula"
            tipo_doc = "C" if id_type == "cedula" else ("P" if id_type == "pasaporte" else "N")

            # Sexo
            gender = (emp.get("gender", "") or "").lower()
            sexo = "M" if gender in ("masculino", "male", "m") else ("F" if gender else "")

            # Fecha nacimiento
            birth = (emp.get("birthDate", "") or "").strip()
            if birth:
                try:
                    bd = datetime.strptime(birth[:10], "%Y-%m-%d")
                    birth = bd.strftime("%d/%m/%Y")
                except ValueError:
                    pass

            # Nombres
            nombres = " ".join(p for p in [
                emp.get("firstName", ""), emp.get("middleName", "")
            ] if p).strip()

            total_income = pl.get("totalIncome", 0)
            afp_cap = emp.get("afpSalaryCap", 0) or _AFP_SALARY_CAP
            sfs_cap = emp.get("sfsSalaryCap", 0) or _SFS_SALARY_CAP
            salario_cotizable = min(total_income, afp_cap)

            # Aportes individuales
            afp_emp_val = pl.get("afpEmployee", 0)
            sfs_emp_val = pl.get("sfsEmployee", 0)
            afp_empl_val = pl.get("afpEmployer", 0)
            sfs_empl_val = pl.get("sfsEmployer", 0)
            srl_val = pl.get("srlEmployer", 0)
            infotep_val = pl.get("infotepEmployer", 0)

            # Otras remuneraciones = comisiones + bonos + otros ingresos
            otras_rem = (pl.get("commission", 0) + pl.get("bonus", 0) +
                        pl.get("otherIncome", 0))

            row_data = [
                "",  # A: empty
                emp.get("tssKey", ""),              # B: Clave Nómina
                tipo_doc,                            # C: Tipo Doc.
                emp.get("cedula", "") or emp.get("idNumber", ""),  # D: Documento
                nombres,                             # E: Nombres
                emp.get("firstLastName", "") or emp.get("lastName", ""),  # F: 1er Apellido
                emp.get("secondLastName", ""),       # G: 2do Apellido
                sexo,                                # H: Sexo
                birth,                               # I: Fecha Nacimiento
                salario_cotizable,                   # J: Salario Cotizable
                0.0,                                 # K: Aporte Voluntario
                total_income,                        # L: Salario ISR
                "01",                                # M: Tipo Ingreso (01=Normal)
                otras_rem,                           # N: Otras Remuneraciones
                "",                                  # O: RNC/Céd. Agente Ret
                0.0,                                 # P: Remun. Otros Agentes
                total_income,                        # Q: Remuneración del período
                0.0,                                 # R: Saldo a favor
                0.0,                                 # S: Regalía Pascual
                0.0,                                 # T: Preaviso/Cesantía/Indemniz.
                0.0,                                 # U: Retención Pensión Alimenticia
                total_income,                        # V: Salario INFOTEP
            ]

            for col_idx, val in enumerate(row_data):
                cell = ws.cell(row=row, column=col_idx + 1, value=val)
                cell.border = thin_border
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'

            total_afp_emp += afp_emp_val
            total_sfs_emp += sfs_emp_val
            total_afp_empl += afp_empl_val
            total_sfs_empl += sfs_empl_val
            total_srl += srl_val
            total_infotep += infotep_val

        # Actualizar conteo
        ws["C5"] = emp_count

        # ── Save to bytes ──
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"TSS_Autodeterminacion_{periodo_label.replace(' ', '_')}.xlsx"

        return {
            "content": output.getvalue(),
            "filename": filename,
            "periodo": periodo_label,
            "periodo_tss": periodo_mmaaaa,
            "total_empleados": emp_count,
            "resumen": {
                "afp_empleado": round(total_afp_emp, 2),
                "sfs_empleado": round(total_sfs_emp, 2),
                "afp_empleador": round(total_afp_empl, 2),
                "sfs_empleador": round(total_sfs_empl, 2),
                "srl": round(total_srl, 2),
                "infotep": round(total_infotep, 2),
                "total": round(total_afp_emp + total_sfs_emp + total_afp_empl +
                              total_sfs_empl + total_srl + total_infotep, 2),
            },
        }

    @classmethod
    def _split_name(cls, emp: dict) -> dict:
        """Separa nombres y apellidos para el formato TSS."""
        primer_nombre = (emp.get("firstName", "") or "").strip()
        segundo_nombre = (emp.get("middleName", "") or "").strip()
        primer_apellido = (emp.get("firstLastName", "") or emp.get("lastName", "") or "").strip()
        segundo_apellido = (emp.get("secondLastName", "") or "").strip()

        # Fallback: si no hay nombres separados, intentar separar fullName
        if not primer_nombre and not primer_apellido:
            full = (emp.get("fullName", "") or "").strip()
            parts = full.split()
            if len(parts) >= 2:
                primer_nombre = parts[0]
                primer_apellido = parts[-1]
                if len(parts) >= 3:
                    segundo_nombre = parts[1] if len(parts) > 2 else ""
                    if len(parts) >= 4:
                        segundo_apellido = parts[-2]

        return {
            "primer_nombre": primer_nombre,
            "segundo_nombre": segundo_nombre,
            "primer_apellido": primer_apellido,
            "segundo_apellido": segundo_apellido,
        }

    @classmethod
    def _calcular_estado_tss(cls, hire_date_str: str, termination_date_str: str,
                              year: int, month: int) -> str:
        """
        Determina el estado TSS del empleado para el período:
          I = Ingreso (fecha de entrada dentro del mes)
          P = Permanencia (activo, entró antes del mes)
          E = Egreso (fecha de salida dentro del mes)
        """
        from datetime import datetime

        # Primer día del período
        try:
            periodo_inicio = datetime(year, month, 1).date()
            if month == 12:
                periodo_fin = datetime(year + 1, 1, 1).date()
            else:
                periodo_fin = datetime(year, month + 1, 1).date()
        except (ValueError, TypeError):
            return "P"

        try:
            if hire_date_str:
                hd = datetime.strptime(hire_date_str[:10], "%Y-%m-%d").date()
                # Si entró en este mes, es Ingreso
                if periodo_inicio <= hd < periodo_fin:
                    return "I"

            if termination_date_str:
                td = datetime.strptime(termination_date_str[:10], "%Y-%m-%d").date()
                # Si salió en este mes, es Egreso
                if periodo_inicio <= td < periodo_fin:
                    return "E"
        except (ValueError, TypeError):
            pass

        return "P"
