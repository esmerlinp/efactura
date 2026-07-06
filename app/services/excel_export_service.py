import io
from datetime import datetime


class ExcelExportService:
    @staticmethod
    def _create_workbook():
        import openpyxl
        wb = openpyxl.Workbook()
        return wb

    @staticmethod
    def _style_header(ws, headers: list, row: int = 1):
        import openpyxl.styles as styles

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = styles.Font(bold=True, size=11, color="FFFFFF")
            cell.fill = styles.PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
            cell.alignment = styles.Alignment(horizontal="center", vertical="center")
            cell.border = styles.Border(
                bottom=styles.Side(style="thin", color="CBD5E1"),
            )

    @staticmethod
    def _style_data_cell(cell, is_currency: bool = False, is_number: bool = False):
        import openpyxl.styles as styles

        cell.font = styles.Font(size=10)
        cell.border = styles.Border(
            bottom=styles.Side(style="hair", color="E2E8F0"),
        )
        if is_currency:
            cell.number_format = '#,##0.00'
            cell.alignment = styles.Alignment(horizontal="right")
        elif is_number:
            cell.alignment = styles.Alignment(horizontal="center")

    @staticmethod
    def _auto_width(ws, min_width: int = 10, max_width: int = 50):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    @staticmethod
    def _write_currency_row(ws, row: int, data: list):
        import openpyxl.styles as styles

        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if isinstance(val, (int, float)) and col_idx > 1:
                ExcelExportService._style_data_cell(cell, is_currency=True)
            elif isinstance(val, (int, float)):
                ExcelExportService._style_data_cell(cell, is_number=True)
            else:
                ExcelExportService._style_data_cell(cell)

    @classmethod
    def export_trial_balance(cls, accounts: list, entries: list, period_label: str = "") -> io.BytesIO:
        wb = cls._create_workbook()
        ws = wb.active
        ws.title = "Balanza de Comprobación"

        if period_label:
            ws.cell(row=1, column=1, value=f"Balanza de Comprobación — {period_label}")
            ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

        headers = ["Código", "Cuenta", "Débito", "Crédito", "Saldo"]
        cls._style_header(ws, headers, row=3)

        row = 4
        total_debit = 0.0
        total_credit = 0.0

        for acc in sorted(accounts, key=lambda a: a.get("code", "")):
            code = acc.get("code", "")
            name = acc.get("name", "")
            if acc.get("type") != "movimiento":
                continue

            debit = sum(float(l.get("debit", 0)) for e in entries if e.get("status") != "voided"
                       for l in e.get("lines", []) if l.get("accountId") == acc.get("id"))
            credit = sum(float(l.get("credit", 0)) for e in entries if e.get("status") != "voided"
                        for l in e.get("lines", []) if l.get("accountId") == acc.get("id"))
            balance = round(debit - credit, 2)
            total_debit += debit
            total_credit += credit

            if abs(debit) < 0.01 and abs(credit) < 0.01:
                continue

            cls._write_currency_row(ws, row, [code, name, round(debit, 2), round(credit, 2), balance])
            row += 1

        import openpyxl.styles as styles
        for col_idx, val in enumerate(["", "TOTAL", round(total_debit, 2), round(total_credit, 2), ""], 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = styles.Font(bold=True, size=11)
            cell.border = styles.Border(top=styles.Side(style="double", color="2D3748"))

        cls._auto_width(ws)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    def export_aging_report(cls, data: list, report_type: str = "CxC") -> io.BytesIO:
        wb = cls._create_workbook()
        ws = wb.active
        ws.title = f"Aging {report_type}"

        ws.cell(row=1, column=1, value=f"Reporte de Antigüedad — {report_type}")
        ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

        headers = ["Cliente/Proveedor", "Documento", "Fecha", "Vencimiento", "Total", "Sin Vencer", "1-30 días", "31-60 días", "61-90 días", "91+ días"]
        cls._style_header(ws, headers, row=3)

        row = 4
        for item in data:
            cls._write_currency_row(ws, row, [
                item.get("name", ""),
                item.get("document_number", ""),
                item.get("date", ""),
                item.get("due_date", ""),
                item.get("total", 0.0),
                item.get("sin_vencer", 0.0),
                item.get("dias_1_30", 0.0),
                item.get("dias_31_60", 0.0),
                item.get("dias_61_90", 0.0),
                item.get("dias_91_plus", 0.0),
            ])
            row += 1

        cls._auto_width(ws)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    def export_financial_statement(cls, sections: list, title: str) -> io.BytesIO:
        wb = cls._create_workbook()
        ws = wb.active
        ws.title = title[:31]

        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=16)

        headers = ["Cuenta", "Monto (RD$)"]
        cls._style_header(ws, headers, row=3)

        row = 4
        for section in sections:
            section_name = section.get("name", "")
            if section_name:
                cell = ws.cell(row=row, column=1, value=section_name)
                cell.font = openpyxl.styles.Font(bold=True, size=11, color="2D3748")
                cell.fill = openpyxl.styles.PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                row += 1

            for line in section.get("lines", []):
                cls._write_currency_row(ws, row, [
                    f"  {line.get('name', '')}",
                    line.get("amount", 0.0),
                ])
                row += 1

            if section.get("total"):
                total_row = row
                for col_idx, val in enumerate([f"Total {section_name}", section["total"]], 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = openpyxl.styles.Font(bold=True, size=11)
                    cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style="thin", color="2D3748"))

                row += 1

        cls._auto_width(ws)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
