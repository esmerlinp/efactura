"""RRHH module — auto-extracted."""

from datetime import date
from flask import render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
from app.web.rrhh import (
    web_rrhh_bp, _get_owner_uid_and_sandbox, _login_required,
    _is_hr_role, _sanitize_for_role, MONTHS_ES,
    _filter_employees_by_period, _generate_periods,
)
from app.web.invoices import web_invoices_bp
from app.utils.module_gate import require_module
from app.services import hr_data_service as hr
from app.utils.hr_utils import is_active_equivalent
from app.services.payroll_ytd_service import get_ytd
from app.services.payroll_service import PayrollService
import csv, io


def _projection_filters(employees):
    """Lee filtros comunes y devuelve opciones para todas las proyecciones."""
    return {
        "employee_id": request.args.get("employee_id", ""),
        "department": request.args.get("department", ""),
        "area": request.args.get("area", ""),
        "group_id": request.args.get("group_id", ""),
        "status": request.args.get("status", ""),
        "include_inactive": request.args.get("include_inactive") == "1",
        "employees": employees,
        "departments": sorted({e.get("department", e.get("area", "General")) for e in employees if e.get("department") or e.get("area")}),
        "areas": sorted({e.get("area", "") for e in employees if e.get("area")}),
        "groups": [],
        "statuses": ["activo", "vacaciones", "licencia", "suspendido", "inactivo"],
    }


def _projection_groups(company_id, sandbox):
    try:
        return hr.get_payroll_groups(company_id, sandbox=sandbox)
    except (AttributeError, TypeError):
        return []


def _projection_year():
    try:
        return int(request.args.get("year", date.today().year))
    except (TypeError, ValueError):
        return date.today().year


def _projection_csv(title, headers, rows, total_row):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([title])
    writer.writerow(headers)
    writer.writerows(rows)
    writer.writerow([])
    writer.writerow(total_row)
    result = io.BytesIO(b"\xef\xbb\xbf" + output.getvalue().encode("utf-8"))
    result.seek(0)
    return result


def _projection_pdf(template, data, filename):
    from app.utils.pdf import pdf_write_options
    from weasyprint import HTML as WeasyprintHTML
    data["is_pdf"] = True
    rendered = render_template(template, active_page="rrhh_reports", **data)
    pdf_bytes = WeasyprintHTML(string=rendered, base_url=request.host_url).write_pdf(**pdf_write_options())
    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _projection_data(kind):
    from app.services import hr_data_service as hr
    from app.services.payroll_projection_service import PayrollProjectionService
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    employees = hr.get_employees(company_id, sandbox=sandbox)
    filters = _projection_filters(employees)
    filters["groups"] = _projection_groups(company_id, sandbox)
    year = _projection_year()
    if kind == "benefits":
        cutoff = request.args.get("cutoff_date", f"{year}-12-31")
        result = PayrollProjectionService.project_benefits(employees, cutoff, company_id=company_id, sandbox=sandbox, **{k: filters[k] for k in ("employee_id", "department", "area", "group_id", "status", "include_inactive")})
        result.update(filters=filters, year=year, cutoff_date=cutoff)
        return result
    result = PayrollProjectionService.project_payroll(employees, year, tax_rates=hr.get_tax_rates(company_id, sandbox=sandbox), **{k: filters[k] for k in ("employee_id", "department", "area", "group_id", "status", "include_inactive")})
    result["filters"] = filters
    return result


def _projection_view(kind, title, template):
    data = _projection_data(kind)
    return render_template(template, active_page="rrhh_reports", title=title, **data)


@web_invoices_bp.route('/reports/rrhh/payroll-projection')
@web_rrhh_bp.route('/rrhh/reports/payroll-projection')
@require_module('nomina')
def report_payroll_projection():
    if _login_required(): return redirect(url_for('web_auth.login'))
    return _projection_view('payroll', 'Proyección anual de nómina', 'rrhh/reports/projection.html')


@web_invoices_bp.route('/reports/rrhh/benefits-projection')
@web_rrhh_bp.route('/rrhh/reports/benefits-projection')
@require_module('nomina')
def report_benefits_projection():
    if _login_required(): return redirect(url_for('web_auth.login'))
    return _projection_view('benefits', 'Proyección de prestaciones laborales', 'rrhh/reports/benefits_projection.html')


def _contribution_projection(kind, title):
    data = _projection_data('payroll')
    metric = 'afpEmployee' if kind == 'afp' else 'sfsEmployee' if kind == 'sfs' else 'isrRetention'
    employer = 'afpEmployer' if kind == 'afp' else 'sfsEmployer' if kind == 'sfs' else None
    data['metric'] = metric
    data['employer_metric'] = employer
    data['projection_title'] = title
    return render_template('rrhh/reports/contribution_projection.html', active_page='rrhh_reports', **data)


@web_invoices_bp.route('/reports/rrhh/afp-projection')
@web_rrhh_bp.route('/rrhh/reports/afp-projection')
@require_module('nomina')
def report_afp_projection():
    if _login_required(): return redirect(url_for('web_auth.login'))
    return _contribution_projection('afp', 'Proyección anual de AFP')


@web_invoices_bp.route('/reports/rrhh/sfs-projection')
@web_rrhh_bp.route('/rrhh/reports/sfs-projection')
@require_module('nomina')
def report_sfs_projection():
    if _login_required(): return redirect(url_for('web_auth.login'))
    return _contribution_projection('sfs', 'Proyección anual de SFS')


@web_invoices_bp.route('/reports/rrhh/isr-projection')
@web_rrhh_bp.route('/rrhh/reports/isr-projection')
@require_module('nomina')
def report_isr_projection():
    if _login_required(): return redirect(url_for('web_auth.login'))
    return _contribution_projection('isr', 'Proyección anual de ISR')


@web_invoices_bp.route('/reports/rrhh/projection/export')
@web_rrhh_bp.route('/rrhh/reports/projection/export')
@require_module('nomina')
def report_projection_export():
    if _login_required(): return redirect(url_for('web_auth.login'))
    kind = request.args.get('kind', 'payroll')
    fmt = request.args.get('format', 'csv')
    data = _projection_data(kind)
    if kind == 'benefits':
        headers = ['Empleado', 'Cédula', 'Departamento', 'Salario promedio', 'Preaviso', 'Cesantía', 'Vacaciones', 'Salario Navidad', 'Salario proporcional', 'Descuentos', 'Total', 'Neto a pagar']
        rows = [[r['employeeName'], r['cedula'], r['department'], r['salarioPromedio'], r['preaviso'], r['cesantia'], r['vacaciones'], r['salarioNavidad'], r['salarioProporcional'], r['descuentos'], r['total'], r['netoAPagar']] for r in data['rows']]
        total = ['TOTAL', '', '', '', '', '', '', '', '', '', data['total'], data['totalNeto']]
        filename = 'proyeccion_prestaciones'
    else:
        headers = ['Mes', 'Bruto', 'AFP empleado', 'SFS empleado', 'ISR', 'Neto', 'Aportes patronales', 'Costo total']
        rows = [[r['month'], r['totalIncome'], r['afpEmployee'], r['sfsEmployee'], r['isrRetention'], r['netSalary'], r['totalEmployerContrib'], r['totalCost']] for r in data['months']]
        total = ['TOTAL', data['totals']['totalIncome'], data['totals']['afpEmployee'], data['totals']['sfsEmployee'], data['totals']['isrRetention'], data['totals']['netSalary'], data['totals']['totalEmployerContrib'], data['totals']['totalCost']]
        filename = 'proyeccion_nomina'
    if fmt == 'pdf':
        template = 'rrhh/reports/benefits_projection.html' if kind == 'benefits' else 'rrhh/reports/projection.html'
        return _projection_pdf(template, data, f'{filename}_{data.get("year", "")}.pdf')
    return send_file(_projection_csv(filename, headers, rows, total), mimetype='text/csv', as_attachment=True, download_name=f'{filename}_{data.get("year", "")}.csv')


def _enrich_periods(periods, owner_uid, sandbox):
    """Inyecta líneas desde subcolección a cada período para compatibilidad con templates."""
    for p in periods:
        p["lines"] = PayrollService.get_period_lines(p, company_id=company_id, sandbox=sandbox)
    return periods


def _enrich_period(period, owner_uid, sandbox):
    """Inyecta líneas desde subcolección a un período."""
    if period:
        period["lines"] = PayrollService.get_period_lines(period, company_id=company_id, sandbox=sandbox)
    return period


# ═══════════════════════════════════════════════════════════════════════════
# REPORTES
# ═══════════════════════════════════════════════════════════════════════════

@web_invoices_bp.route('/reports/rrhh/ir18')
@web_rrhh_bp.route("/rrhh/reports/ir18")
@require_module('nomina')
def report_ir18_list():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr
    from app.services.payroll_ytd_service import get_ytd

    employees = hr.get_employees(company_id, sandbox=sandbox)
    try:
        year = int(request.args.get("year", date.today().year))
    except (ValueError, TypeError):
        year = date.today().year

    employee_ytds = []
    for emp in employees:
        if not is_active_equivalent(emp.get("status", "")):
            continue
        ytd = get_ytd(company_id, emp["id"], year, sandbox=sandbox)
        if ytd.get("grossIncome", 0) > 0:
            employee_ytds.append({
                "employee": emp,
                "ytd": ytd,
            })

    return render_template("rrhh/reports/ir18_list.html", active_page="rrhh_reports",
                           employee_ytds=employee_ytds, year=year)


@web_invoices_bp.route('/reports/rrhh/ir18/<employee_id>')
@web_rrhh_bp.route("/rrhh/reports/ir18/<employee_id>")
@require_module('nomina')
def report_ir18_view(employee_id):
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr
    from app.services.payroll_ytd_service import get_ytd

    employee = hr.get_employee(company_id, employee_id, sandbox=sandbox)
    if not employee:
        flash("Empleado no encontrado.", "error")
        return redirect(url_for("web_invoices.report_ir18_list"))

    try:
        year = int(request.args.get("year", date.today().year))
    except (ValueError, TypeError):
        year = date.today().year

    ytd = get_ytd(company_id, employee_id, year, sandbox=sandbox)
    return render_template("rrhh/reports/ir18_detail.html", active_page="rrhh_reports",
                           employee=employee, ytd=ytd, year=year, today=date.today())


@web_invoices_bp.route('/reports/rrhh')
@web_rrhh_bp.route("/rrhh/reports")
@require_module('nomina')
def reports_index():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    # Legacy RRHH reports landing page: keep old URLs working while using the
    # unified reports category navigation.
    return redirect(url_for("web_invoices.reports_category", category_key="nomina"))


@web_invoices_bp.route('/reports/rrhh/department')
@web_rrhh_bp.route("/rrhh/reports/department")
@require_module('nomina')
def report_department():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr
    periods = hr.get_payroll_periods(company_id, sandbox=sandbox)
    periods = _enrich_periods(periods, owner_uid, sandbox)
    periods = _enrich_periods(periods, owner_uid, sandbox)
    period_key = request.args.get("period", "")
    period = next((p for p in periods if p.get("periodKey") == period_key), None) if period_key else None
    by_dept = {}
    if period:
        for l in period.get("lines", []):
            dept = l.get("department", "Sin depto")
            if dept not in by_dept:
                by_dept[dept] = {"count": 0, "gross": 0.0, "net": 0.0, "employer": 0.0}
            by_dept[dept]["count"] += 1
            by_dept[dept]["gross"] += l.get("totalIncome", 0)
            by_dept[dept]["net"] += l.get("netSalary", 0)
            by_dept[dept]["employer"] += l.get("totalEmployerContrib", 0)
    period_keys = sorted(set(p.get("periodKey", "") for p in periods), reverse=True)
    return render_template("rrhh/reports/department.html", active_page="rrhh_reports",
                           by_dept=by_dept, period_keys=period_keys, selected=period_key)


@web_invoices_bp.route('/reports/rrhh/tss')
@web_rrhh_bp.route("/rrhh/reports/tss")
@require_module('nomina')
def report_tss():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr
    periods = hr.get_payroll_periods(company_id, sandbox=sandbox)
    periods = _enrich_periods(periods, owner_uid, sandbox)
    try:
        year = int(request.args.get("year", date.today().year))
    except (ValueError, TypeError):
        year = date.today().year
    monthly = {}
    for m in range(1, 13):
        key = f"{year}-{m:02d}"
        monthly[key] = {"afp_emp": 0, "sfs_emp": 0, "afp_empl": 0, "sfs_empl": 0, "srl": 0, "infotep": 0}
    for p in periods:
        pk = p.get("periodKey", "")
        if str(year) not in pk:
            continue
        base_key = pk[:7] if len(pk) >= 7 else pk
        if base_key in monthly:
            for l in p.get("lines", []):
                monthly[base_key]["afp_emp"] += l.get("afpEmployee", 0)
                monthly[base_key]["sfs_emp"] += l.get("sfsEmployee", 0)
                monthly[base_key]["afp_empl"] += l.get("afpEmployer", 0)
                monthly[base_key]["sfs_empl"] += l.get("sfsEmployer", 0)
                monthly[base_key]["srl"] += l.get("srlEmployer", 0)
                monthly[base_key]["infotep"] += l.get("infotepEmployer", 0)
    return render_template("rrhh/reports/tss.html", active_page="rrhh_reports",
                           monthly=monthly, year=year, months_es=MONTHS_ES)


@web_invoices_bp.route('/reports/rrhh/comparative')
@web_rrhh_bp.route("/rrhh/reports/comparative")
@require_module('nomina')
def report_comparative():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr
    periods = hr.get_payroll_periods(company_id, sandbox=sandbox)
    periods = _enrich_periods(periods, owner_uid, sandbox)
    periods.sort(key=lambda p: p.get("periodKey", ""), reverse=True)
    p1_key = request.args.get("p1", "")
    p2_key = request.args.get("p2", "")
    p1 = next((p for p in periods if p.get("periodKey") == p1_key), None) if p1_key else None
    p2 = next((p for p in periods if p.get("periodKey") == p2_key), None) if p2_key else None
    comparison = None
    if p1 and p2:
        def pt(p):
            return {
                "gross": round(sum(l.get("totalIncome", 0) for l in p.get("lines", [])), 2),
                "net": round(sum(l.get("netSalary", 0) for l in p.get("lines", [])), 2),
                "employer": round(sum(l.get("totalEmployerContrib", 0) for l in p.get("lines", [])), 2),
                "count": len(p.get("lines", [])),
            }
        t1 = pt(p1)
        t2 = pt(p2)
        comparison = {
            "p1_label": p1.get("periodRange") or p1.get("periodKey"),
            "p2_label": p2.get("periodRange") or p2.get("periodKey"),
            "count_diff": t1["count"] - t2["count"],
            "gross_diff": round(t1["gross"] - t2["gross"], 2),
            "gross_pct": round((t1["gross"] - t2["gross"]) / t2["gross"] * 100, 1) if t2["gross"] else 0,
            "net_diff": round(t1["net"] - t2["net"], 2),
            "net_pct": round((t1["net"] - t2["net"]) / t2["net"] * 100, 1) if t2["net"] else 0,
            "employer_diff": round(t1["employer"] - t2["employer"], 2),
            "t1": t1, "t2": t2,
        }
    period_keys = [p.get("periodKey", "") for p in periods]
    return render_template("rrhh/reports/comparative.html", active_page="rrhh_reports",
                           period_keys=period_keys, comparison=comparison,
                           p1_key=p1_key, p2_key=p2_key)


# ═══════════════════════════════════════════════════════════════════════════
# REPORTE: NÓMINA NETA SIN PROVISIONES
# ═══════════════════════════════════════════════════════════════════════════

@web_invoices_bp.route('/reports/rrhh/net-payroll')
@web_rrhh_bp.route("/rrhh/reports/net-payroll")
@require_module('nomina')
def report_net_payroll():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr

    periods = hr.get_payroll_periods(company_id, sandbox=sandbox)
    periods = _enrich_periods(periods, owner_uid, sandbox)
    periods.sort(key=lambda p: p.get("periodKey", ""), reverse=True)

    # Filtro por período
    period_key = request.args.get("period", "")
    selected = None
    if period_key:
        selected = next((p for p in periods if p.get("periodKey") == period_key), None)

    period_keys = [p.get("periodKey", "") for p in periods]
    lines = selected.get("lines", []) if selected else []

    return render_template("rrhh/reports/net_payroll.html", active_page="rrhh_reports",
                           period_keys=period_keys, selected=selected,
                           lines=lines, period_key=period_key)


@web_invoices_bp.route('/reports/rrhh/net-payroll/export')
@web_rrhh_bp.route("/rrhh/reports/net-payroll/export")
@require_module('nomina')
def report_net_payroll_export():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr

    period_key = request.args.get("period", "")
    period = hr.get_payroll_period_by_key(company_id, period_key, sandbox=sandbox)
    period = _enrich_period(period, owner_uid, sandbox)
    if not period:
        flash("Período no encontrado.", "error")
        return redirect(url_for("web_invoices.report_net_payroll"))

    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Empleado", "Cedula", "Cargo", "Salario Base", "Horas Extra",
                      "Comisiones", "Total Bruto", "Otras Ded.", "Neto sin Provisiones"])
    for l in period.get("lines", []):
        neto_sin_provisiones = round(l.get("netSalary", 0) + l.get("afpEmployee", 0) +
                                     l.get("sfsEmployee", 0) + l.get("infotepEmployee", 0) +
                                     l.get("isrRetention", 0), 2)
        writer.writerow([
            l.get("employeeName", ""), l.get("cedula", ""), l.get("position", ""),
            f"{l.get('baseSalary', 0):.2f}", f"{l.get('overtimeHours', 0):.2f}",
            f"{l.get('commission', 0):.2f}", f"{l.get('totalIncome', 0):.2f}",
            f"{l.get('otherDeductions', 0):.2f}", f"{neto_sin_provisiones:.2f}",
        ])
    dest = io.BytesIO()
    dest.write(b"\xef\xbb\xbf")
    dest.write(output.getvalue().encode("utf-8"))
    dest.seek(0)
    period_label = period_key or "nomina"
    filename = f"nomina_neta_{period_label}.csv"
    return send_file(dest, mimetype="text/csv", as_attachment=True, download_name=filename)


# ═══════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN CSV DE NÓMINA GENERAL CONSOLIDADA
# ═══════════════════════════════════════════════════════════════════════════

@web_rrhh_bp.route("/rrhh/payroll/<period_id>/export-csv")
def payroll_export_csv(period_id):
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr

    period = hr.get_payroll_period(company_id, period_id, sandbox=sandbox)
    period = _enrich_period(period, owner_uid, sandbox)
    if not period:
        flash("Período no encontrado.", "error")
        return redirect(url_for("web_rrhh.payroll_list"))

    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Empleado", "Cedula", "Cargo", "Salario Base", "Bruto",
                      "AFP Emp.", "SFS Emp.", "INFOTEP Emp.", "ISR", "Otras Ded.",
                      "Neto", "AFP Empl.", "SFS Empl.", "SRL", "INFOTEP Empl.",
                      "Aportes Empleador"])
    for l in period.get("lines", []):
        writer.writerow([
            l.get("employeeName", ""), l.get("cedula", ""), l.get("position", ""),
            f"{l.get('baseSalary', 0):.2f}", f"{l.get('totalIncome', 0):.2f}",
            f"{l.get('afpEmployee', 0):.2f}", f"{l.get('sfsEmployee', 0):.2f}",
            f"{l.get('infotepEmployee', 0):.2f}", f"{l.get('isrRetention', 0):.2f}",
            f"{l.get('otherDeductions', 0):.2f}", f"{l.get('netSalary', 0):.2f}",
            f"{l.get('afpEmployer', 0):.2f}", f"{l.get('sfsEmployer', 0):.2f}",
            f"{l.get('srlEmployer', 0):.2f}", f"{l.get('infotepEmployer', 0):.2f}",
            f"{l.get('totalEmployerContrib', 0):.2f}",
        ])
    # Totales
    writer.writerow([])
    writer.writerow(["TOTALES", "", "", "",
                     f"{period.get('totalGross', 0):.2f}", "", "", "", "", "",
                     f"{period.get('totalNet', 0):.2f}", "", "", "", "",
                     f"{period.get('totalEmployerContrib', 0):.2f}"])

    dest = io.BytesIO()
    dest.write(b"\xef\xbb\xbf")
    dest.write(output.getvalue().encode("utf-8"))
    dest.seek(0)
    period_label = period.get("periodKey", "nomina")
    filename = f"nomina_general_{period_label}.csv"
    return send_file(dest, mimetype="text/csv", as_attachment=True, download_name=filename)


# ═══════════════════════════════════════════════════════════════════════════
# REPORTE: RETENCIONES ISR NÓMINA (suma quincenas para DGII)
# ═══════════════════════════════════════════════════════════════════════════

@web_invoices_bp.route('/reports/rrhh/isr-retentions')
@web_rrhh_bp.route("/rrhh/reports/isr-retentions")
@require_module('nomina')
def report_isr_retentions():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr

    periods = hr.get_payroll_periods(company_id, sandbox=sandbox)
    periods = _enrich_periods(periods, owner_uid, sandbox)
    try:
        year = int(request.args.get("year", date.today().year))
    except (ValueError, TypeError):
        year = date.today().year

    # Agrupar por mes: sumar quincenas si existen
    monthly = {}
    for m in range(1, 13):
        key = f"{year}-{m:02d}"
        monthly[key] = {"isr": 0.0, "afp_emp": 0.0, "sfs_emp": 0.0, "employees": 0, "lines": []}

    for p in periods:
        pk = p.get("periodKey", "")
        if str(year) not in pk:
            continue
        base_key = pk[:7]
        if base_key in monthly:
            for l in p.get("lines", []):
                monthly[base_key]["isr"] += l.get("isrRetention", 0)
                monthly[base_key]["afp_emp"] += l.get("afpEmployee", 0)
                monthly[base_key]["sfs_emp"] += l.get("sfsEmployee", 0)
                monthly[base_key]["employees"] += 1
                monthly[base_key]["lines"].append(l)

    return render_template("rrhh/reports/isr_retentions.html", active_page="rrhh_reports",
                           monthly=monthly, year=year, months_es=MONTHS_ES)


@web_invoices_bp.route('/reports/rrhh/isr-retentions/export')
@web_rrhh_bp.route("/rrhh/reports/isr-retentions/export")
@require_module('nomina')
def report_isr_retentions_export():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr

    periods = hr.get_payroll_periods(company_id, sandbox=sandbox)
    periods = _enrich_periods(periods, owner_uid, sandbox)
    try:
        year = int(request.args.get("year", date.today().year))
        month = int(request.args.get("month", date.today().month))
    except (ValueError, TypeError):
        year = date.today().year
        month = date.today().month

    # Agrupar para el mes seleccionado
    base_key = f"{year}-{month:02d}"
    isr_lines = []
    for p in periods:
        pk = p.get("periodKey", "")
        if pk[:7] == base_key:
            for l in p.get("lines", []):
                if l.get("isrRetention", 0) > 0:
                    isr_lines.append(l)

    import csv, io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["RNC Agente", "Período", "Cédula Retenido", "Nombre Retenido",
                      "ISR Retenido", "Salario Bruto", "Período Nómina"])
    period_label = f"{year:04d}{month:02d}"
    for l in isr_lines:
        writer.writerow([
            "", period_label, l.get("cedula", ""), l.get("employeeName", ""),
            f"{l.get('isrRetention', 0):.2f}", f"{l.get('totalIncome', 0):.2f}",
            l.get("periodType", ""),
        ])
    writer.writerow([])
    total_isr = sum(l.get("isrRetention", 0) for l in isr_lines)
    writer.writerow(["TOTAL", "", "", "", f"{total_isr:.2f}", "", ""])

    dest = io.BytesIO()
    dest.write(b"\xef\xbb\xbf")
    dest.write(output.getvalue().encode("utf-8"))
    dest.seek(0)
    filename = f"isr_retenciones_{period_label}.csv"
    return send_file(dest, mimetype="text/csv", as_attachment=True, download_name=filename)


# ═══════════════════════════════════════════════════════════════════════════
# WHAT-IF ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════

@web_invoices_bp.route('/reports/rrhh/what-if')
@web_rrhh_bp.route("/rrhh/reports/what-if")
@require_module('nomina')
def report_what_if():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr
    from app.services.payroll_service import PayrollService

    scenario_type = request.args.get("type", "")
    value = float(request.args.get("value", 0) or 0)
    filter_dept = request.args.get("department", "")
    filter_area = request.args.get("area", "")

    all_employees = hr.get_employees(company_id, sandbox=sandbox)
    tax_rates = hr.get_tax_rates(company_id, sandbox=sandbox)
    result = None
    departments = sorted(set(e.get("department", e.get("area", "General")) for e in all_employees if e.get("department") or e.get("area")))

    if scenario_type and value:
        scenario = {
            "type": scenario_type,
            "value": value,
            "filter_department": filter_dept,
            "filter_area": filter_area,
        }
        result = PayrollService.what_if_analysis(all_employees, scenario, tax_rates=tax_rates)

    return render_template("rrhh/reports/what_if.html", active_page="rrhh_reports",
                           result=result, departments=departments,
                           scenario_type=scenario_type, value=value,
                           filter_dept=filter_dept, filter_area=filter_area)


# ═══════════════════════════════════════════════════════════════════════════
# RETROACTIVE PAY CALCULATOR
# ═══════════════════════════════════════════════════════════════════════════

@web_rrhh_bp.route("/rrhh/employees/<employee_id>/retroactive", methods=["GET", "POST"])
def employee_retroactive_pay(employee_id):
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services import hr_data_service as hr
    from app.services.payroll_service import PayrollService

    employee = hr.get_employee(company_id, employee_id, sandbox=sandbox)
    if not employee:
        flash("Empleado no encontrado.", "error")
        return redirect(url_for("web_rrhh.employee_list"))

    result = None
    if request.method == "POST":
        new_salary = float(request.form.get("new_salary", 0) or 0)
        effective_date = request.form.get("effective_date", "")
        tax_rates = hr.get_tax_rates(company_id, sandbox=sandbox)
        salary_history = hr.get_salary_history(company_id, employee_id, sandbox=sandbox)

        result = PayrollService.calculate_retroactive_pay(
            employee, salary_history, new_salary, effective_date, tax_rates=tax_rates
        )

    salary_history = hr.get_salary_history(company_id, employee_id, sandbox=sandbox)
    return render_template("rrhh/employee_retroactive.html", active_page="rrhh_employees",
                           employee=employee, result=result, salary_history=salary_history)


# ═══════════════════════════════════════════════════════════════════════════
# VALIDACIÓN IR-18
# ═══════════════════════════════════════════════════════════════════════════

@web_invoices_bp.route('/reports/rrhh/ir18/validation')
@web_rrhh_bp.route("/rrhh/reports/ir18/validation")
@require_module('nomina')
def report_ir18_validation():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services.payroll_service import PayrollService
    from datetime import date

    year = int(request.args.get("year", date.today().year))
    result = PayrollService.validate_ir18_readiness(company_id, year=year, sandbox=sandbox)

    return render_template("rrhh/reports/ir18_validation.html", active_page="rrhh_reports",
                           result=result, year=year)


# ═══════════════════════════════════════════════════════════════════════════
# IR-13 — DECLARACIÓN JURADA ANUAL DEL AGENTE DE RETENCIÓN DE ASALARIADOS
# ═══════════════════════════════════════════════════════════════════════════

@web_invoices_bp.route('/reports/rrhh/ir13')
@web_rrhh_bp.route("/rrhh/reports/ir13")
@require_module('nomina')
def report_ir13():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services.ir13_service import calculate_ir13
    from app.services.db_service import DatabaseService

    try:
        year = int(request.args.get("year", date.today().year))
    except (ValueError, TypeError):
        year = date.today().year

    data = calculate_ir13(company_id, year=year, sandbox=sandbox)

    company = DatabaseService.get_company_profile(owner_uid, company_id=company_id) or {}
    data["company"] = company
    data["fecha_emision"] = date.today().strftime("%d/%m/%Y")

    return render_template("rrhh/reports/ir13.html", active_page="rrhh_reports", **data)


@web_invoices_bp.route('/reports/rrhh/ir13/pdf')
@web_rrhh_bp.route("/rrhh/reports/ir13/pdf")
@require_module('nomina')
def report_ir13_pdf():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services.ir13_service import calculate_ir13
    from app.services.db_service import DatabaseService
    from app.utils.pdf import pdf_write_options
    from weasyprint import HTML as WeasyprintHTML

    try:
        year = int(request.args.get("year", date.today().year))
    except (ValueError, TypeError):
        year = date.today().year

    data = calculate_ir13(company_id, year=year, sandbox=sandbox)

    company = DatabaseService.get_company_profile(owner_uid, company_id=company_id) or {}
    data["company"] = company
    data["fecha_emision"] = date.today().strftime("%d/%m/%Y")
    data["is_pdf"] = True

    rendered = render_template("rrhh/reports/ir13.html", active_page="rrhh_reports", **data)
    pdf_bytes = WeasyprintHTML(string=rendered, base_url=request.host_url).write_pdf(**pdf_write_options())

    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="IR-13_{year}.pdf"'
    return response


# ═══════════════════════════════════════════════════════════════
#  Reporte: Listado de Aniversarios
# ═══════════════════════════════════════════════════════════════

MESES_ES_MAP = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _build_anniversary_data(company_id, sandbox, owner_uid):
    from app.services.db_service import DatabaseService

    employees = hr.get_employees(company_id, sandbox=sandbox)
    employees = [e for e in employees if is_active_equivalent(e.get("status", "")) and e.get("hireDate")]

    branches = DatabaseService.get_branches(owner_uid, sandbox=sandbox, company_id=company_id)
    branch_map = {b["id"]: b.get("name", b.get("code", b["id"])) for b in branches}

    today = date.today()
    results = []
    for emp in employees:
        try:
            hd = date.fromisoformat(emp["hireDate"][:10])
        except (ValueError, TypeError):
            continue

        years = today.year - hd.year
        if today.month < hd.month or (today.month == hd.month and today.day < hd.day):
            years -= 1

        results.append({
            "id": emp.get("id", ""),
            "code": emp.get("code", ""),
            "fullName": emp.get("fullName", ""),
            "branchName": branch_map.get(emp.get("branchId", ""), ""),
            "department": emp.get("department", "") or emp.get("area", ""),
            "position": emp.get("position", ""),
            "hireDate": hd,
            "anniversaryDay": hd.day,
            "anniversaryMonth": hd.month,
            "anniversaryYear": hd.year,
            "yearsInCompany": max(0, years),
        })

    results.sort(key=lambda r: (r["anniversaryMonth"], r["anniversaryDay"]))
    return results, branch_map


@web_invoices_bp.route('/reports/empleados/anniversary')
@web_rrhh_bp.route("/rrhh/reports/empleados/anniversary")
@require_module('nomina')
def report_anniversary():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()

    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
    except (ValueError, TypeError):
        year = today.year
    try:
        month = int(request.args.get("month", today.month))
    except (ValueError, TypeError):
        month = today.month
    try:
        day = int(request.args.get("day", today.day))
    except (ValueError, TypeError):
        day = today.day

    results, _ = _build_anniversary_data(company_id, sandbox, owner_uid)

    if day > 0:
        filtered = [r for r in results if r["anniversaryMonth"] == month and r["anniversaryDay"] == day]
    else:
        filtered = [r for r in results if r["anniversaryMonth"] == month]

    is_today = (year == today.year and month == today.month and day == today.day)

    return render_template(
        "rrhh/reports/anniversary.html",
        active_page="rrhh_reports",
        results=filtered,
        year=year,
        month=month,
        day=day,
        is_today=is_today,
        today=today,
        meses=MESES_ES_MAP,
        is_pdf=False,
    )


@web_invoices_bp.route('/reports/empleados/anniversary/export')
@web_rrhh_bp.route("/rrhh/reports/empleados/anniversary/export")
@require_module('nomina')
def report_anniversary_export():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()

    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
    except (ValueError, TypeError):
        year = today.year
    try:
        month = int(request.args.get("month", today.month))
    except (ValueError, TypeError):
        month = today.month
    try:
        day = int(request.args.get("day", today.day))
    except (ValueError, TypeError):
        day = today.day

    results, _ = _build_anniversary_data(company_id, sandbox, owner_uid)
    if day > 0:
        filtered = [r for r in results if r["anniversaryMonth"] == month and r["anniversaryDay"] == day]
    else:
        filtered = [r for r in results if r["anniversaryMonth"] == month]

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aniversarios"
        ws.append(["Codigo", "Nombre", "Sucursal", "Departamento", "Puesto",
                    "Dia de Aniversario", "Anos en la Empresa"])
        for r in filtered:
            ws.append([
                r["code"],
                r["fullName"],
                r["branchName"],
                r["department"],
                r["position"],
                f"{r['anniversaryDay']:02d}/{r['anniversaryMonth']:02d}",
                r["yearsInCompany"],
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"aniversarios_{year}_{month:02d}_{day:02d}.xlsx"
        return send_file(output,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except ImportError:
        csv_out = io.StringIO()
        csv_out.write("Codigo,Nombre,Sucursal,Departamento,Puesto,Dia de Aniversario,Anos en la Empresa\n")
        for r in filtered:
            csv_out.write(
                 f"{r['code']},{r['fullName']},{r['branchName']},{r['department']},"
                f"{r['position']},{r['anniversaryDay']:02d}/{r['anniversaryMonth']:02d},{r['yearsInCompany']}\n"
            )
        buf = io.BytesIO()
        buf.write(b"\xef\xbb\xbf")
        buf.write(csv_out.getvalue().encode("utf-8"))
        buf.seek(0)
        filename = f"aniversarios_{year}_{month:02d}_{day:02d}.csv"
        return send_file(buf, mimetype="text/csv", as_attachment=True, download_name=filename)


@web_invoices_bp.route('/reports/empleados/anniversary/pdf')
@web_rrhh_bp.route("/rrhh/reports/empleados/anniversary/pdf")
@require_module('nomina')
def report_anniversary_pdf():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services.db_service import DatabaseService
    from app.utils.pdf import pdf_write_options
    from weasyprint import HTML as WeasyprintHTML

    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
    except (ValueError, TypeError):
        year = today.year
    try:
        month = int(request.args.get("month", today.month))
    except (ValueError, TypeError):
        month = today.month
    try:
        day = int(request.args.get("day", today.day))
    except (ValueError, TypeError):
        day = today.day

    results, _ = _build_anniversary_data(company_id, sandbox, owner_uid)
    if day > 0:
        filtered = [r for r in results if r["anniversaryMonth"] == month and r["anniversaryDay"] == day]
    else:
        filtered = [r for r in results if r["anniversaryMonth"] == month]

    company = DatabaseService.get_company_profile(owner_uid, company_id=company_id) or {}
    is_today = (year == today.year and month == today.month and day == today.day)

    rendered = render_template(
        "rrhh/reports/anniversary_pdf.html",
        results=filtered,
        year=year,
        month=month,
        day=day,
        is_today=is_today,
        today=today,
        meses=MESES_ES_MAP,
        company=company,
    )

    pdf_bytes = WeasyprintHTML(string=rendered, base_url=request.host_url).write_pdf(**pdf_write_options())
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="aniversarios_{year}_{month:02d}_{day:02d}.pdf"'
    return response


# ═══════════════════════════════════════════════════════════════
#  Reporte: Listado de Cumpleaños
# ═══════════════════════════════════════════════════════════════


def _build_birthday_data(company_id, sandbox, owner_uid):
    from app.services.db_service import DatabaseService

    employees = hr.get_employees(company_id, sandbox=sandbox)
    employees = [e for e in employees if is_active_equivalent(e.get("status", "")) and e.get("birthDate")]

    branches = DatabaseService.get_branches(owner_uid, sandbox=sandbox, company_id=company_id)
    branch_map = {b["id"]: b.get("name", b.get("code", b["id"])) for b in branches}

    today = date.today()
    results = []
    for emp in employees:
        try:
            bd = date.fromisoformat(emp["birthDate"][:10])
        except (ValueError, TypeError):
            continue

        age = today.year - bd.year
        if today.month < bd.month or (today.month == bd.month and today.day < bd.day):
            age -= 1

        results.append({
            "id": emp.get("id", ""),
            "code": emp.get("code", ""),
            "fullName": emp.get("fullName", ""),
            "branchName": branch_map.get(emp.get("branchId", ""), ""),
            "department": emp.get("department", "") or emp.get("area", ""),
            "position": emp.get("position", ""),
            "birthDate": bd,
            "birthDay": bd.day,
            "birthMonth": bd.month,
            "birthYear": bd.year,
            "age": max(0, age),
        })

    results.sort(key=lambda r: (r["birthMonth"], r["birthDay"]))
    return results, branch_map


@web_invoices_bp.route('/reports/empleados/birthday')
@web_rrhh_bp.route("/rrhh/reports/empleados/birthday")
@require_module('nomina')
def report_birthday():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()

    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
    except (ValueError, TypeError):
        year = today.year
    try:
        month = int(request.args.get("month", today.month))
    except (ValueError, TypeError):
        month = today.month
    try:
        day = int(request.args.get("day", today.day))
    except (ValueError, TypeError):
        day = today.day

    results, _ = _build_birthday_data(company_id, sandbox, owner_uid)

    if day > 0:
        filtered = [r for r in results if r["birthMonth"] == month and r["birthDay"] == day]
    else:
        filtered = [r for r in results if r["birthMonth"] == month]

    is_today = (year == today.year and month == today.month and day == today.day)

    return render_template(
        "rrhh/reports/birthday.html",
        active_page="rrhh_reports",
        results=filtered,
        year=year,
        month=month,
        day=day,
        is_today=is_today,
        today=today,
        meses=MESES_ES_MAP,
    )


@web_invoices_bp.route('/reports/empleados/birthday/export')
@web_rrhh_bp.route("/rrhh/reports/empleados/birthday/export")
@require_module('nomina')
def report_birthday_export():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()

    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
    except (ValueError, TypeError):
        year = today.year
    try:
        month = int(request.args.get("month", today.month))
    except (ValueError, TypeError):
        month = today.month
    try:
        day = int(request.args.get("day", today.day))
    except (ValueError, TypeError):
        day = today.day

    results, _ = _build_birthday_data(company_id, sandbox, owner_uid)
    if day > 0:
        filtered = [r for r in results if r["birthMonth"] == month and r["birthDay"] == day]
    else:
        filtered = [r for r in results if r["birthMonth"] == month]

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cumpleanos"
        ws.append(["Codigo", "Nombre", "Sucursal", "Departamento", "Puesto",
                    "Dia de Nacimiento", "Edad"])
        for r in filtered:
            ws.append([
                r["code"],
                r["fullName"],
                r["branchName"],
                r["department"],
                r["position"],
                f"{r['birthDay']:02d}/{r['birthMonth']:02d}",
                r["age"],
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"cumpleanos_{year}_{month:02d}_{day:02d}.xlsx"
        return send_file(output,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except ImportError:
        csv_out = io.StringIO()
        csv_out.write("Codigo,Nombre,Sucursal,Departamento,Puesto,Dia de Nacimiento,Edad\n")
        for r in filtered:
            csv_out.write(
                 f"{r['code']},{r['fullName']},{r['branchName']},{r['department']},"
                f"{r['position']},{r['birthDay']:02d}/{r['birthMonth']:02d},{r['age']}\n"
            )
        buf = io.BytesIO()
        buf.write(b"\xef\xbb\xbf")
        buf.write(csv_out.getvalue().encode("utf-8"))
        buf.seek(0)
        filename = f"cumpleanos_{year}_{month:02d}_{day:02d}.csv"
        return send_file(buf, mimetype="text/csv", as_attachment=True, download_name=filename)


@web_invoices_bp.route('/reports/empleados/birthday/pdf')
@web_rrhh_bp.route("/rrhh/reports/empleados/birthday/pdf")
@require_module('nomina')
def report_birthday_pdf():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services.db_service import DatabaseService
    from app.utils.pdf import pdf_write_options
    from weasyprint import HTML as WeasyprintHTML

    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
    except (ValueError, TypeError):
        year = today.year
    try:
        month = int(request.args.get("month", today.month))
    except (ValueError, TypeError):
        month = today.month
    try:
        day = int(request.args.get("day", today.day))
    except (ValueError, TypeError):
        day = today.day

    results, _ = _build_birthday_data(company_id, sandbox, owner_uid)
    if day > 0:
        filtered = [r for r in results if r["birthMonth"] == month and r["birthDay"] == day]
    else:
        filtered = [r for r in results if r["birthMonth"] == month]

    company = DatabaseService.get_company_profile(owner_uid, company_id=company_id) or {}
    is_today = (year == today.year and month == today.month and day == today.day)

    rendered = render_template(
        "rrhh/reports/birthday_pdf.html",
        results=filtered,
        year=year,
        month=month,
        day=day,
        is_today=is_today,
        today=today,
        meses=MESES_ES_MAP,
        company=company,
    )

    pdf_bytes = WeasyprintHTML(string=rendered, base_url=request.host_url).write_pdf(**pdf_write_options())
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="cumpleanos_{year}_{month:02d}_{day:02d}.pdf"'
    return response


# ═══════════════════════════════════════════════════════════════
#  Reporte: Saldo de Vacaciones
# ═══════════════════════════════════════════════════════════════


def _build_vacation_balance_data(company_id, sandbox, owner_uid):
    from app.services.db_service import DatabaseService

    employees = hr.get_employees(company_id, sandbox=sandbox)
    employees = [e for e in employees if is_active_equivalent(e.get("status", "")) and e.get("hireDate")]

    emp_map = {e.get("id", ""): e for e in employees}

    vacation_requests = hr.get_vacation_requests(company_id, sandbox=sandbox)

    branches = DatabaseService.get_branches(owner_uid, sandbox=sandbox, company_id=company_id)
    branch_map = {b["id"]: b.get("name", b.get("code", b["id"])) for b in branches}

    today = date.today()
    results = []
    for emp in employees:
        emp_id = emp.get("id", "")
        hire_date_str = emp.get("hireDate", "")

        supervisor_name = ""
        sup_id = emp.get("reportsTo", "")
        if sup_id and sup_id in emp_map:
            supervisor_name = emp_map[sup_id].get("fullName", "")

        taken = sum(
            r.get("days", 0) for r in vacation_requests
            if r.get("employeeId") == emp_id and r.get("status") == "aprobada"
        )
        accrued = PayrollService.calculate_vacation_days(hire_date_str, today=today, taken_days=0)
        remaining = PayrollService.calculate_vacation_days(hire_date_str, today=today, taken_days=taken)

        try:
            hd = date.fromisoformat(hire_date_str[:10])
        except (ValueError, TypeError):
            hd = None

        years_in_company = 0
        if hd:
            years_in_company = today.year - hd.year
            if today.month < hd.month or (today.month == hd.month and today.day < hd.day):
                years_in_company -= 1
            years_in_company = max(0, years_in_company)

        results.append({
            "id": emp_id,
            "code": emp.get("code", ""),
            "fullName": emp.get("fullName", ""),
            "branchName": branch_map.get(emp.get("branchId", ""), ""),
            "department": emp.get("department", "") or emp.get("area", ""),
            "position": emp.get("position", ""),
            "supervisorName": supervisor_name,
            "hireDate": hire_date_str,
            "yearsInCompany": years_in_company,
            "accruedDays": accrued,
            "takenDays": taken,
            "remainingDays": remaining,
        })

    results.sort(key=lambda r: r["fullName"].lower())
    return results, branch_map


@web_invoices_bp.route('/reports/empleados/vacation-balance')
@web_rrhh_bp.route("/rrhh/reports/empleados/vacation-balance")
@require_module('nomina')
def report_vacation_balance():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()

    all_results, branch_map = _build_vacation_balance_data(company_id, sandbox, owner_uid)

    filter_branch = request.args.get("branch", "").strip()
    filter_department = request.args.get("department", "").strip()
    filter_position = request.args.get("position", "").strip()
    filter_supervisor = request.args.get("supervisor", "").strip()
    filter_employee = request.args.get("employee", "").strip()

    results = all_results
    if filter_branch:
        results = [r for r in results if r["branchName"] == filter_branch]
    if filter_department:
        results = [r for r in results if r["department"] == filter_department]
    if filter_position:
        results = [r for r in results if r["position"] == filter_position]
    if filter_supervisor:
        results = [r for r in results if r["supervisorName"] == filter_supervisor]
    if filter_employee:
        results = [r for r in results if filter_employee.lower() in r["fullName"].lower()]

    departments_set = sorted(set(r["department"] for r in all_results if r["department"]))
    positions_set = sorted(set(r["position"] for r in all_results if r["position"]))
    supervisors_set = sorted(set(r["supervisorName"] for r in all_results if r["supervisorName"]))
    employees_list = sorted(set(r["fullName"] for r in all_results))
    branches_list = sorted(set(branch_map.values()))

    return render_template(
        "rrhh/reports/vacation_balance.html",
        active_page="rrhh_reports",
        results=results,
        branches=branches_list,
        departments=departments_set,
        positions=positions_set,
        supervisors=supervisors_set,
        employees=employees_list,
        filter_branch=filter_branch,
        filter_department=filter_department,
        filter_position=filter_position,
        filter_supervisor=filter_supervisor,
        filter_employee=filter_employee,
        today=date.today(),
    )


@web_invoices_bp.route('/reports/empleados/vacation-balance/export')
@web_rrhh_bp.route("/rrhh/reports/empleados/vacation-balance/export")
@require_module('nomina')
def report_vacation_balance_export():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()

    results, branch_map = _build_vacation_balance_data(company_id, sandbox, owner_uid)

    filter_branch = request.args.get("branch", "").strip()
    filter_department = request.args.get("department", "").strip()
    filter_position = request.args.get("position", "").strip()
    filter_supervisor = request.args.get("supervisor", "").strip()
    filter_employee = request.args.get("employee", "").strip()

    if filter_branch:
        results = [r for r in results if r["branchName"] == filter_branch]
    if filter_department:
        results = [r for r in results if r["department"] == filter_department]
    if filter_position:
        results = [r for r in results if r["position"] == filter_position]
    if filter_supervisor:
        results = [r for r in results if r["supervisorName"] == filter_supervisor]
    if filter_employee:
        results = [r for r in results if filter_employee.lower() in r["fullName"].lower()]

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Saldo Vacaciones"
        ws.append(["Codigo", "Nombre", "Sucursal", "Departamento", "Puesto",
                    "Fecha Ingreso", "Anos Antiguedad", "Dias Acumulados",
                    "Dias Tomados", "Dias Pendientes"])
        for r in results:
            ws.append([
                r["code"], r["fullName"], r["branchName"], r["department"],
                r["position"], r["hireDate"], r["yearsInCompany"],
                r["accruedDays"], r["takenDays"], r["remainingDays"],
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        today = date.today()
        filename = f"saldo_vacaciones_{today.strftime('%Y%m%d')}.xlsx"
        return send_file(output,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except ImportError:
        csv_out = io.StringIO()
        csv_out.write("Codigo,Nombre,Sucursal,Departamento,Puesto,Fecha Ingreso,Anos Antiguedad,Dias Acumulados,Dias Tomados,Dias Pendientes\n")
        for r in results:
            csv_out.write(
                 f"{r['code']},{r['fullName']},{r['branchName']},{r['department']},"
                f"{r['position']},{r['hireDate']},{r['yearsInCompany']},"
                f"{r['accruedDays']},{r['takenDays']},{r['remainingDays']}\n"
            )
        buf = io.BytesIO()
        buf.write(b"\xef\xbb\xbf")
        buf.write(csv_out.getvalue().encode("utf-8"))
        buf.seek(0)
        today = date.today()
        filename = f"saldo_vacaciones_{today.strftime('%Y%m%d')}.csv"
        return send_file(buf, mimetype="text/csv", as_attachment=True, download_name=filename)


@web_invoices_bp.route('/reports/empleados/vacation-balance/pdf')
@web_rrhh_bp.route("/rrhh/reports/empleados/vacation-balance/pdf")
@require_module('nomina')
def report_vacation_balance_pdf():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services.db_service import DatabaseService
    from app.utils.pdf import pdf_write_options
    from weasyprint import HTML as WeasyprintHTML

    today = date.today()
    results, _ = _build_vacation_balance_data(company_id, sandbox, owner_uid)

    filter_branch = request.args.get("branch", "").strip()
    filter_department = request.args.get("department", "").strip()
    filter_position = request.args.get("position", "").strip()
    filter_supervisor = request.args.get("supervisor", "").strip()
    filter_employee = request.args.get("employee", "").strip()

    if filter_branch:
        results = [r for r in results if r["branchName"] == filter_branch]
    if filter_department:
        results = [r for r in results if r["department"] == filter_department]
    if filter_position:
        results = [r for r in results if r["position"] == filter_position]
    if filter_supervisor:
        results = [r for r in results if r["supervisorName"] == filter_supervisor]
    if filter_employee:
        results = [r for r in results if filter_employee.lower() in r["fullName"].lower()]

    company = DatabaseService.get_company_profile(owner_uid, company_id=company_id) or {}

    rendered = render_template(
        "rrhh/reports/vacation_balance_pdf.html",
        results=results,
        today=today,
        company=company,
        filter_branch=filter_branch,
    )

    pdf_bytes = WeasyprintHTML(string=rendered, base_url=request.host_url).write_pdf(**pdf_write_options())
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="saldo_vacaciones_{today.strftime("%Y%m%d")}.pdf"'
    return response


# ═══════════════════════════════════════════════════════════════
#  Reporte: Períodos de Vacaciones
# ═══════════════════════════════════════════════════════════════


def _build_vacation_periods_data(company_id, sandbox, owner_uid):
    from app.services.db_service import DatabaseService

    employees = hr.get_employees(company_id, sandbox=sandbox)
    emp_map = {e.get("id", ""): e for e in employees}

    vacation_requests = hr.get_vacation_requests(company_id, sandbox=sandbox)

    branches = DatabaseService.get_branches(owner_uid, sandbox=sandbox, company_id=company_id)
    branch_map = {b["id"]: b.get("name", b.get("code", b["id"])) for b in branches}

    results = []
    for req in vacation_requests:
        if req.get("status") != "aprobada":
            continue
        emp = emp_map.get(req.get("employeeId", ""), {})
        if not is_active_equivalent(emp.get("status", "")):
            continue

        try:
            sd = date.fromisoformat(req["startDate"][:10])
        except (ValueError, TypeError):
            sd = None
        try:
            ed = date.fromisoformat(req["endDate"][:10])
        except (ValueError, TypeError):
            ed = None

        results.append({
            "id": req.get("id", ""),
            "code": emp.get("code", ""),
            "fullName": req.get("employeeName", emp.get("fullName", "")),
            "branchName": branch_map.get(emp.get("branchId", ""), ""),
            "department": emp.get("department", "") or emp.get("area", ""),
            "position": emp.get("position", ""),
            "startDate": sd,
            "endDate": ed,
            "days": req.get("days", 0),
            "status": req.get("status", ""),
            "startMonth": sd.month if sd else 0,
            "startYear": sd.year if sd else 0,
        })

    results.sort(key=lambda r: (r.get("startDate") or date.min))
    return results, branch_map


@web_invoices_bp.route('/reports/empleados/vacation-periods')
@web_rrhh_bp.route("/rrhh/reports/empleados/vacation-periods")
@require_module('nomina')
def report_vacation_periods():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()

    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
    except (ValueError, TypeError):
        year = today.year
    try:
        month = int(request.args.get("month", 0))
    except (ValueError, TypeError):
        month = 0

    results, _ = _build_vacation_periods_data(company_id, sandbox, owner_uid)

    if month > 0:
        filtered = [r for r in results if r["startYear"] == year and r["startMonth"] == month]
    else:
        filtered = [r for r in results if r["startYear"] == year]

    return render_template(
        "rrhh/reports/vacation_periods.html",
        active_page="rrhh_reports",
        results=filtered,
        year=year,
        month=month,
        today=today,
        meses=MESES_ES_MAP,
    )


@web_invoices_bp.route('/reports/empleados/vacation-periods/export')
@web_rrhh_bp.route("/rrhh/reports/empleados/vacation-periods/export")
@require_module('nomina')
def report_vacation_periods_export():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()

    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
    except (ValueError, TypeError):
        year = today.year
    try:
        month = int(request.args.get("month", 0))
    except (ValueError, TypeError):
        month = 0

    results, _ = _build_vacation_periods_data(company_id, sandbox, owner_uid)

    if month > 0:
        filtered = [r for r in results if r["startYear"] == year and r["startMonth"] == month]
    else:
        filtered = [r for r in results if r["startYear"] == year]

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Periodos Vacaciones"
        ws.append(["Codigo", "Nombre", "Sucursal", "Departamento", "Puesto",
                    "Desde", "Hasta", "Dias", "Estado"])
        for r in filtered:
            ws.append([
                r["code"], r["fullName"], r["branchName"], r["department"],
                r["position"],
                r["startDate"].strftime("%d/%m/%Y") if r["startDate"] else "",
                r["endDate"].strftime("%d/%m/%Y") if r["endDate"] else "",
                r["days"], r["status"],
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"periodos_vacaciones_{year}_{month:02d}.xlsx" if month > 0 else f"periodos_vacaciones_{year}.xlsx"
        return send_file(output,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except ImportError:
        csv_out = io.StringIO()
        csv_out.write("Codigo,Nombre,Sucursal,Departamento,Puesto,Desde,Hasta,Dias,Estado\n")
        for r in filtered:
            sd = r["startDate"].strftime("%d/%m/%Y") if r["startDate"] else ""
            ed = r["endDate"].strftime("%d/%m/%Y") if r["endDate"] else ""
            csv_out.write(
                 f"{r['code']},{r['fullName']},{r['branchName']},{r['department']},"
                f"{r['position']},{sd},{ed},{r['days']},{r['status']}\n"
            )
        buf = io.BytesIO()
        buf.write(b"\xef\xbb\xbf")
        buf.write(csv_out.getvalue().encode("utf-8"))
        buf.seek(0)
        filename = f"periodos_vacaciones_{year}_{month:02d}.csv" if month > 0 else f"periodos_vacaciones_{year}.csv"
        return send_file(buf, mimetype="text/csv", as_attachment=True, download_name=filename)


@web_invoices_bp.route('/reports/empleados/vacation-periods/pdf')
@web_rrhh_bp.route("/rrhh/reports/empleados/vacation-periods/pdf")
@require_module('nomina')
def report_vacation_periods_pdf():
    if _login_required():
        return redirect(url_for("web_auth.login"))
    owner_uid, sandbox, company_id = _get_owner_uid_and_sandbox()
    from app.services.db_service import DatabaseService
    from app.utils.pdf import pdf_write_options
    from weasyprint import HTML as WeasyprintHTML

    today = date.today()
    try:
        year = int(request.args.get("year", today.year))
    except (ValueError, TypeError):
        year = today.year
    try:
        month = int(request.args.get("month", 0))
    except (ValueError, TypeError):
        month = 0

    results, _ = _build_vacation_periods_data(company_id, sandbox, owner_uid)

    if month > 0:
        filtered = [r for r in results if r["startYear"] == year and r["startMonth"] == month]
    else:
        filtered = [r for r in results if r["startYear"] == year]

    company = DatabaseService.get_company_profile(owner_uid, company_id=company_id) or {}

    rendered = render_template(
        "rrhh/reports/vacation_periods_pdf.html",
        results=filtered,
        year=year,
        month=month,
        today=today,
        meses=MESES_ES_MAP,
        company=company,
    )

    pdf_bytes = WeasyprintHTML(string=rendered, base_url=request.host_url).write_pdf(**pdf_write_options())
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    filename = f"periodos_vacaciones_{year}_{month:02d}.pdf" if month > 0 else f"periodos_vacaciones_{year}.pdf"
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
