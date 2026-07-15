import io
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file, g
from app.services.db_service import DatabaseService
from app.utils.decorators import check_permission

web_reports_606_bp = Blueprint('web_reports_606', __name__)


@web_reports_606_bp.before_request
def restrict_to_do():
    if session.get('company_country', 'DO') != 'DO':
        return render_template('auth/restricted.html',
            feature_name="Reporte 606 DGII (solo disponible para República Dominicana)",
            required_permission="")

TIPO_GASTO_606 = {
    "01": "Gastos de Personal",
    "02": "Gastos por Trabajos, Suministros y Servicios",
    "03": "Arrendamientos",
    "04": "Gastos de Activos Fijos",
    "05": "Gastos de Representación",
    "06": "Otras Deducciones Admitidas",
    "07": "Gastos Financieros",
    "08": "Gastos Extraordinarios",
    "09": "Compras y Gastos que Formarán Parte del Costo de Venta",
    "10": "Adquisiciones de Activos",
    "11": "Gastos de Seguros",
}

FORMA_PAGO_606 = {
    "01": "Efectivo",
    "02": "Cheques/Transferencias/Depósito",
    "03": "Tarjeta Crédito/Débito",
    "04": "Compra a Crédito",
    "05": "Permuta",
    "06": "Notas de Crédito",
    "07": "Mixto",
}

TIPO_RETENCION_ISR_606 = {
    "01": "Alquileres",
    "02": "Honorarios por Servicios",
    "03": "Otras Rentas",
    "04": "Otras Rentas (Rentas Presuntas)",
    "05": "Intereses Pagados a Personas Jurídicas Residentes",
    "06": "Intereses Pagados a Personas Físicas Residentes",
    "07": "Retención por Proveedores del Estado",
    "08": "Juegos Telefónicos",
    "09": "Retenciones Subsector de Ganadería de Carne Bovina",
}

from app.models.fiscal_document_type import report_labels as _report_labels_606
ECF_TYPE_LABELS = _report_labels_606("606")


def _parse_int(val, default=0):
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def _get_expenses_for_period(owner_uid, sandbox, year, month):
    prefix = f"{year:04d}-{month:02d}"
    all_items = []
    branch_id = g.get('branch_id')
    project_id = g.get('project_id')

    expenses = DatabaseService.get_expenses(owner_uid, sandbox=sandbox, branch_id=branch_id, project_id=project_id)
    for exp in expenses:
        d = (exp.get("date") or exp.get("createdAt") or "")[:7]
        if d == prefix:
            all_items.append(exp)

    try:
        from app.services.supplier_invoice_service import SupplierInvoiceService
        invoices = SupplierInvoiceService.get_all(owner_uid, sandbox=sandbox)
        for inv in invoices:
            if branch_id and inv.get("branchId") != branch_id:
                continue
            if project_id == '__no_project__':
                if inv.get("projectId"):
                    continue
            elif project_id:
                if inv.get("projectId") != project_id:
                    continue
            d = (inv.get("date") or inv.get("createdAt") or "")[:7]
            if d == prefix:
                concept = inv.get("comentario") or inv.get("notes") or ""
                if not concept and inv.get("supplierInvoiceNumber"):
                    concept = f"Compra {inv.get('supplierInvoiceNumber', '')} - {inv.get('supplierName', '')}"
                mapped = {
                    "id": inv.get("id", ""),
                    "concept": concept,
                    "amount": float(inv.get("total", 0)),
                    "itbisAmount": float(inv.get("itbis", 0)),
                    "date": inv.get("date", ""),
                    "dueDate": inv.get("dueDate", ""),
                    "rncEmisor": inv.get("supplierRnc", ""),
                    "providerName": inv.get("supplierName", ""),
                    "ncf": inv.get("ncf", ""),
                    "encf": inv.get("ncf", ""),
                    "ecfType": inv.get("ecfType", ""),
                    "tipoGastoDGII": inv.get("tipoGastoDGII", "02"),
                    "supplierId": inv.get("supplierId", ""),
                    "isDeductible": inv.get("isDeductible", True),
                    "isITBISDeductible": inv.get("isITBISDeductible", True),
                    "isrWithheld": float(inv.get("retainedISR", 0) or 0),
                    "itbisWithheld": float(inv.get("retainedITBIS", 0) or 0),
                    "formaPago": "02",
                    "dgiiStatus": inv.get("dgiiStatus", ""),
                    "createdAt": inv.get("createdAt", ""),
                }
                all_items.append(mapped)
    except Exception as e:
        print(f"Error al obtener facturas proveedor para 606: {e}")

    return all_items


def _filter_expenses(expenses, supplier_id, tipo_gasto, ecf_type, search):
    result = expenses
    if supplier_id:
        result = [e for e in result if e.get("supplierId") == supplier_id]
    if tipo_gasto:
        result = [e for e in result if e.get("tipoGastoDGII") == tipo_gasto]
    if ecf_type:
        result = [e for e in result if e.get("ecfType") == ecf_type]
    if search:
        q = search.lower()
        result = [
            e
            for e in result
            if q in (e.get("providerName") or "").lower()
            or q in (e.get("rncEmisor") or "").lower()
            or q in (e.get("ncf") or "").lower()
            or q in (e.get("concept") or "").lower()
        ]
    return result


def _to_dgii_date(date_str):
    if not date_str:
        return ""
    clean = str(date_str).strip()[:10]
    return clean.replace("-", "")


def _resolve_montos(exp):
    monto_serv = float(exp.get("montoServicios", 0) or 0)
    monto_bienes = float(exp.get("montoBienes", 0) or 0)
    total = float(exp.get("amount", 0) or 0)
    if monto_serv + monto_bienes == 0:
        monto_bienes = total
    else:
        total = monto_serv + monto_bienes
    return monto_serv, monto_bienes, total


def _generate_606_xlsx(owner_rnc, period, expenses):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Herramienta Formato 606"

    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(size=10)

    ws.merge_cells('E2:K2')
    c = ws['E2']
    c.value = "Formato de Envío de Compras de Bienes y Servicios"
    c.font = title_font

    ws.merge_cells('E3:K3')
    ws['A3'] = "Versión 2025"
    ws['E3'] = "Herramienta de Distribucion Gratuita"

    ws.merge_cells('A4:B4')
    ws['A4'] = "RNC o Cédula"
    ws['A4'].font = header_font
    ws['C4'] = owner_rnc

    ws.merge_cells('A5:B5')
    ws['A5'] = "Periodo"
    ws['A5'].font = header_font
    ws['C5'] = period

    ws.merge_cells('A6:B6')
    ws['A6'] = "Cantidad Registros"
    ws['A6'].font = header_font
    ws['C6'] = len(expenses)

    headers = [
        "Líneas", "RNC o Cédula", "Tipo Id", "Tipo Bienes y Servicios Comprados",
        "NCF", "NCF ó Documento Modificado", "Fecha Comprobante", "",
        "Fecha Pago", "", "Monto Facturado en Servicios", "Monto Facturado en Bienes",
        "Total Monto Facturado", "ITBIS Facturado", "ITBIS Retenido",
        "ITBIS sujeto a Proporcionalidad (Art. 349)", "ITBIS llevado al Costo",
        "ITBIS por Adelantar", "ITBIS percibido en compras", "Tipo de Retención en ISR",
        "Monto Retención Renta", "ISR Percibido en compras", "Impuesto Selectivo al Consumo",
        "Otros Impuesto/Tasas", "Monto Propina Legal", "Forma de Pago", "Estatus",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_idx, value=h)
        cell.font = Font(bold=True, size=9)

    numbering_cols = {
        2: 1, 3: 2, 4: 3, 5: 4, 6: 5, 7: 6,
        9: 7, 11: 8, 12: 9, 13: 10, 14: 11, 15: 12, 16: 13,
        17: 14, 18: 15, 19: 16, 20: 17, 21: 18,
        22: 19, 23: 20, 24: 21, 25: 22, 26: 23,
    }
    num_font = Font(name="Calibri", size=9)
    for col_idx, num in numbering_cols.items():
        cell = ws.cell(row=10, column=col_idx, value=num)
        cell.font = num_font
        cell.alignment = Alignment(horizontal="center")

    for i, exp in enumerate(expenses):
        row = i + 12
        monto_serv, monto_bienes, total = _resolve_montos(exp)
        itbis_fact = float(exp.get("itbisAmount", 0) or 0)
        itbis_costo = float(exp.get("itbisLlevadoCosto", 0) or 0)
        itbis_adelantar = itbis_fact - itbis_costo
        date_str = _to_dgii_date(exp.get("date"))
        due_str = _to_dgii_date(exp.get("dueDate"))
        ncf = exp.get("ncf") or exp.get("encf") or ""

        valores = [
            i + 1,
            exp.get("rncEmisor", ""),
            "1",
            exp.get("tipoGastoDGII", "02"),
            ncf,
            exp.get("ncfModificado", ""),
            date_str,
            "",
            due_str,
            "",
            monto_serv,
            monto_bienes,
            total,
            itbis_fact,
            float(exp.get("itbisWithheld", 0) or 0),
            float(exp.get("itbisProporcionalidad", 0) or 0),
            itbis_costo,
            itbis_adelantar,
            "",
            exp.get("tipoRetencionISR", ""),
            float(exp.get("isrWithheld", 0) or 0),
            "",
            float(exp.get("iscMonto", 0) or 0),
            float(exp.get("otrosImpuestos", 0) or 0),
            float(exp.get("propinaLegal", 0) or 0),
            exp.get("formaPago", "02"),
            "",
        ]
        for col_idx, val in enumerate(valores, 1):
            ws.cell(row=row, column=col_idx, value=val)

    col_widths = {
        'A': 8, 'B': 16, 'C': 8, 'D': 20, 'E': 20, 'F': 20, 'G': 14, 'H': 4,
        'I': 14, 'J': 4, 'K': 16, 'L': 16, 'M': 16, 'N': 14, 'O': 14,
        'P': 16, 'Q': 14, 'R': 14, 'S': 14, 'T': 16, 'U': 16, 'V': 14,
        'W': 16, 'X': 16, 'Y': 14, 'Z': 14, 'AA': 20,
    }
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    dest = io.BytesIO()
    wb.save(dest)
    dest.seek(0)
    return dest


def _generate_606_txt(owner_rnc, period, expenses):
    output = io.StringIO()
    n = len(expenses)
    output.write(f"606|{owner_rnc}|{period}|{n}\n")

    for exp in expenses:
        monto_serv, monto_bienes, total = _resolve_montos(exp)
        itbis_fact = float(exp.get("itbisAmount", 0) or 0)
        itbis_costo = float(exp.get("itbisLlevadoCosto", 0) or 0)
        itbis_adelantar = itbis_fact - itbis_costo
        date_str = _to_dgii_date(exp.get("date"))
        due_str = _to_dgii_date(exp.get("dueDate"))
        ncf = exp.get("ncf") or exp.get("encf") or ""

        fields = [
            exp.get("rncEmisor", ""),
            "1",
            exp.get("tipoGastoDGII", "06"),
            ncf,
            exp.get("ncfModificado", ""),
            date_str,
            due_str,
            f'{monto_serv:.2f}',
            f'{monto_bienes:.2f}',
            f'{total:.2f}',
            f'{itbis_fact:.2f}',
            f'{float(exp.get("itbisWithheld", 0) or 0):.2f}',
            f'{float(exp.get("itbisProporcionalidad", 0) or 0):.2f}',
            f'{itbis_costo:.2f}',
            f'{itbis_adelantar:.2f}',
            "",
            exp.get("tipoRetencionISR", ""),
            f'{float(exp.get("isrWithheld", 0) or 0):.2f}',
            "",
            f'{float(exp.get("iscMonto", 0) or 0):.2f}',
            f'{float(exp.get("otrosImpuestos", 0) or 0):.2f}',
            f'{float(exp.get("propinaLegal", 0) or 0):.2f}',
            exp.get("formaPago", "02"),
        ]
        output.write('|'.join(fields) + '\n')

    return output.getvalue()


# ═════════════════════════════════════════════════════════════════════
# REPORTE 606 — Main View
# ═════════════════════════════════════════════════════════════════════


@web_reports_606_bp.route("/reports/606")
def reporte_606():
    if "user" not in session:
        return redirect(url_for("login"))
    if not check_permission("canExpenses"):
        return render_template("auth/restricted.html", active_page="reporte_606")
    owner_uid = session["user"]["ownerUID"]
    sandbox = session.get("is_sandbox_mode", True)

    now = datetime.now(timezone.utc)
    year = _parse_int(request.args.get("year"), now.year)
    month = _parse_int(request.args.get("month"), now.month)
    supplier_id = request.args.get("supplier_id", "").strip()
    tipo_gasto = request.args.get("tipo_gasto", "").strip()
    ecf_type = request.args.get("ecf_type", "").strip()
    search = request.args.get("search", "").strip()

    expenses = _get_expenses_for_period(owner_uid, sandbox, year, month)
    filtered = _filter_expenses(expenses, supplier_id, tipo_gasto, ecf_type, search)
    filtered.sort(key=lambda x: (x.get("date") or x.get("createdAt") or ""))

    total_monto = sum(float(e.get("amount", 0)) for e in filtered)
    total_itbis = sum(float(e.get("itbisAmount", 0)) for e in filtered)
    total_itbis_deducible = sum(
        float(e.get("itbisAmount", 0))
        for e in filtered
        if e.get("isITBISDeductible")
    )
    total_deducible = sum(
        float(e.get("amount", 0)) for e in filtered if e.get("isDeductible")
    )

    # Group by supplier for subtotals
    by_supplier = defaultdict(lambda: {"count": 0, "monto": 0.0, "itbis": 0.0})
    for e in filtered:
        key = e.get("supplierId") or e.get("rncEmisor") or "sin-id"
        by_supplier[key]["count"] += 1
        by_supplier[key]["monto"] += float(e.get("amount", 0))
        by_supplier[key]["itbis"] += float(e.get("itbisAmount", 0))
        if not by_supplier[key].get("name"):
            by_supplier[key]["name"] = e.get("providerName", "—")
        if not by_supplier[key].get("rnc"):
            by_supplier[key]["rnc"] = e.get("rncEmisor", "—")

    return render_template(
        "reports/reporte_606.html",
        filtered=filtered,
        by_supplier=dict(by_supplier),
        year=year,
        month=month,
        supplier_id=supplier_id,
        tipo_gasto=tipo_gasto,
        ecf_type=ecf_type,
        search=search,
        total_monto=total_monto,
        total_itbis=total_itbis,
        total_itbis_deducible=total_itbis_deducible,
        total_deducible=total_deducible,
        tipo_gasto_map=TIPO_GASTO_606,
        ecf_labels=ECF_TYPE_LABELS,
        active_page="reporte_606",
        now=now,
    )


# ═════════════════════════════════════════════════════════════════════
# REPORTE 606 — Export (XLSX / TXT)
# ═════════════════════════════════════════════════════════════════════


@web_reports_606_bp.route("/reports/606/export")
def reporte_606_export():
    if "user" not in session:
        return redirect(url_for("login"))
    if not check_permission("canExpenses"):
        return render_template("auth/restricted.html"), 403
    owner_uid = session["user"]["ownerUID"]
    sandbox = session.get("is_sandbox_mode", True)

    now = datetime.now(timezone.utc)
    year = _parse_int(request.args.get("year"), now.year)
    month = _parse_int(request.args.get("month"), now.month)
    supplier_id = request.args.get("supplier_id", "").strip()
    tipo_gasto = request.args.get("tipo_gasto", "").strip()
    ecf_type = request.args.get("ecf_type", "").strip()
    search = request.args.get("search", "").strip()
    fmt = request.args.get("format", "dgii-xlsx")

    expenses = _get_expenses_for_period(owner_uid, sandbox, year, month)
    filtered = _filter_expenses(expenses, supplier_id, tipo_gasto, ecf_type, search)
    filtered.sort(key=lambda x: (x.get("date") or x.get("createdAt") or ""))

    profile = DatabaseService.get_company_profile(owner_uid)
    owner_rnc = (profile or {}).get("companyRNC", "").replace("-", "")
    period = f"{year:04d}{month:02d}"

    if fmt == "dgii-txt":
        content = _generate_606_txt(owner_rnc, period, filtered)
        dest = io.BytesIO()
        dest.write(content.encode("utf-8"))
        dest.seek(0)
        filename = f"DGII_F_606_{owner_rnc}_{period}.TXT"
        return send_file(
            dest,
            mimetype="text/plain",
            as_attachment=True,
            download_name=filename,
        )

    dest = _generate_606_xlsx(owner_rnc, period, filtered)
    filename = f"DGII_F_606_{owner_rnc}_{period}.xlsx"
    return send_file(
        dest,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ═════════════════════════════════════════════════════════════════════
# REPORTE 606 — KPIs Dashboard
# ═════════════════════════════════════════════════════════════════════


@web_reports_606_bp.route("/reports/606/dashboard")
def reporte_606_dashboard():
    if "user" not in session:
        return redirect(url_for("login"))
    if not check_permission("canExpenses"):
        return render_template("auth/restricted.html", active_page="dashboard_606")
    owner_uid = session["user"]["ownerUID"]
    sandbox = session.get("is_sandbox_mode", True)

    now = datetime.now(timezone.utc)
    month = _parse_int(request.args.get("month"), now.month)
    year = _parse_int(request.args.get("year"), now.year)

    expenses = _get_expenses_for_period(owner_uid, sandbox, year, month)

    total_monto = sum(float(e.get("amount", 0)) for e in expenses)
    total_itbis = sum(float(e.get("itbisAmount", 0)) for e in expenses)
    total_itbis_deducible = sum(
        float(e.get("itbisAmount", 0)) for e in expenses if e.get("isITBISDeductible")
    )
    total_deducible = sum(
        float(e.get("amount", 0)) for e in expenses if e.get("isDeductible")
    )
    count = len(expenses)

    unique_suppliers = set()
    for e in expenses:
        sid = e.get("supplierId") or e.get("rncEmisor")
        if sid:
            unique_suppliers.add(sid)

    # Top 10 suppliers
    by_supplier = defaultdict(lambda: {"monto": 0.0, "itbis": 0.0, "count": 0})
    for e in expenses:
        key = e.get("supplierId") or e.get("rncEmisor") or "sin-id"
        by_supplier[key]["monto"] += float(e.get("amount", 0))
        by_supplier[key]["itbis"] += float(e.get("itbisAmount", 0))
        by_supplier[key]["count"] += 1
        if not by_supplier[key].get("name"):
            by_supplier[key]["name"] = e.get("providerName", "Sin nombre")
    top_suppliers = sorted(by_supplier.items(), key=lambda x: x[1]["monto"], reverse=True)[:10]

    # Tipo gasto breakdown
    by_tipo = defaultdict(lambda: {"monto": 0.0, "count": 0})
    for e in expenses:
        tg = e.get("tipoGastoDGII", "06")
        by_tipo[tg]["monto"] += float(e.get("amount", 0))
        by_tipo[tg]["count"] += 1
    tipo_breakdown = [{"code": k, "label": TIPO_GASTO_606.get(k, k), "monto": v["monto"], "count": v["count"]} for k, v in by_tipo.items()]
    tipo_breakdown.sort(key=lambda x: x["monto"], reverse=True)

    # e-CF type distribution
    by_ecf = defaultdict(lambda: {"monto": 0.0, "count": 0})
    for e in expenses:
        et = e.get("ecfType", "Otro")
        by_ecf[et]["monto"] += float(e.get("amount", 0))
        by_ecf[et]["count"] += 1
    ecf_dist = [{"code": k, "label": ECF_TYPE_LABELS.get(k, k), "monto": v["monto"], "count": v["count"]} for k, v in by_ecf.items()]
    ecf_dist.sort(key=lambda x: x["monto"], reverse=True)

    return render_template(
        "reports/dashboard_606.html",
        year=year,
        month=month,
        total_monto=total_monto,
        total_itbis=total_itbis,
        total_itbis_deducible=total_itbis_deducible,
        total_deducible=total_deducible,
        count=count,
        supplier_count=len(unique_suppliers),
        top_suppliers=top_suppliers,
        tipo_breakdown=tipo_breakdown,
        ecf_dist=ecf_dist,
        now=now,
        active_page="dashboard_606",
    )
