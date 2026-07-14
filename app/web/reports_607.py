import io
from datetime import datetime, timezone
from collections import defaultdict
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, g
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from app.services.db_service import DatabaseService
from app.utils.decorators import check_permission

web_reports_607_bp = Blueprint('web_reports_607', __name__)


@web_reports_607_bp.before_request
def restrict_to_do():
    if session.get('company_country', 'DO') != 'DO':
        return render_template('auth/restricted.html',
            feature_name="Reporte 607 DGII (solo disponible para República Dominicana)",
            required_permission="")

from app.models.fiscal_document_type import report_labels as _report_labels_607
ECF_TYPE_LABELS_607 = _report_labels_607("607")

INCOME_TYPE_LABELS = {
    "01": "Ingresos por operaciones (No financieros)",
    "02": "Ingresos financieros",
    "03": "Ingresos extraordinarios",
    "04": "Ingresos por arrendamientos",
    "05": "Ingresos por venta de activo depreciable",
    "06": "Otros ingresos",
}


def _parse_int(val, default=0):
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def _get_invoices_for_period(owner_uid, sandbox, year, month):
    invoices = DatabaseService.get_invoices(owner_uid, sandbox=sandbox)
    real = [inv for inv in invoices
            if not inv.get('isQuotation')
            and inv.get('status') not in ('Anulada', 'Borrador', 'Consolidada')
            and inv.get('dgiiStatus', '') in ('ACCEPTED', 'ACCEPTED_CONDITIONAL')]
    prefix = f"{year:04d}-{month:02d}"
    filtered = []
    for inv in real:
        d = (inv.get("date") or inv.get("createdAt") or "")[:7]
        if d == prefix:
            filtered.append(inv)
    return filtered


def _filter_invoices(invoices, client_id, ecf_type, income_type, search):
    result = invoices
    if client_id:
        result = [e for e in result if e.get("clientId") == client_id]
    if ecf_type:
        ecf_short = ecf_type.split("(")[-1].replace(")", "").strip() if "(" in ecf_type else ecf_type
        result = [e for e in result if (e.get("ecfType") or "").upper().startswith(ecf_short)]
    if income_type:
        result = [e for e in result if (e.get("incomeType") or "").startswith(income_type)]
    if search:
        q = search.lower()
        result = [
            e for e in result
            if q in (e.get("clientName") or "").lower()
            or q in (e.get("clientRNC") or "").lower()
            or q in (e.get("encf") or e.get("invoiceNumber") or "").lower()
        ]
    return result


# ═════════════════════════════════════════════════════════════════════
# REPORTE 607 — Main View
# ═════════════════════════════════════════════════════════════════════


@web_reports_607_bp.route("/reports/607")
def reporte_607():
    if "user" not in session:
        return redirect(url_for("login"))
    if not check_permission("canInvoice"):
        return render_template("auth/restricted.html", active_page="reporte_607")
    owner_uid = session["user"]["ownerUID"]
    sandbox = session.get("is_sandbox_mode", True)

    now = datetime.now(timezone.utc)
    year = _parse_int(request.args.get("year"), now.year)
    month = _parse_int(request.args.get("month"), now.month)
    client_id = request.args.get("client_id", "").strip()
    ecf_type = request.args.get("ecf_type", "").strip()
    income_type = request.args.get("income_type", "").strip()
    search = request.args.get("search", "").strip()

    invoices = _get_invoices_for_period(owner_uid, sandbox, year, month)
    filtered = _filter_invoices(invoices, client_id, ecf_type, income_type, search)
    filtered.sort(key=lambda x: (x.get("date") or x.get("createdAt") or ""))

    total_monto = sum(float(e.get("subtotal", 0)) for e in filtered)
    total_itbis = sum(float(e.get("totalITBIS", 0)) for e in filtered)
    total_neto = sum(float(e.get("total", 0)) for e in filtered)
    total_retenciones = sum(float(e.get("retainedISR", 0)) + float(e.get("retainedITBIS", 0)) for e in filtered)

    # Group by client
    by_client = defaultdict(lambda: {"count": 0, "monto": 0.0, "itbis": 0.0, "name": "", "rnc": ""})
    for e in filtered:
        key = e.get("clientId") or e.get("clientRNC") or "sin-id"
        by_client[key]["count"] += 1
        by_client[key]["monto"] += float(e.get("subtotal", 0))
        by_client[key]["itbis"] += float(e.get("totalITBIS", 0))
        if not by_client[key].get("name"):
            by_client[key]["name"] = e.get("clientName", "Consumidor Final")
        if not by_client[key].get("rnc"):
            by_client[key]["rnc"] = e.get("clientRNC", "—")

    clients_list = DatabaseService.get_clients(owner_uid, sandbox=sandbox, branch_id=g.get('branch_id'), project_id=g.get('project_id')) or []

    return render_template(
        "reports/reporte_607.html",
        filtered=filtered,
        by_client=dict(by_client),
        year=year,
        month=month,
        client_id=client_id,
        ecf_type=ecf_type,
        income_type=income_type,
        search=search,
        total_monto=total_monto,
        total_itbis=total_itbis,
        total_neto=total_neto,
        total_retenciones=total_retenciones,
        ecf_labels=ECF_TYPE_LABELS_607,
        income_labels=INCOME_TYPE_LABELS,
        clients_list=clients_list,
        active_page="reporte_607",
        now=now,
    )


# ═════════════════════════════════════════════════════════════════════
# REPORTE 607 — XLS Export
# ═════════════════════════════════════════════════════════════════════


@web_reports_607_bp.route("/reports/607/export")
def reporte_607_export():
    if "user" not in session:
        return redirect(url_for("login"))
    if not check_permission("canInvoice"):
        return render_template("auth/restricted.html"), 403
    owner_uid = session["user"]["ownerUID"]
    sandbox = session.get("is_sandbox_mode", True)

    now = datetime.now(timezone.utc)
    year = _parse_int(request.args.get("year"), now.year)
    month = _parse_int(request.args.get("month"), now.month)
    client_id = request.args.get("client_id", "").strip()
    ecf_type = request.args.get("ecf_type", "").strip()
    income_type = request.args.get("income_type", "").strip()
    search = request.args.get("search", "").strip()
    fmt = request.args.get("format", "simple")

    invoices = _get_invoices_for_period(owner_uid, sandbox, year, month)
    filtered = _filter_invoices(invoices, client_id, ecf_type, income_type, search)
    filtered.sort(key=lambda x: (x.get("date") or x.get("createdAt") or ""))

    profile = DatabaseService.get_company_profile(owner_uid)
    owner_rnc = (profile or {}).get("companyRNC", "").replace("-", "")

    wb = Workbook()

    if fmt == "dgii":
        ws = wb.active
        ws.title = "Herramienta Formato 607"

        header_font = Font(name="Calibri", bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        money_fmt = '#,##0.00'

        period = f"{year:04d}{month:02d}"

        ws.merge_cells("A1:Y1")
        ws["A1"] = "Formato de Envío de Ventas de Bienes y Servicios"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12)

        ws.merge_cells("A2:Y2")
        ws["A2"] = "Herramienta de Distribución Gratuita — Derechos Reservados DGII 2023"
        ws["A2"].font = Font(name="Calibri", size=9, italic=True)

        ws["A4"] = "RNC o Cédula"
        ws["B4"] = owner_rnc
        ws["A5"] = "Periodo"
        ws["B5"] = period
        ws["A6"] = "Cantidad Registros"
        ws["B6"] = len(filtered)

        dgii_headers = [
            "No.",
            "RNC/Cédula o Pasaporte",
            "Tipo Identificación",
            "Número Comprobante Fiscal",
            "Número Comprobante Fiscal Modificado",
            "Tipo de Ingreso",
            "Fecha Comprobante",
            "Fecha de Retención",
            "Monto Facturado",
            "ITBIS Facturado",
            "ITBIS Retenido por Terceros",
            "ITBIS Percibido",
            "Retención Renta por Terceros",
            "ISR Percibido",
            "Impuesto Selectivo al Consumo",
            "Otros Impuestos/Tasas",
            "Monto Propina Legal",
            "Efectivo",
            "Cheque/ Transferencia/ Depósito",
            "Tarjeta Débito/Crédito",
            "Venta a Crédito",
            "Bonos o Certificados de Regalo",
            "Permuta",
            "Otras Formas de Ventas",
            "Estatus",
        ]

        numbering_row = 10
        numbering_font = Font(name="Calibri", size=9)
        for num, col_idx in enumerate(range(2, 25), 1):
            cell = ws.cell(row=numbering_row, column=col_idx, value=num)
            cell.font = numbering_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        header_row = 11
        for col_idx, h in enumerate(dgii_headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        data_start_row = header_row + 1
        for i, e in enumerate(filtered):
            row = data_start_row + i
            seq = i + 1

            rnc = (e.get("clientRNC") or "").replace("-", "")
            rnc_len = len(rnc) if rnc else 0
            if rnc_len == 9:
                tipo_id = 1
            elif rnc_len == 11:
                tipo_id = 2
            else:
                tipo_id = 3 if rnc else ""

            ncf = e.get("encf") or e.get("invoiceNumber") or ""
            ncf_modified = e.get("ncfModified") or ""

            date_str = (e.get("date") or "")[:10].replace("-", "")

            retention_date = (e.get("retentionDate") or "").replace("-", "")
            if not retention_date:
                has_retention = float(e.get("retainedISR", 0)) > 0 or float(e.get("retainedITBIS", 0)) > 0
                if has_retention:
                    retention_date = date_str

            subtotal = float(e.get("subtotal", 0))
            total_itbis = float(e.get("totalITBIS", 0))
            itbis_retenido = float(e.get("retainedITBIS", 0))
            isr_retenido = float(e.get("retainedISR", 0))
            isc = float(e.get("totalISCEspecifico", 0)) + float(e.get("totalISCAdValorem", 0))
            otros_impuestos = float(e.get("totalOtrosImpuestos", 0) or 0)
            propina = float(e.get("propinaLegal", 0))

            total_factura = e.get("total", 0)
            if not total_factura or float(total_factura) == 0:
                total_factura = subtotal + total_itbis + isc + otros_impuestos + propina

            pm = (e.get("paymentMethod") or "").lower()
            efectivo = float(total_factura) if "efectivo" in pm else 0.00
            transferencia = float(total_factura) if pm in ("transferencia", "cheque", "deposito") or any(k in pm for k in ("transfer", "cheque", "deposito")) else 0.00
            tarjeta = float(total_factura) if "tarjeta" in pm else 0.00
            credito = float(total_factura) if "credito" in pm else 0.00

            row_data = [
                seq,
                rnc,
                tipo_id,
                ncf,
                ncf_modified,
                (e.get("incomeType") or "01")[:2],
                date_str,
                retention_date,
                subtotal,
                total_itbis,
                itbis_retenido,
                0.00,
                isr_retenido,
                0.00,
                isc,
                otros_impuestos,
                propina,
                efectivo,
                transferencia,
                tarjeta,
                credito,
                0.00,
                0.00,
                0.00,
                "",
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in (9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in (2, 4, 5, 7, 8):
                    cell.alignment = Alignment(horizontal="left")
                elif col_idx in (3, 6):
                    cell.alignment = Alignment(horizontal="center")

        col_widths = {
            1: 6, 2: 24, 3: 14, 4: 20, 5: 20, 6: 14,
            7: 15, 8: 15, 9: 15, 10: 15, 11: 15, 12: 15,
            13: 15, 14: 15, 15: 15, 16: 15, 17: 15,
            18: 15, 19: 15, 20: 15, 21: 15, 22: 15,
            23: 15, 24: 15, 25: 15,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

    else:
        ws = wb.active
        ws.title = "Reporte 607"

        header_font = Font(name="Calibri", bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        money_fmt = '#,##0.00'

        ws["A1"] = f"Reporte 607 — {year:04d}-{month:02d}"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12)

        ws["A3"] = "RNC Emisor"
        ws["B3"] = owner_rnc
        ws["A4"] = "Periodo"
        ws["B4"] = f"{year:04d}{month:02d}"
        ws["A5"] = "Registros"
        ws["B5"] = len(filtered)

        simple_headers = [
            "Fecha", "RNC Cliente", "Cliente", "e-CF", "NCF/e-CF",
            "Monto", "ITBIS", "ISR Ret.", "ITBIS Ret.", "Neto", "Tipo Ingreso",
        ]

        header_row = 7
        for col_idx, h in enumerate(simple_headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        data_start_row = header_row + 1
        for i, e in enumerate(filtered):
            row = data_start_row + i
            row_data = [
                (e.get("date") or "")[:10],
                (e.get("clientRNC") or "").replace("-", ""),
                e.get("clientName", ""),
                e.get("ecfType", ""),
                e.get("encf") or e.get("invoiceNumber") or "",
                float(e.get("subtotal", 0)),
                float(e.get("totalITBIS", 0)),
                float(e.get("retainedISR", 0)),
                float(e.get("retainedITBIS", 0)),
                float(e.get("total", 0)),
                e.get("incomeType", "01"),
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in (6, 7, 8, 9, 10):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right")

    dest = io.BytesIO()
    wb.save(dest)
    dest.seek(0)

    filename = f"reporte_607_{year:04d}{month:02d}_{now.strftime('%Y%m%d')}.xlsx"
    return send_file(
        dest,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
