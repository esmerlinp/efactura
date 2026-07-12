import io
import csv
from datetime import datetime, timezone
from collections import defaultdict
from flask import Blueprint, render_template, request, redirect, url_for, session, send_file, g
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from app.services.db_service import DatabaseService
from app.utils.decorators import check_permission

web_reports_608_bp = Blueprint('web_reports_608', __name__)


@web_reports_608_bp.before_request
def restrict_to_do():
    if session.get('company_country', 'DO') != 'DO':
        return render_template('auth/restricted.html',
            feature_name="Reporte 608 DGII (solo disponible para República Dominicana)",
            required_permission="")

CANCELLATION_TYPE_LABELS = {
    "01": "Deterioro de Factura Pre-Impresa",
    "02": "Errores de Impresión (Factura Pre-Impresa)",
    "03": "Impresión Defectuosa",
    "04": "Corrección de la Información",
    "05": "Cambio de Productos",
    "06": "Devolución de Productos",
    "07": "Omisión de Productos",
    "08": "Errores en Secuencias de NCF",
    "09": "Por Cese de Operaciones",
    "10": "Pérdida O Hurto De Talonario(S)",
}


def _parse_int(val, default=0):
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


def _get_annulled_invoices(owner_uid, sandbox, year, month):
    prefix = f"{year:04d}-{month:02d}"
    invoices = DatabaseService.get_invoices(owner_uid, sandbox=sandbox,
        branch_id=g.get('branch_id'), project_id=g.get('project_id'))
    filtered = []
    for inv in invoices:
        if inv.get('status') != 'Anulada':
            continue
        d = (inv.get('cancelledAt') or inv.get('updatedAt') or inv.get('date') or inv.get('createdAt') or "")[:7]
        if d == prefix:
            filtered.append(inv)
    return filtered


def _filter_annulled(invoices, cancellation_type, search):
    result = invoices
    if cancellation_type:
        result = [e for e in result if e.get("cancellationType", "") == cancellation_type]
    if search:
        q = search.lower()
        result = [
            e for e in result
            if q in (e.get("clientName") or "").lower()
            or q in (e.get("clientRNC") or "").lower()
            or q in (e.get("encf") or e.get("invoiceNumber") or "").lower()
        ]
    return result


@web_reports_608_bp.route("/reports/608")
def reporte_608():
    if "user" not in session:
        return redirect(url_for("login"))
    if not check_permission("canReports"):
        return render_template("auth/restricted.html", active_page="reporte_608")
    owner_uid = session["user"]["ownerUID"]
    sandbox = session.get("is_sandbox_mode", True)

    now = datetime.now(timezone.utc)
    year = _parse_int(request.args.get("year"), now.year)
    month = _parse_int(request.args.get("month"), now.month)
    cancellation_type = request.args.get("cancellation_type", "").strip()
    search = request.args.get("search", "").strip()

    invoices = _get_annulled_invoices(owner_uid, sandbox, year, month)
    filtered = _filter_annulled(invoices, cancellation_type, search)
    filtered.sort(key=lambda x: (x.get("cancelledAt") or x.get("updatedAt") or x.get("date") or x.get("createdAt") or ""))

    total_count = len(filtered)

    by_type = defaultdict(int)
    for e in filtered:
        ct = e.get("cancellationType", "04")
        by_type[ct] += 1

    profile = DatabaseService.get_company_profile(owner_uid)
    owner_rnc = (profile or {}).get("companyRNC", "")

    return render_template(
        "reports/reporte_608.html",
        filtered=filtered,
        by_type=dict(by_type),
        year=year,
        month=month,
        cancellation_type=cancellation_type,
        search=search,
        total_count=total_count,
        cancellation_labels=CANCELLATION_TYPE_LABELS,
        owner_rnc=owner_rnc,
        active_page="reporte_608",
        now=now,
    )


def _build_xlsx(owner_rnc, period, filtered):
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato 608"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_font = Font(name='Calibri', size=11, bold=True)
    title_font = Font(name='Calibri', size=12, bold=True)
    normal_font = Font(name='Calibri', size=10)
    mono_font = Font(name='Consolas', size=9)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    col_widths = {'A': 7, 'B': 30, 'C': 14, 'D': 12, 'E': 14, 'F': 14, 'G': 14, 'H': 5, 'I': 14}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for r in range(1, 4):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = normal_font

    ws.merge_cells('A1:G1')
    ws.cell(row=1, column=1, value="DIRECCIÓN GENERAL DE IMPUESTOS INTERNOS").font = Font(name='Calibri', size=13, bold=True)
    ws.cell(row=1, column=1).alignment = center_align

    ws.merge_cells('A2:G2')
    ws.cell(row=2, column=1, value="FORMATO DE ENVÍO DE COMPROBANTES FISCALES ANULADOS").font = title_font
    ws.cell(row=2, column=1).alignment = center_align

    ws.merge_cells('A3:G3')
    ws.cell(row=3, column=1, value="Reporte 608 - DGII").font = Font(name='Calibri', size=10, italic=True)
    ws.cell(row=3, column=1).alignment = center_align

    ws.merge_cells('A5:B5')
    ws.cell(row=5, column=1, value="RNC o Cédula:").font = header_font
    ws.cell(row=5, column=1).alignment = left_align
    ws.merge_cells('C5:G5')
    ws.cell(row=5, column=3, value=owner_rnc).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=5, column=3).alignment = left_align

    ws.merge_cells('A6:B6')
    ws.cell(row=6, column=1, value="Periodo:").font = header_font
    ws.cell(row=6, column=1).alignment = left_align
    ws.merge_cells('C6:G6')
    ws.cell(row=6, column=3, value=period).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=6, column=3).alignment = left_align

    ws.merge_cells('A7:B7')
    ws.cell(row=7, column=1, value="Cantidad Registros:").font = header_font
    ws.cell(row=7, column=1).alignment = left_align
    ws.cell(row=7, column=3, value=len(filtered)).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=7, column=3).alignment = center_align
    ws.cell(row=7, column=3).border = thin_border

    ws.merge_cells('E7:F7')
    ws.cell(row=7, column=5, value="Total Errores:").font = header_font
    ws.cell(row=7, column=5).alignment = left_align
    ws.cell(row=7, column=7, value=0).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=7, column=7).alignment = center_align
    ws.cell(row=7, column=7).border = thin_border

    ws.merge_cells('B9:F9')
    ws.cell(row=9, column=2, value="DETALLE").font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=9, column=2).alignment = center_align
    ws.cell(row=9, column=2).fill = header_fill

    ws.cell(row=10, column=2, value=1).font = Font(name='Calibri', size=9, bold=True)
    ws.cell(row=10, column=2).alignment = center_align
    ws.cell(row=10, column=2).fill = yellow_fill
    ws.merge_cells('B10:C10')
    ws.cell(row=10, column=4, value=2).font = Font(name='Calibri', size=9, bold=True)
    ws.cell(row=10, column=4).alignment = center_align
    ws.cell(row=10, column=4).fill = yellow_fill
    ws.merge_cells('E10:F10')
    ws.cell(row=10, column=5, value=3).font = Font(name='Calibri', size=9, bold=True)
    ws.cell(row=10, column=5).alignment = center_align
    ws.cell(row=10, column=5).fill = yellow_fill

    ws.cell(row=11, column=1, value="Líneas").font = header_font
    ws.cell(row=11, column=1).alignment = center_align
    ws.cell(row=11, column=1).fill = header_fill
    ws.cell(row=11, column=1).border = thin_border
    ws.merge_cells('B11:C11')
    ws.cell(row=11, column=2, value="Número de Comprobante Fiscal").font = header_font
    ws.cell(row=11, column=2).alignment = center_align
    ws.cell(row=11, column=2).fill = header_fill
    ws.cell(row=11, column=2).border = thin_border
    ws.cell(row=11, column=3).border = thin_border
    ws.cell(row=11, column=4, value="Fecha de Comprobante").font = header_font
    ws.cell(row=11, column=4).alignment = center_align
    ws.cell(row=11, column=4).fill = header_fill
    ws.cell(row=11, column=4).border = thin_border
    ws.merge_cells('E11:F11')
    ws.cell(row=11, column=5, value="Tipo de Anulación").font = header_font
    ws.cell(row=11, column=5).alignment = center_align
    ws.cell(row=11, column=5).fill = header_fill
    ws.cell(row=11, column=5).border = thin_border
    ws.cell(row=11, column=6).border = thin_border
    ws.cell(row=11, column=7, value="Estatus").font = header_font
    ws.cell(row=11, column=7).alignment = center_align
    ws.cell(row=11, column=7).fill = header_fill
    ws.cell(row=11, column=7).border = thin_border

    for idx, inv in enumerate(filtered):
        row = 12 + idx
        line_num = idx + 1
        ncf = inv.get("encf") or inv.get("invoiceNumber") or ""
        date_val = (inv.get("cancelledAt") or inv.get("updatedAt") or inv.get("date") or "")[:10]
        canc_type = inv.get("cancellationType", "04")
        canc_display = f"{canc_type} - {CANCELLATION_TYPE_LABELS.get(canc_type, '')}"

        ws.cell(row=row, column=1, value=line_num).font = mono_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        ws.merge_cells(f'B{row}:C{row}')
        ws.cell(row=row, column=2, value=ncf).font = mono_font
        ws.cell(row=row, column=2).alignment = center_align
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value=date_val).font = normal_font
        ws.cell(row=row, column=4).alignment = center_align
        ws.cell(row=row, column=4).border = thin_border
        ws.merge_cells(f'E{row}:F{row}')
        ws.cell(row=row, column=5, value=canc_display).font = normal_font
        ws.cell(row=row, column=5).alignment = left_align
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=7, value="Anulado").font = normal_font
        ws.cell(row=row, column=7).alignment = center_align
        ws.cell(row=row, column=7).border = thin_border

    ref_row = max(12 + len(filtered) + 2, 14)
    ws.merge_cells(f'I{ref_row}:I{ref_row}')
    ws.cell(row=ref_row, column=9, value="CÓDIGOS DE ANULACIÓN DGII").font = Font(name='Calibri', size=10, bold=True)
    for i, (code, label) in enumerate(CANCELLATION_TYPE_LABELS.items()):
        r = ref_row + 1 + i
        ws.cell(row=r, column=9, value=f"{code} {label}").font = Font(name='Calibri', size=9)

    wb.properties.creator = "VykOne ERP"
    wb.properties.title = "Formato 608 - Comprobantes Anulados"

    return wb


@web_reports_608_bp.route("/reports/608/export")
def reporte_608_export():
    if "user" not in session:
        return redirect(url_for("login"))
    if not check_permission("canReports"):
        return render_template("auth/restricted.html"), 403
    owner_uid = session["user"]["ownerUID"]
    sandbox = session.get("is_sandbox_mode", True)

    now = datetime.now(timezone.utc)
    year = _parse_int(request.args.get("year"), now.year)
    month = _parse_int(request.args.get("month"), now.month)
    cancellation_type = request.args.get("cancellation_type", "").strip()
    search = request.args.get("search", "").strip()
    fmt = request.args.get("format", "simple")

    invoices = _get_annulled_invoices(owner_uid, sandbox, year, month)
    filtered = _filter_annulled(invoices, cancellation_type, search)
    filtered.sort(key=lambda x: (x.get("cancelledAt") or x.get("updatedAt") or x.get("date") or x.get("createdAt") or ""))

    profile = DatabaseService.get_company_profile(owner_uid)
    owner_rnc = (profile or {}).get("companyRNC", "")

    if fmt == "xlsx":
        period = f"{year:04d}{month:02d}"
        wb = _build_xlsx(owner_rnc, period, filtered)
        dest = io.BytesIO()
        wb.save(dest)
        dest.seek(0)
        filename = f"reporte_608_{year:04d}{month:02d}_{now.strftime('%Y%m%d')}.xlsx"
        return send_file(
            dest,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)

    if fmt == "dgii":
        writer.writerow([
            "RNC Emisor",
            "Periodo",
            "NCF/e-CF",
            "Fecha de Comprobante",
            "Tipo Anulación (Código)",
            "Tipo Anulación",
            "Estatus",
        ])
        period = f"{year:04d}{month:02d}"
        for e in filtered:
            canc_type = e.get("cancellationType", "04")
            writer.writerow([
                owner_rnc,
                period,
                e.get("encf") or e.get("invoiceNumber") or "",
                (e.get("cancelledAt") or e.get("updatedAt") or e.get("date") or "")[:10],
                canc_type,
                CANCELLATION_TYPE_LABELS.get(canc_type, ""),
                "Anulado",
            ])
    else:
        writer.writerow([
            "Fecha Anulación", "NCF/e-CF", "RNC Cliente", "Nombre Cliente",
            "Tipo Anulación", "Estatus",
        ])
        for e in filtered:
            canc_type = e.get("cancellationType", "04")
            writer.writerow([
                (e.get("cancelledAt") or e.get("updatedAt") or e.get("date") or "")[:10],
                e.get("encf") or e.get("invoiceNumber") or "",
                e.get("clientRNC", ""),
                e.get("clientName", ""),
                CANCELLATION_TYPE_LABELS.get(canc_type, ""),
                "Anulado",
            ])
        writer.writerow([])
        writer.writerow(["TOTAL ANULADOS", str(len(filtered)), "", "", "", ""])

    dest = io.BytesIO()
    dest.write(b"\xef\xbb\xbf")
    dest.write(output.getvalue().encode("utf-8"))
    dest.seek(0)

    filename = f"reporte_608_{year:04d}{month:02d}_{now.strftime('%Y%m%d')}.csv"
    return send_file(
        dest,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )
