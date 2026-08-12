#!/usr/bin/env python3
"""
Orquestador de Certificacion DGII - Fase II (Pruebas de Datos e-CF)

Lee el Set de Pruebas de DGII (Excel), genera cada e-CF, lo firma
digitalmente y lo envia a los endpoints de recepcion de DGII en sandbox.

Uso:
    python scripts/run_phase2_certification.py \\
        --excel test_data.xlsx \\
        --owner-uid <OWNER_UID> \\
        --company-id <COMPANY_ID>

Los parametros se obtienen de Firestore:
  - owner-uid: ID del usuario dueno de la empresa
  - company-id: ID de la empresa en Firestore

El script:
  1. Carga el perfil de la empresa con su certificado digital
  2. Lee el Excel con los 25 casos de prueba
  3. Para cada caso: construye XML, firma, envia a DGII
  4. Para E32 < RD$250,000: envia primero a RFCE, luego carga factura
  5. Consulta el estado de cada envio
  6. Genera un archivo JSON con los resultados
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import create_app
from app.services.db_service import DatabaseService
from app.services.dgii_signer import DgiiSigner
from app.services.dgii_test_data_loader import DgiiTestDataLoader

RFCE_THRESHOLD = 250000.00


def resolve_env():
    """Determina el entorno de ejecucion."""
    env = os.getenv("APP_ENVIRONMENT", "sandbox").lower()
    sandbox = env != "production"
    signing_mode = os.getenv("DGII_SIGNING_MODE", "real" if not sandbox else "real")
    return sandbox, signing_mode


def build_endpoints(sandbox):
    """Construye URLs de DGII segun entorno."""
    env = os.getenv("DGII_ENVIRONMENT", "testecf").lower()
    base_ecf = f"https://ecf.dgii.gov.do/{env}"
    base_fc = f"https://fc.dgii.gov.do/{env}"

    return {
        "auth_semilla": f"{base_ecf}/autenticacion/api/autenticacion/semilla",
        "auth_validar": f"{base_ecf}/autenticacion/api/autenticacion/validarsemilla",
        "recepcion": f"{base_ecf}/recepcion/api/facturaselectronicas",
        "rfce_recepcion": f"{base_fc}/recepcionfc/api/recepcion/ecf",
        "consulta": f"{base_ecf}/consultaresultado/api/consultas/estado",
        "acecf": f"{base_ecf}/aprobacioncomercial/api/aprobacioncomercial",
    }


def get_token(company_profile, sandbox):
    """
    Autenticacion CerteCF segun XSD Semilla v.1.0:
    GET  semilla → <SemillaModel><valor>...</valor><fecha>...</fecha></SemillaModel>
    POST validarsemilla → <SemillaModel> firmado con XMLDSig
    """
    from app.services.dgii_direct import DgiiDirectService
    import requests, re

    env = os.getenv("DGII_ENVIRONMENT", "testecf").lower()
    semilla_url = f"https://ecf.dgii.gov.do/{env}/autenticacion/api/autenticacion/semilla"
    validar_url = f"https://ecf.dgii.gov.do/{env}/autenticacion/api/autenticacion/validarsemilla"

    print(f"  Semilla: {semilla_url}")
    print(f"  Validar: {validar_url}")

    cert_path = DgiiDirectService._prepare_tls_cert(company_profile)
    try:
        headers = {"accept": "application/json", "User-Agent": os.getenv("DGII_USER_AGENT", "VykOne/1.0")}

        # ── 1. GET semilla ──
        print("  [1] GET semilla...")
        resp = requests.get(semilla_url, headers=headers, cert=cert_path, timeout=30)
        print(f"      HTTP {resp.status_code}")

        if resp.status_code >= 400:
            raise RuntimeError(f"Semilla HTTP {resp.status_code}: {resp.text[:200]}")

        # Extract valor and fecha
        semilla_xml_raw = resp.text
        valor = re.search(r"<valor[^>]*>([^<]+)</valor>", semilla_xml_raw, re.IGNORECASE)
        fecha = re.search(r"<fecha[^>]*>([^<]+)</fecha>", semilla_xml_raw, re.IGNORECASE)
        if not valor or not fecha:
            raise RuntimeError(f"No se encontro valor/fecha en: {semilla_xml_raw[:200]}")
        valor_val = valor.group(1).strip()
        fecha_val = fecha.group(1).strip()
        print(f"      valor={valor_val[:30]}...  fecha={fecha_val}")

        # ── 2. Construir SemillaModel XML ──
        semilla_model_xml = (
            '<?xml version="1.0" encoding="utf-8"?>'
            f'<SemillaModel>'
            f'<valor>{valor_val}</valor>'
            f'<fecha>{fecha_val}</fecha>'
            f'</SemillaModel>'
        ).encode("utf-8")

        # ── 3. Firmar SemillaModel con XMLDSig ──
        print("  [2] Firmando SemillaModel con XMLDSig...")
        signed_semilla = DgiiSigner.sign_xml(semilla_model_xml, company_profile)
        print(f"      XML firmado: {len(signed_semilla)} bytes")

        # ── 4. POST validarsemilla ──
        print("  [3] POST validarsemilla...")
        files = {"xml": ("signed_seed.xml", signed_semilla, "text/xml")}
        resp2 = requests.post(validar_url, files=files, headers=headers, cert=cert_path, timeout=30)
        print(f"      HTTP {resp2.status_code} | {resp2.text[:400]}")

        if resp2.status_code >= 400:
            raise RuntimeError(f"Validar HTTP {resp2.status_code}: {resp2.text[:300]}")

        # Extract token
        text2 = resp2.text
        token_match = re.search(r"eyJ[a-zA-Z0-9_-]+\.[a-zA-Z0-9_-]+\.[a-zA-Z0-9_-]+", text2)
        if token_match:
            return token_match.group(0)

        try:
            d = resp2.json()
            for k in ("token", "Token", "jwt", "access_token"):
                if d.get(k):
                    return d[k]
        except:
            pass

        for tag in ("Token", "token"):
            m = re.search(rf"<{tag}[^>]*>([^<]+)</{tag}>", text2, re.IGNORECASE)
            if m:
                return m.group(1).strip()

        raise RuntimeError(f"No token en respuesta: {text2[:300]}")

    finally:
        DgiiDirectService._cleanup_tls_cert(cert_path)


def send_to_dgii(signed_xml, token, company_profile, invoice_row, endpoints, encf, rfce=False):
    """Envia el XML firmado a DGII y retorna la respuesta."""
    from app.services.dgii_direct import DgiiDirectService
    import requests
    import uuid

    company_rnc = str(company_profile.get("companyRNC", "")).replace("-", "").strip()
    filename = f"{encf}.xml"
    recepcion_url = endpoints["rfce_recepcion"] if rfce else endpoints["recepcion"]

    cert_path = DgiiDirectService._prepare_tls_cert(company_profile)
    try:
        max_retries = 3
        for attempt in range(max_retries):
            try:
                response = DgiiDirectService._multipart_post(
                    recepcion_url, signed_xml, token=token,
                    filename=filename, cert_path=cert_path
                )
                break
            except Exception as e:
                err_name = type(e).__name__
                if attempt < max_retries - 1 and ("Connection" in err_name or "ProtocolError" in err_name):
                    wait = (attempt + 1) * 5
                    print(f"    Retry {attempt+1}/{max_retries} en {wait}s: {err_name}")
                    time.sleep(wait)
                else:
                    raise
        response_text = response.text if response is not None else ""
        status_code = response.status_code if response is not None else 0
        response_data = DgiiDirectService._safe_json(response)

        # Mostrar respuesta real para debug
        print(f"    DGII resp: HTTP {status_code} | {response_text[:400]}")

        dgii_status = DgiiDirectService._extract_status(response_data, response_text)
        track_id = DgiiDirectService._extract_track_id(response_data, response_text) or f"dgii_{uuid.uuid4().hex[:12]}"

        result = {
            "success": 200 <= status_code < 300,
            "status_code": status_code,
            "dgii_status": dgii_status,
            "track_id": track_id,
            "response": response_text[:500] if response_text else "",
            "rfce": rfce,
        }
        if isinstance(response_data, dict):
            result["response_data"] = response_data
            if response_data.get("codigo"):
                result["codigo_rfce"] = response_data["codigo"]
            if response_data.get("estado"):
                result["estado_rfce"] = response_data["estado"]
            if response_data.get("secuenciaUtilizada"):
                result["secuencia"] = response_data["secuenciaUtilizada"]

        return result
    finally:
        DgiiDirectService._cleanup_tls_cert(cert_path)


def check_dgii_status(token, company_profile, track_id, endpoints, sandbox):
    """Consulta el estado de un envio por trackId."""
    from app.services.dgii_direct import DgiiDirectService
    try:
        result = DgiiDirectService.check_status(company_profile, track_id, sandbox=sandbox)
        return result
    except Exception as e:
        return {"error": str(e), "track_id": track_id}


def run_paso3_acecf(excel_path, profile, sandbox, signing_mode, dry_run):
    """Procesa el Set de Pruebas de Aprobaciones Comerciales (Paso 3)."""
    from app.services.dgii_direct import DgiiDirectService
    import openpyxl as xl

    os.environ["DGII_SIGNING_MODE"] = signing_mode
    output_dir = "evidencia_fase2"

    print(f"\n{'=' * 60}")
    print("PASO 3 — Aprobaciones Comerciales (ACECF)")
    print(f"{'=' * 60}")

    wb = xl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = {}
    for cell in ws[1]:
        if cell.value is not None:
            def cl(n):
                letters = []
                while n > 0: n -= 1; letters.append(chr(65 + n % 26)); n //= 26
                return ''.join(reversed(letters))
            headers[cl(cell.column)] = str(cell.value).strip()

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        row_dict = {}
        for cell in row:
            if cell.value is not None:
                row_dict[cl(cell.column)] = str(cell.value).strip()
        if row_dict:
            rows.append((row_dict, headers))
    wb.close()

    print(f"\n  Excel: {excel_path}")
    print(f"  Casos de prueba: {len(rows)}")

    print("\n  Autenticando con DGII sandbox...")
    endpoints = build_endpoints(sandbox)
    token = get_token(profile, sandbox)

    results = []
    xml_dir = os.path.join(output_dir, "xml")
    os.makedirs(xml_dir, exist_ok=True)

    for idx, (row_dict, headers) in enumerate(rows):
        encf = row_dict.get("C", f"ACE{idx+1}")
        estado = row_dict.get("G", "1")
        label = "Aceptado" if estado == "1" else "Rechazado"
        print(f"\n  [{idx+1}/{len(rows)}] ACECF eNCF={encf} Estado={label}")

        try:
            raw_xml = DgiiTestDataLoader.build_acecf_xml_from_row(row_dict, headers)
            signed_xml = DgiiSigner.sign_xml(raw_xml, profile)
        except Exception as e:
            print(f"    ERROR: {e}")
            results.append({"encf": encf, "success": False, "error": str(e)})
            continue

        signed_path = os.path.join(xml_dir, f"ACECF_{encf}_signed.xml")
        with open(signed_path, "wb") as f:
            f.write(signed_xml)
        print(f"    XML: {signed_path} ({len(signed_xml)} bytes)")

        if dry_run:
            results.append({"encf": encf, "success": True, "signed_xml": signed_path})
            continue

        cert_path = DgiiDirectService._prepare_tls_cert(profile)
        try:
            acecf_url = endpoints["acecf"]
            response = DgiiDirectService._multipart_post(
                acecf_url, signed_xml, token=token,
                filename=f"aprobacion_comercial_{encf}.xml", cert_path=cert_path
            )
            resp_text = response.text if response is not None else ""
            sc = response.status_code if response is not None else 0
            print(f"    DGII resp: HTTP {sc} | {resp_text[:300]}")

            results.append({
                "encf": encf, "success": 200 <= sc < 300,
                "status_code": sc, "response": resp_text[:500],
                "signed_xml": signed_path,
            })
        except Exception as e:
            print(f"    ERROR envio: {e}")
            results.append({"encf": encf, "success": False, "error": str(e)})
        finally:
            DgiiDirectService._cleanup_tls_cert(cert_path)

        time.sleep(0.5)

    ok = sum(1 for r in results if r.get("success"))
    print(f"\n{'=' * 60}")
    print(f"RESUMEN Paso 3: {ok}/{len(results)} aceptados")
    print(f"{'=' * 60}")

def main():
    parser = argparse.ArgumentParser(description="Fase II - Certificacion DGII")
    parser.add_argument("--excel", required=True, help="Ruta al archivo Excel del Set de Pruebas")
    parser.add_argument("--owner-uid", required=True, help="UID del usuario dueno")
    parser.add_argument("--company-id", required=True, help="ID de la empresa en Firestore")
    parser.add_argument("--dry-run", action="store_true", help="Solo generar XML, no enviar a DGII")
    parser.add_argument("--groups", default="1,2,3,4", help="Grupos a ejecutar (ej: 1,2 o 1,2,3,4)")
    parser.add_argument("--paso-3", dest="paso3_excel", default=None, help="Excel de Aprobaciones Comerciales (Paso 3)")
    parser.add_argument("--wait", action="store_true", help="Esperar y consultar estado hasta que todos sean ACEPTADOS")
    parser.add_argument("--skip-accepted", action="store_true", help="Omitir eNCFs ya aceptados en ejecucion anterior")
    parser.add_argument("--results-json", default="evidencia_fase2/resultados_fase2.json", help="JSON de resultados previos")
    parser.add_argument("--output-dir", default="evidencia_fase2", help="Directorio de salida")
    args = parser.parse_args()

    selected_groups = set(int(g.strip()) for g in args.groups.split(",") if g.strip().isdigit())

    # ── Cargar eNCFs ya aceptados de ejecuciones anteriores ──
    accepted_enfc = set()
    if args.skip_accepted and os.path.exists(args.results_json):
        with open(args.results_json, "r", encoding="utf-8") as f:
            prev = json.load(f)
        for r in prev.get("resultados", []):
            encf = r.get("encf", "")
            success = r.get("success", False)
            tid = str(r.get("track_id", ""))
            dgii_st = str(r.get("dgii_status", "")).upper()
            # Aceptado si: success=True, trackId real, estado ACCEPTED o PENDING (no REJECTED)
            if success and tid and not tid.startswith("dgii_"):
                status_check = r.get("status_check", {})
                check_st = str(status_check.get("dgiiStatus", "")).upper()
                if check_st in ("ACCEPTED", "ACEPTADO", "PENDING") or dgii_st in ("ACCEPTED", "ACEPTADO", "PENDING"):
                    accepted_enfc.add(encf)
        if accepted_enfc:
            print(f"  [INFO] Omitiendo {len(accepted_enfc)} eNCFs ya aceptados: {sorted(accepted_enfc)}")
        else:
            print(f"  [INFO] No se encontraron eNCFs aceptados en {args.results_json}")

    # ================ 1. Inicializar app y cargar perfil ================
    app = create_app()
    sandbox, signing_mode = resolve_env()

    print("=" * 60)
    print("FASE II - Certificacion DGII (Pruebas de Datos e-CF)")
    print("=" * 60)
    print(f"  Entorno: {'SANDBOX' if sandbox else 'PRODUCCION'}")
    print(f"  Modo firma: {signing_mode}")
    print(f"  Excel: {args.excel}")
    print(f"  Dry run: {args.dry_run}")
    print()

    os.environ["DGII_SIGNING_MODE"] = signing_mode

    with app.app_context():
        # Cargar perfil de empresa
        print("[1/6] Cargando perfil de empresa...")
        profile = DatabaseService.get_company_profile(args.owner_uid, company_id=args.company_id)
        if not profile:
            print("ERROR: No se encontro el perfil de empresa.")
            sys.exit(1)

        # ── Paso 3: Aprobaciones Comerciales (ACECF) — flujo independiente ──
        if args.paso3_excel:
            run_paso3_acecf(args.paso3_excel, profile, sandbox, signing_mode, args.dry_run)
            return

        rnc = profile.get("companyRNC", "").replace("-", "")
        company_name = profile.get("companyName", "N/A")
        has_cert = bool(profile.get("certificateContent"))
        print(f"  Empresa: {company_name} (RNC: {rnc})")
        print(f"  Certificado: {'CARGADO' if has_cert else 'NO CARGADO - ERROR'}")
        if not has_cert:
            print("ERROR: No hay certificado digital cargado.")
            sys.exit(1)

        # ================ 2. Cargar Excel ================
        print("\n[2/6] Cargando Set de Pruebas...")
        sheet1_rows, sheet2_rows = DgiiTestDataLoader.load_workbook(args.excel)
        print(f"  Hoja1: {len(sheet1_rows)} casos de prueba")
        print(f"  Hoja2: {len(sheet2_rows)} resumenes RFCE")

        # ================ 3. Autenticar con DGII ================
        print("\n[3/6] Autenticando con DGII sandbox...")
        endpoints = build_endpoints(sandbox)
        print(f"  Auth URL: {endpoints['auth_semilla']}")
        print(f"  Recepcion URL: {endpoints['recepcion']}")
        print(f"  RFCE URL: {endpoints['rfce_recepcion']}")

        token = None
        if not args.dry_run:
            token = get_token(profile, sandbox)
            print("  Token obtenido exitosamente.")
        else:
            print("  [DRY RUN] Sin autenticacion.")

        # ================ 4. Procesar casos en orden DGII ================
        os.makedirs(args.output_dir, exist_ok=True)
        results = []
        xml_dir = os.path.join(args.output_dir, "xml")
        os.makedirs(xml_dir, exist_ok=True)

        # ── Clasificar todos los casos ──
        DGII_ORDER_GROUP1 = ["31", "32", "41", "43", "44", "45", "46", "47"]  # E32 >=250K va aqui
        DGII_ORDER_GROUP2 = ["33", "34"]
        # E32 <250K va en grupo 3 (RFCE) y grupo 4 (factura completa)

        grupos = {1: [], 2: [], 3: [], 4: []}
        casos_map = {}  # encf -> (row_dict, headers, tipo, total)

        for row_dict, headers in sheet1_rows:
            tipo = row_dict.get("C", "?")
            encf = row_dict.get("D", f"E{tipo}??????")
            total_str = row_dict.get("EW", row_dict.get(
                next((c for c, h in headers.items() if h.strip() == "MontoTotal"), ""), "0"))
            total = float(total_str) if total_str else 0.0

            casos_map[encf] = (row_dict, headers, tipo, total)

        for encf, (row_dict, headers, tipo, total) in casos_map.items():
            if args.skip_accepted and encf in accepted_enfc:
                continue
            is_e32 = tipo == "32"
            is_rfce = is_e32 and total < RFCE_THRESHOLD

            if is_rfce:
                grupos[3].append((encf, row_dict, headers, tipo, total, "rfce"))
                grupos[4].append((encf, row_dict, headers, tipo, total, "e32_completa"))
            elif tipo in DGII_ORDER_GROUP1:
                grupos[1].append((encf, row_dict, headers, tipo, total, "e-cf"))
            elif tipo in DGII_ORDER_GROUP2:
                grupos[2].append((encf, row_dict, headers, tipo, total, "e-cf"))

        # Ordenar cada grupo por el orden DGII
        def sort_key(item):
            encf, _rd, _hdrs, tipo, _tot, _tag = item
            order_list = DGII_ORDER_GROUP1 if tipo in DGII_ORDER_GROUP1 else DGII_ORDER_GROUP2
            return (order_list.index(tipo) if tipo in order_list else 99, encf)

        for g in [1, 2, 3, 4]:
            grupos[g].sort(key=sort_key)

        total_casos = sum(len(g) for g in grupos.values())

        def process_case(encf, row_dict, headers, tipo, total, tag, group_label, counter, rfce=False):
            nonlocal token
            print(f"\n  [{counter}/{total_casos}] Grupo {group_label} | tipo={tipo} eNCF={encf} total={total:,.2f}")

            try:
                if rfce and tag == "rfce":
                    raw_xml = DgiiTestDataLoader.build_rfce_xml_from_row(row_dict, headers)
                else:
                    raw_xml = DgiiTestDataLoader.build_xml_from_row(row_dict, headers)
            except Exception as e:
                print(f"    ERROR al construir XML: {e}")
                results.append({"encf": encf, "tipo": tipo, "total": total, "grupo": group_label,
                               "tag": tag, "success": False, "error": f"XML build: {e}"})
                return

            raw_path = os.path.join(xml_dir, f"{encf}_raw.xml")
            with open(raw_path, "wb") as f:
                f.write(raw_xml)

            try:
                signed_xml = DgiiSigner.sign_xml(raw_xml, profile)
            except Exception as e:
                print(f"    ERROR al firmar: {e}")
                results.append({"encf": encf, "tipo": tipo, "total": total, "grupo": group_label,
                               "tag": tag, "success": False, "error": f"Sign: {e}", "raw_xml": raw_path})
                return

            signed_path = os.path.join(xml_dir, f"{encf}_signed.xml")
            with open(signed_path, "wb") as f:
                f.write(signed_xml)
            print(f"    XML: {signed_path} ({len(signed_xml)} bytes)")

            if args.dry_run:
                results.append({"encf": encf, "tipo": tipo, "total": total, "grupo": group_label,
                               "tag": tag, "success": True, "dry_run": True,
                               "raw_xml": raw_path, "signed_xml": signed_path})
                return

            result = send_to_dgii(signed_xml, token, profile, row_dict, endpoints, encf, rfce=rfce)
            result["tipo"] = tipo
            result["encf"] = encf
            result["total"] = total
            result["grupo"] = group_label
            result["tag"] = tag
            result["signed_xml"] = signed_path

            status = "OK" if result.get("success") else "FAIL"
            track = result.get("track_id", "N/A")
            dgii_st = result.get("dgii_status", "?")
            print(f"    Envio: {status} | trackId={track} | DGII={dgii_st}")
            results.append(result)

        counter = 0
        total_casos = sum(len(grupos[g]) for g in selected_groups)

        # Grupo 1
        if 1 in selected_groups:
            print(f"\n[4a/6] Grupo 1 — ECF principales ({len(grupos[1])} casos)")
            for encf, row_dict, headers, tipo, total, tag in grupos[1]:
                counter += 1
                process_case(encf, row_dict, headers, tipo, total, tag, 1, counter)
                time.sleep(0.3)
        else:
            print(f"\n[4a/6] Grupo 1 — Omitido (--groups={args.groups})")

        # Grupo 2: E33, E34 (Notas de Credito/Debito)
        if 2 in selected_groups:
            print(f"\n[4b/6] Grupo 2 — Notas ({len(grupos[2])} casos)")
            for encf, row_dict, headers, tipo, total, tag in grupos[2]:
                counter += 1
                process_case(encf, row_dict, headers, tipo, total, tag, 2, counter)
                time.sleep(0.3)
        else:
            print(f"\n[4b/6] Grupo 2 — Omitido (--groups={args.groups})")

        # Grupo 3: E32 <250K — RFCE summaries
        if 3 in selected_groups:
            print(f"\n[4c/6] Grupo 3 — Resumenes RFCE ({len(grupos[3])} casos)")
            if sheet2_rows:
                rfce_map = {}
                for rfce_row_dict, rfce_headers in sheet2_rows:
                    rfce_encf = DgiiTestDataLoader._v(rfce_row_dict, rfce_headers, "ENCF") or rfce_row_dict.get("D", "")
                    if rfce_encf:
                        rfce_map[rfce_encf] = (rfce_row_dict, rfce_headers)

                for encf, row_dict, headers, tipo, total, tag in grupos[3]:
                    counter += 1
                    # ── 1. Extraer SignatureValue del E32 original firmado ──
                    codigo_seg = None
                    e32_signed_path = os.path.join(xml_dir, f"{encf}_e32_firmado.xml")
                    if encf in casos_map:
                        e32_row, e32_headers, _, _ = casos_map[encf]
                        try:
                            e32_raw = DgiiTestDataLoader.build_xml_from_row(e32_row, e32_headers)
                            e32_signed = DgiiSigner.sign_xml(e32_raw, profile)
                            # Guardar el E32 firmado para reuso en Grupo 4
                            with open(e32_signed_path, "wb") as f:
                                f.write(e32_signed)
                            sig_val = DgiiSigner.extract_signature_value(e32_signed)
                            if sig_val:
                                codigo_seg = sig_val[:6]
                                print(f"  [{counter}/{total_casos}] CodigoSeguridadeCF={codigo_seg} (E32 guardado en {e32_signed_path})")
                        except Exception as e:
                            print(f"  [{counter}/{total_casos}] WARN: no se pudo extraer SignatureValue: {e}")

                    print(f"\n  [{counter}/{total_casos}] Grupo 3 | tipo={tipo} eNCF={encf} total={total:,.2f}")

                    if encf in rfce_map:
                        rfce_row, rfce_hdrs = rfce_map[encf]
                        try:
                            raw_xml = DgiiTestDataLoader.build_rfce_xml_from_row(rfce_row, rfce_hdrs, codigo_seguridad=codigo_seg)
                        except Exception as e:
                            print(f"    ERROR al construir RFCE: {e}")
                            results.append({"encf": encf, "tipo": tipo, "total": total, "grupo": 3,
                                           "tag": tag, "success": False, "error": f"RFCE build: {e}"})
                            continue
                    else:
                        try:
                            raw_xml = DgiiTestDataLoader.build_rfce_xml_from_row(row_dict, headers, codigo_seguridad=codigo_seg)
                        except Exception as e:
                            print(f"    ERROR al construir RFCE: {e}")
                            results.append({"encf": encf, "tipo": tipo, "total": total, "grupo": 3,
                                           "tag": tag, "success": False, "error": f"RFCE build: {e}"})
                            continue

                    raw_path = os.path.join(xml_dir, f"{encf}_rfce_raw.xml")
                    with open(raw_path, "wb") as f:
                        f.write(raw_xml)

                    try:
                        signed_xml = DgiiSigner.sign_xml(raw_xml, profile)
                    except Exception as e:
                        print(f"    ERROR al firmar: {e}")
                        results.append({"encf": encf, "tipo": tipo, "total": total, "grupo": 3,
                                       "tag": tag, "success": False, "error": f"Sign: {e}", "raw_xml": raw_path})
                        continue

                    signed_path = os.path.join(xml_dir, f"{encf}_rfce_signed.xml")
                    with open(signed_path, "wb") as f:
                        f.write(signed_xml)
                    print(f"    XML: {signed_path} ({len(signed_xml)} bytes)")

                    if args.dry_run:
                        results.append({"encf": encf, "tipo": tipo, "total": total, "grupo": 3,
                                       "tag": tag, "success": True, "dry_run": True,
                                       "raw_xml": raw_path, "signed_xml": signed_path,
                                       "codigo_seguridad": codigo_seg})
                        continue

                    result = send_to_dgii(signed_xml, token, profile,
                                         rfce_row if encf in rfce_map else row_dict,
                                         endpoints, encf, rfce=True)
                    result["tipo"] = tipo
                    result["encf"] = encf
                    result["total"] = total
                    result["grupo"] = 3
                    result["tag"] = tag
                    result["signed_xml"] = signed_path
                    result["codigo_seguridad"] = codigo_seg

                    status = "OK" if result.get("success") else "FAIL"
                    track = result.get("track_id", "N/A")
                    dgii_st = result.get("dgii_status", "?")
                    print(f"    Envio: {status} | trackId={track} | DGII={dgii_st}")
                    results.append(result)
                    time.sleep(0.3)
        else:
            print(f"\n[4c/6] Grupo 3 — Omitido (--groups={args.groups})")

        # Grupo 4: E32 <250K — Facturas completas (subida manual al portal)
        if 4 in selected_groups:
            print(f"\n[4d/6] Grupo 4 — Facturas E32 <250K (subida manual — usa el mismo XML firmado del RFCE) ({len(grupos[4])} casos)")
            for encf, row_dict, headers, tipo, total, tag in grupos[4]:
                counter += 1
                print(f"\n  [{counter}/{total_casos}] Grupo 4 | tipo={tipo} eNCF={encf} total={total:,.2f} | SUBIDA MANUAL")
                try:
                    e32_signed_path = os.path.join(xml_dir, f"{encf}_e32_firmado.xml")
                    manual_path = os.path.join(xml_dir, f"{encf}_manual_signed.xml")

                    use_existing = False
                    if os.path.exists(e32_signed_path):
                        with open(e32_signed_path, "rb") as f:
                            content = f.read()
                            if b"<ECF>" in content:
                                use_existing = True
                                import shutil
                                shutil.copy(e32_signed_path, manual_path)
                                sv = DgiiSigner.extract_signature_value(content)
                                print(f"    XML reusado de RFCE: {manual_path} ({len(content)} bytes) | SignatureValue[:6]={sv[:6] if sv else 'N/A'}")

                    if not use_existing:
                        raw_xml = DgiiTestDataLoader.build_xml_from_row(row_dict, headers)
                        signed_xml = DgiiSigner.sign_xml(raw_xml, profile)
                        with open(manual_path, "wb") as f:
                            f.write(signed_xml)
                        sv = DgiiSigner.extract_signature_value(signed_xml)
                        print(f"    XML generado (eCF 32 <250K): {manual_path} ({len(signed_xml)} bytes) | SignatureValue[:6]={sv[:6] if sv else 'N/A'}")

                    results.append({"encf": encf, "tipo": tipo, "total": total, "grupo": 4,
                                   "tag": "manual_upload", "success": True, "signed_xml": manual_path,
                                   "nota": "Subir manualmente en portal DGII > Facturas de consumo < 250Mil"})
                except Exception as e:
                    print(f"    ERROR: {e}")
                    results.append({"encf": encf, "tipo": tipo, "total": total, "grupo": 4,
                                   "tag": "manual_upload", "success": False, "error": str(e)})
        else:
            print(f"\n[4d/6] Grupo 4 — Omitido (--groups={args.groups})")

        # ================ 5. Consultar estados (con --wait polling) ================
        track_results = [r for r in results if not r.get("dry_run") and r.get("success") and r.get("track_id") and not str(r.get("track_id", "")).startswith("dgii_")]

        if track_results:
            if args.wait:
                print(f"\n[5/6] Esperando que DGII procese {len(track_results)} envios (--wait)...")
                max_attempts = 30
                for attempt in range(1, max_attempts + 1):
                    pending = 0
                    accepted = 0
                    rejected = 0
                    for r in track_results:
                        current = r.get("status_check", {}).get("dgiiStatus", "PENDING")
                        # Only check non-final states
                        if current not in ("ACCEPTED", "REJECTED", "ACEPTADO", "RECHAZADO"):
                            track_id = r["track_id"]
                            status_result = check_dgii_status(token, profile, track_id, endpoints, sandbox)
                            if status_result.get("error") and "auth" in str(status_result.get("error", "")).lower():
                                token = get_token(profile, sandbox)
                                status_result = check_dgii_status(token, profile, track_id, endpoints, sandbox)
                            r["status_check"] = status_result
                            current = status_result.get("dgiiStatus", status_result.get("error", "?"))
                        if current in ("ACCEPTED", "ACEPTADO"):
                            accepted += 1
                        elif current in ("REJECTED", "RECHAZADO"):
                            rejected += 1
                        else:
                            pending += 1
                    print(f"  Intento {attempt}/{max_attempts}: {accepted} aceptados, {rejected} rechazados, {pending} pendientes")
                    if pending == 0:
                        break
                    if attempt < max_attempts:
                        wait_sec = min(attempt * 10, 60)
                        print(f"  Esperando {wait_sec}s...")
                        time.sleep(wait_sec)
            else:
                print(f"\n[5/6] Consultando estados de {len(track_results)} envios...")
                for r in track_results:
                    track_id = r["track_id"]
                    print(f"  Consultando trackId={track_id}...")
                    status_result = check_dgii_status(token, profile, track_id, endpoints, sandbox)
                    r["status_check"] = status_result
                    st = status_result.get("dgiiStatus", status_result.get("error", "?"))
                    print(f"    Estado: {st}")
                    time.sleep(0.5)
        else:
            print(f"\n[5/6] Sin trackIds reales para consultar.")

        # ================ 6. Guardar resultados ================
        print(f"\n[6/6] Guardando resultados...")
        summary = {
            "fecha": datetime.now(timezone.utc).isoformat(),
            "empresa": company_name,
            "rnc": rnc,
            "total_casos": len(results),
            "exitosos": sum(1 for r in results if r.get("success")),
            "fallidos": sum(1 for r in results if not r.get("success") and not r.get("dry_run")),
            "dry_run": sum(1 for r in results if r.get("dry_run")),
            "resultados": results,
        }

        resumen_path = os.path.join(args.output_dir, "resultados_fase2.json")
        with open(resumen_path, "w", encoding="utf-8") as f:
            json.dump(summary, f, indent=2, ensure_ascii=False, default=str)

        print(f"\n{'=' * 60}")
        print(f"RESUMEN")
        print(f"{'=' * 60}")
        print(f"  Total casos:      {summary['total_casos']}")
        print(f"  Exitosos:         {summary['exitosos']}")
        print(f"  Fallidos:         {summary['fallidos']}")
        print(f"  Dry run:          {summary['dry_run']}")
        print(f"  Resultados:       {resumen_path}")
        print(f"  XML generados en: {xml_dir}")
        print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
